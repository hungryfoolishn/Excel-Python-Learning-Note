# coding: utf-8

import requests
import base64
import pandas as pd
import matplotlib.pyplot as plt
import calendar
import json
import smtplib
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

charts_path = r'F:\Learn\工作\数据明细\charts'
filename_total_80 = r"F:\Learn\工作\数据明细\通报"
city = '杭州'
start_time = '2023-11-01'
end_time = '2023-11-30'
def get_df_from_db(sql):
    data = {
        "type": "query",
        "tableName": sql['tableName'],
        "where": (base64.b64encode(sql['where'].encode())).decode(),
        "columns": sql['columns'],
        "isEncry": "1"
    }
    url = 'https://yhxc.jtyst.zj.gov.cn:7443/zc-interface/db/excuteSql'
    headers = {'content-type': 'application/json'}
    res = requests.post(url, json=data, headers=headers)
    res = json.loads(res.text)
    data = res['data']
    data = pd.DataFrame(data)
    return data


def 总重80吨以上数据分析(start_time, end_time):
    ##合规率
    t_sys_station = {
        "tableName": "t_sys_station ",
        "where": " station_status=0 and station_type =31 and is_deleted= 0 ",
        "columns": "station_name,station_code,station_status,station_type,area_county  "
    }
    t_sys_station = get_df_from_db(t_sys_station)
    station_code = t_sys_station['station_code']

    sql = {
        "tableName": "t_bas_over_data_31 a left join t_bas_over_data_collection_31 b on a.record_code=b.record_code  left join t_bas_over_data_opinion_31 d on a.record_code = d.data_id",
        "where": "a.out_station_time between '{} 00:00:00' and '{} 23:59:59' and a.total_weight>80    and a.allow is null ".format(start_time, end_time),
        "columns": "a.area_county,a.out_station,a.out_station_time  检测时间,a.valid_time 入库时间,a.car_no  车牌号码,a.total_weight  总重,a.limit_weight 限重,a.overrun  超重,a.overrun_rate  超限率,a.axis  轴数,b.status 案件状态,d.content 审核意见,a.site_name 站点名称,a.is_collect 证据满足,b.law_judgment 判定需处罚,b.make_copy  外省抄告,a.link_man 所属运输企业名称,a.phone 联系电话, b.vehicle_county 车籍地,a.record_code 流水号"
    }
    t_bas_over_data_31_80 = get_df_from_db(sql)
    t_bas_over_data_31_80 = pd.DataFrame(t_bas_over_data_31_80,
                          columns=["out_station", "站点名称", "检测时间", "车牌号码", "总重", "限重",
                                   "超重", "轴数", "超限率", "证据满足", "所属运输企业名称", "联系电话", "车籍地", "流水号", "入库时间", "案件状态", "审核意见","判定需处罚",
                                   "外省抄告", "area_county"])

    sql = {
        "tableName": "t_code_area",
        "where": "city_code='330100' ",
        "columns": " city 地市,county 区县,county_code"
    }
    t_code_area = get_df_from_db(sql)
    t_bas_over_data_31_80 = pd.merge(t_bas_over_data_31_80, t_code_area, left_on='area_county', right_on='county_code', how='left')

    df_big = t_bas_over_data_31_80[t_bas_over_data_31_80.loc[:, 'out_station'].isin(station_code)]
    df_big = pd.DataFrame(df_big,
                          columns=["地市", "区县", "站点名称", "检测时间", "车牌号码", "总重", "限重",
                                   "超重", "轴数", "超限率", "证据满足", "所属运输企业名称", "联系电话", "车籍地", "流水号", "入库时间", "案件状态", "审核意见","判定需处罚",
                                   "外省抄告", "area_county"])
    df_big.drop_duplicates(subset=['流水号'], inplace=True)
    总重80吨以上明细 = df_big.sort_values(by=['area_county'],
                                   ascending=True).reset_index(drop=True)
    return 总重80吨以上明细

def creat_charts():

    df = 总重80吨以上数据分析(start_time, end_time)
    df = df[df['地市'] == '{}'.format(city)]
    df['检测时间'] = pd.to_datetime(df['检测时间'])
    df['检测时间日'] = df['检测时间'].map(lambda x: x.strftime('%d'))
    df['检测时间时'] = df['检测时间'].map(lambda x: x.strftime('%H'))
    df['案件状态'] = df['案件状态'].fillna('99')
    df['案件状态'] = df['案件状态'].astype('int')
    df.loc[df['案件状态'] == 99, '案件状态'] = '未采集'
    df.loc[df['案件状态'] == 0, '案件状态'] = '审核不通过'
    df.loc[df['案件状态'] == 1, '案件状态'] = '待初审'
    df.loc[df['案件状态'] == 2, '案件状态'] = '待审核'
    df.loc[df['案件状态'] == 3, '案件状态'] = '待判定'
    df.loc[df['案件状态'] == 4, '案件状态'] = '已告知'
    df.loc[df['案件状态'] == 5, '案件状态'] = '免处理'
    df.loc[df['案件状态'] == 6, '案件状态'] = '已立案'
    df.loc[df['案件状态'] == 12, '案件状态'] = '待告知'
    df.loc[df['案件状态'] == 13, '案件状态'] = '已结案'
    df.loc[df['案件状态'] == 9, '案件状态'] = '判定不处理'
    df.loc[df['案件状态'] == 15, '案件状态'] = '初审不通过'
    df['判定需处罚'] = df['判定需处罚'].fillna('2')
    df['判定需处罚'] = df['判定需处罚'].astype('int')
    df.loc[df['判定需处罚'] == 2, '判定需处罚'] = '未采集'
    df.loc[df['判定需处罚'] == 99, '判定需处罚'] = '初审'
    df.loc[df['判定需处罚'] == 1, '判定需处罚'] = '复审'
    cor = df['案件状态'].value_counts()
    # cor = cor[0:5]
    x = range(len(cor.index))
    name = cor.index
    y = cor.values
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    plt.figure(figsize=(12, 7))
    # 计算各店铺的商品数量
    cor.plot.bar(width=0.8, alpha=0.6, color='r')
    plt.title('80吨以上(案件状态)', fontsize=18)
    plt.ylabel('案件数', fontsize=14)
    for x, y in zip(x, y):
        plt.text(x, y + 0.1, '%d' % y, ha='center', va='bottom')
    plt.savefig(r'{}\{}80吨以上(案件状态).png'.format(charts_path, city), bbox_inches='tight')
    city_month = df['判定需处罚'].value_counts()
    plt.figure(figsize=(21, 12), dpi=80)
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    plt.subplot(1, 1, 1)
    plt.pie(city_month, labels=city_month.index, autopct='%.2f%%')
    # plt.legend()
    plt.title('审核状态', fontsize=18)
    plt.savefig(r'{}\{}审核状态.png'.format(charts_path, city), bbox_inches='tight')
    city_month = df[df['判定需处罚'] == '初审']['案件状态'].value_counts()
    plt.figure(figsize=(21, 12), dpi=80)
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    plt.subplot(1, 1, 1)
    plt.pie(city_month, labels=city_month.index, autopct='%.2f%%')
    # plt.legend()
    plt.title('初审占比', fontsize=18)
    plt.savefig(r'{}\{}初审占比.png'.format(charts_path, city), bbox_inches='tight')
    city_month = df[df['判定需处罚'] == '复审']['案件状态'].value_counts()
    plt.figure(figsize=(21, 12), dpi=80)
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    plt.subplot(1, 1, 1)
    plt.pie(city_month, labels=city_month.index, autopct='%.2f%%')
    # plt.legend()
    plt.title('复审占比', fontsize=18)
    plt.savefig(r'{}\{}复审占比.png'.format(charts_path, city), bbox_inches='tight')

    city_month = df[df['案件状态'] == '初审不通过']['审核意见'].value_counts().sort_values(ascending=False)
    plt.figure(figsize=(21, 12), dpi=80)
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    plt.subplot(1, 1, 1)
    plt.pie(city_month, labels=city_month.index, autopct='%.2f%%')
    # plt.legend()
    plt.title('审核意见(初始不通过)', fontsize=18)
    plt.savefig(r'{}\{}审核意见(初始不通过).png'.format(charts_path, city), bbox_inches='tight')

    # city_month=df['检测时间日'].value_counts().sort_index()
    city_month = df['检测时间日'].value_counts().sort_index()
    name = city_month.index
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    x = list(range(len(city_month.index)))
    y = city_month.values
    width = 0.3
    plt.figure(figsize=(15, 8), dpi=80)
    plt.plot(x, y, color='lightsalmon')
    plt.xticks(x, name)
    # plt.legend()
    plt.xlabel('天', fontsize=14)
    plt.ylabel('辆次', fontsize=14)
    plt.title('80吨以上(天）', fontsize=18)
    for x, y in zip(x, y):
        plt.text(x, y + 0.1, '%d' % y, ha='center', va='bottom')
    plt.savefig(r'{}\{}80吨以上(天）.png'.format(charts_path, city), bbox_inches='tight')
    # plt.xlabel(u'日', fontproperties=font) # 添加中文字符串
    city_month = df['检测时间时'].value_counts()
    city_month = city_month[0:12]
    name = city_month.index
    x = list(range(len(city_month.index)))
    y = city_month.values
    width = 0.3
    plt.figure(figsize=(15, 8), dpi=80)
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    plt.plot(x, y, color='lightsalmon')
    plt.xticks(x, name)
    # plt.legend()
    plt.xlabel('小时', fontsize=14)
    plt.ylabel('辆次', fontsize=14)
    plt.title('80吨以上(小时)', fontsize=18)
    for x, y in zip(x, y):
        plt.text(x, y + 0.1, '%d' % y, ha='center', va='bottom')
    plt.savefig(r'{}\{}80吨以上(小时).png'.format(charts_path, city), bbox_inches='tight')
    cor = df['站点名称'].value_counts().sort_values(ascending=False)
    cor = cor[0:10]
    cor = cor.sort_values(ascending=True)
    plt.figure(figsize=(8, 15))
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    x = range(len(cor.index))
    name = cor.index
    y = cor.values
    plt.barh(x, y, color='salmon')
    plt.yticks(x, name, fontsize=14)
    for x, y in zip(x, y):
        plt.text(y, x - 0.1, '%.2f' % y, ha='center', va='bottom')
    # plt.ylabel('站点名称', fontproperties=font)
    plt.title('80吨以上(前十站点)', fontsize=18)
    plt.savefig(r'{}\{}80吨以上(前十站点).png'.format(charts_path, city), bbox_inches='tight')
    cor = df['区县'].value_counts().sort_values(ascending=False)
    # cor = cor[0:5]
    x = range(len(cor.index))
    name = cor.index
    y = cor.values
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    plt.figure(figsize=(12, 7))
    # 计算各店铺的商品数量
    cor.plot.bar(width=0.8, alpha=0.6, color='r')
    plt.title('80吨以上(区县排名)', fontsize=18)
    plt.ylabel('辆次', fontsize=14)
    for x, y in zip(x, y):
        plt.text(x, y + 0.1, '%d' % y, ha='center', va='bottom')
    plt.savefig(r'{}\{}80吨以上(区县排名).png'.format(charts_path, city), bbox_inches='tight')
    cor = df[(~((df['车牌号码'].str.contains("牌") == True))) & (~((df['车牌号码'].str.contains("-") == True))) & (
        ~((df['车牌号码'].str.contains("未") == True)))]['车牌号码'].value_counts().sort_values(ascending=False)
    cor = cor[0:10]
    cor = cor.sort_values(ascending=True)
    plt.figure(figsize=(8, 15))
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
    x = range(len(cor.index))
    name = cor.index
    y = cor.values
    plt.barh(x, y, color='salmon')
    plt.yticks(x, name, fontsize=14)
    for x, y in zip(x, y):
        plt.text(y, x - 0.1, '%.2f' % y, ha='center', va='bottom')
    # plt.ylabel('站点名称', fontproperties=font)
    plt.title('80吨以上(前十车牌)', fontsize=18)
    plt.savefig(r'{}\{}80吨以上(前十车牌).png'.format(charts_path, city), bbox_inches='tight')
def add_charts():
    creat_charts()
    workbook = load_workbook(filename=r"{}\{}2023-11月初始数据(2023-11-21).xlsx".format(filename_total_80, city))
    sheet = workbook.create_sheet('80吨以上数据分析')
    # sheet = workbook.active
    image_indices = list(range(len(sheet._images)))
    # 倒序遍历索引列表并删除图片对象
    for index in reversed(image_indices):
        del sheet._images[index]
    logo = Image(r'{}\{}80吨以上(天）.png'.format(charts_path, city))
    logo.height = 600
    logo.width = 1050
    sheet.add_image(logo, "A1")
    logo = Image(r'{}\{}80吨以上(小时).png'.format(charts_path, city))
    logo.height = 600
    logo.width = 1050
    sheet.add_image(logo, "Q1")

    logo = Image(r'{}\{}80吨以上(区县排名).png'.format(charts_path, city))
    logo.height = 600
    logo.width = 1050
    sheet.add_image(logo, "A33")
    logo = Image(r'{}\{}80吨以上(案件状态).png'.format(charts_path, city))
    logo.height = 600
    logo.width = 1050
    sheet.add_image(logo, "Q33")
    logo = Image(r'{}\{}审核状态.png'.format(charts_path, city))
    logo.height = 900
    logo.width = 900
    sheet.add_image(logo, "B66")
    logo = Image(r'{}\{}初审占比.png'.format(charts_path, city))
    logo.height = 900
    logo.width = 900
    sheet.add_image(logo, "Q66")
    logo = Image(r'{}\{}复审占比.png'.format(charts_path, city))
    logo.height = 900
    logo.width = 900
    sheet.add_image(logo, "B115")
    logo = Image(r'{}\{}审核意见(初始不通过).png'.format(charts_path, city))
    logo.height = 1000
    logo.width = 1100
    sheet.add_image(logo, "Q110")
    logo = Image(r'{}\{}80吨以上(前十站点).png'.format(charts_path, city))
    logo.height = 1000
    logo.width = 900
    sheet.add_image(logo, "A163")
    logo = Image(r'{}\{}80吨以上(前十车牌).png'.format(charts_path, city))
    logo.height = 1000
    logo.width = 700
    sheet.add_image(logo, "Q163")
    workbook.save(filename=r"{}\{}2023-11月初始数据(2023-11-21).xlsx.xlsx".format(filename_total_80, city))

if __name__ == "__main__":
    # creat_charts()

    add_charts()