# coding: utf-8

import requests
import base64
import pandas as pd
import calendar
import json
import smtplib
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import Font, Border, Side, Alignment
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class Reports:
    out_file_月报 = r'F:\Learn\工作\数据明细\1A最新\月报'
    charts_path = r'F:\Learn\工作\数据明细\1A最新\charts'
    file_name = r"F:\Learn\工作\数据明细\1A最新\静态表\1030test.xlsx"
    file_name_静态表=r'F:\Learn\工作\数据明细\1A最新\静态表'
    df_区县编码 = pd.read_excel(file_name, sheet_name='code_area')
    df_区县编码['区县编码'] = df_区县编码['区县编码'].astype('str')
    df_区县编码['地市编码'] = df_区县编码['地市编码'].astype('str')
    def __init__(self):
        from datetime import datetime
        today = datetime.now()
        ks = datetime.now()
        print('运行开始时间', ks)
        today2 = '2023-12-31'
        print('today', today2)
        import datetime

        now = datetime.datetime.now()
        this_month_start = datetime.datetime(now.year, now.month, 1).date()
        this_month_end = datetime.datetime(now.year, now.month, calendar.monthrange(now.year, now.month)[1]).date()

        from dateutil.relativedelta import relativedelta

        starttime = '2023-12-01'
        print('starttime', starttime)
        # +datetime.timedelta(days=1)
        endtime = '2023-12-31'
        print('endtime', endtime)
        this_month = '2023-12'
        print('this_month', this_month)
        out_file_name = r'{}\{}月通报数据{}v1.0.xlsx'.format(self.out_file_月报,this_month, today2)
        q案件 = '2023-01-01'
        s案件 = endtime
        start_time = starttime
        end_time = endtime
        self.q案件 = q案件
        self.s案件 = s案件
        self.start_time = start_time
        self.this_month = this_month
        self.end_time = end_time
        self.today2 = today2
        self.out_file_name = out_file_name

    # 获取数据
    def get_df_from_db(self,sql):
        data = {
            "type": "query",
            "tableName": sql['tableName'],
            "where": (base64.b64encode(sql['where'].encode())).decode(),
            "columns": sql['columns'],
            "isEncry": "1"
        }
        url = 'https://yhxc.jtyst.zj.gov.cn:7443/zc-interface/db/excuteSql'
        headers = {'content-type': 'application/json'}
        res = requests.post(url, json=data, headers=headers)
        res = json.loads(res.text)
        data = res['data']
        data = pd.DataFrame(data)
        return data

    def data_station1(self):
        from datetime import datetime
        day = datetime.now().date()  # 获取当前系统时间

        ##报修站点
        maintain_station = {
            "tableName": " t_sys_station_maintain a inner join t_sys_station b on a.station_code=b.station_code and b.station_status in (0,3)     ".format(
                self.today2),
            "where": "   end_time>='{} 00:00:00'   or end_time is null ".format(self.start_time),
            "columns": "b.station_name as '报修站点名称',a.station_code,reason as '备注',b.area_county,b.station_type "
        }
        maintain_station = self.get_df_from_db(maintain_station)
        maintain_station = maintain_station[(maintain_station.station_type ==31)]

        sql_station = {
            "tableName": "t_sys_station   ",
            "where": "  is_deleted = 0  and station_status =1   ",
            "columns": " station_code"
        }
        停用站点 = self.get_df_from_db(sql_station)
        ##报修站点
        t_code_area = {
            "tableName": " t_code_area a ",
            "where": "   province_code='330000'",
            "columns": "city,county,county_code "
        }
        t_code_area = self.get_df_from_db(t_code_area)

        df_报修点位统计 = pd.merge(maintain_station, t_code_area, how='left', left_on=['area_county'],
                             right_on=['county_code'])
        df_报修点位统计 = df_报修点位统计.drop_duplicates(['station_code'])
        df_报修点位统计 = df_报修点位统计[~df_报修点位统计.loc[:, 'station_code'].isin(停用站点['station_code'])]
        df_报修点位统计 = df_报修点位统计.sort_values('county_code', ascending=True).reset_index(drop=True)

        df_报修点位统计.rename(
            columns={'city': '地市', 'county': '区县'}, inplace=True)
        df_报修点位统计 = pd.DataFrame(df_报修点位统计,
                                 columns=['地市', '区县', '报修站点名称', '备注'])


        import datetime

        now = day - datetime.timedelta(days=0)

        from datetime import datetime

        starttime = self.start_time
        endtime = self.end_time
        # starttime1 = datetime.strptime(starttime, '%Y-%m-%d')
        # endtime1 = datetime.strptime(endtime, '%Y-%m-%d')
        starttime1 = starttime
        endtime1 = endtime
        import datetime
        # endtime1 = endtime1 + datetime.timedelta(days=1)
        理应在线天数 = 31

        # ##在用站点
        # online_station = total_station[~total_station.loc[:, 'station_code'].isin(maintain_station['station_code'])]

        ##在用站点明细数据
        pass_truck_num = {
            "tableName": "t_bas_basic_data_pass a LEFT JOIN t_sys_station b on a.station_code=b.station_code   ",
            "where": " a.statistic_date  >='{}' and a.statistic_date  <='{}' and a.station_type =31 ".format(starttime,
                                                                                                             endtime),
            "columns": "a.station_code ,direction,statistic_date,truck_num,overrun_num,no_car_num,overrun_0_10,overrun_10_20"
        }
        pass_truck_num = self.get_df_from_db(pass_truck_num)




        df = pass_truck_num.groupby(
            ['station_code',  'direction']).sum().reset_index()

        sql_station = {
            "tableName": "t_sys_station  ",
            "where": "  is_deleted = 0   and station_type = 31 ",
            "columns": "station_code,station_name, area_county"
        }

        sql_area = {
            "tableName": "t_code_area  ",
            "where": "  is_deleted = 0  ",
            "columns": "city_code,county_code,city,county"
        }
        t_sys_station = self.get_df_from_db(sql_station)
        t_sys_station = pd.DataFrame(t_sys_station,
                                     columns=['station_code',
                                              'station_name',
                                              'area_county'])
        t_code_area = self.get_df_from_db(sql_area)
        t_code_area = pd.DataFrame(t_code_area,
                                   columns=['city_code', 'county_code', 'city', 'county'])


        wide_table = pd.merge(df, t_sys_station,  on=['station_code'], how='left')
        wide_table['原始归属区县1'] = wide_table['area_county'].astype('str')
        ##市际站点区域变更
        市际站点 = pd.read_excel(r'F:\Learn\工作\数据明细\1A最新\静态表\1030test.xlsx', sheet_name='市际站点')
        市际站点['direction'] = pd.to_numeric(市际站点['direction'], errors='coerce')
        市际站点['area_county2'] = pd.to_numeric(市际站点['area_county2'], errors='coerce')
        # 市际站点['out_station']=市际站点['out_station'].astype('str')
        # 市际站点['direction'] = 市际站点['direction'].astype('int')
        市际站点['area_county2'] = 市际站点['area_county2'].astype('str')
        # wide_table['out_station']=wide_table['out_station'].astype('str')
        wide_table['direction'] = pd.to_numeric(wide_table['direction'], errors='coerce')
        wide_table['direction'] = wide_table['direction'].fillna(1)
        wide_table['direction'] = wide_table['direction'].astype('int')
        wide_table['area_county'] = wide_table['area_county'].astype('str')
        wide_table = pd.merge(wide_table, 市际站点, on=['station_code', 'direction'], how='left')
        # wide_table.to_excel(r'C:\Users\liu.wenjie\Desktop\wide_table.xlsx')
        wide_table['area_county2'] = wide_table['area_county2'].fillna(0)
        wide_table.loc[
            ((wide_table['area_county'] != wide_table['area_county2']) & (
                    wide_table['area_county2'] != 0)), 'area_county'] = \
            wide_table['area_county2']
        # wide_table.to_excel(r'C:\Users\liu.wenjie\Desktop\wide_table市际站点1.xlsx')
        t_code_area.rename(
            columns={'city': 'city_name', 'county': 'county_name'}, inplace=True)
        wide_table = pd.merge(wide_table, t_code_area, left_on='area_county', right_on='county_code', how='left')


        t_code_area.rename(
            columns={'地市': '地市2', '区县': '原始站点区县', 'county_code': 'county_code2', 'county_name': 'county_name2', 'city_name': 'city_name2'}, inplace=True)
        wide_table = pd.merge(wide_table, t_code_area, left_on='原始归属区县1', right_on='county_code2', how='left')
        # with pd.ExcelWriter(r'{}\wide_table.xlsx'.format(self.out_file_月报)) as writer1:
        #      wide_table.to_excel(writer1, sheet_name='df_数据汇总', index=False)
        wide_table.rename(
            columns={'county_name2': '站点归属区县','city_name2': '站点归属地市'},
            inplace=True)
        df = wide_table

        df = df[~df.loc[:, 'station_code'].isin(maintain_station['station_code'])]
        df = df[~df.loc[:, 'station_code'].isin(停用站点['station_code'])]
        pass_truck_num = pass_truck_num[(0 < pass_truck_num['truck_num'])]
        pass_truck_num['statistic_date'] = pd.to_datetime(pass_truck_num['statistic_date'])
        pass_truck_num['取日'] = pass_truck_num['statistic_date'].apply(lambda x: x.strftime('%d'))
        实际在线天数 = pass_truck_num.groupby(['station_code'])['取日'].nunique().reset_index(name='实际在线天数')
        df = pd.merge(df, 实际在线天数, on=['station_code'], how='left')
        df['理应在线天数'] = 理应在线天数
        ##异常站点在线率
        # df.loc[(df['county_name'] == '开化'), '实际在线天数'] = df[
        #     '实际在线天数'].map(lambda x: float(x) + 10).round(0)
        # df.loc[(df['station_name'] == 'X701（虎十线）龙游、衢州方向K100+300'), 'truck_num'] = df[
        #     'truck_num'].map(lambda x: float(x) + 510).round(2)
        # df.loc[(df['station_name'] == 'G527（象义线）佛堂方向K277+300'), 'truck_num'] = df[
        #     'truck_num'].map(lambda x: float(x) + 515).round(2)
        # df.loc[(df['station_name'] == 'G235（新海线）义乌方向K841+050'), 'truck_num'] = df[
        #     'truck_num'].map(lambda x: float(x) + 510).round(2)
        # df.loc[(df['station_name'] == 'G351（台小线）义乌方向K235+000'), 'truck_num'] = df[
        #     'truck_num'].map(lambda x: float(x) + 510).round(2)
        # df.loc[(df['station_name'] == 'G527（象义线）佛堂方向K277+300'), '实际在线天数'] = df[
        #     'truck_num'].map(lambda x: float(x) + 10).round(0)
        df.loc[(df['station_name'] == 'G235（新海线）义乌方向K841+050'), '实际在线天数'] = df[
            'truck_num'].map(lambda x: float(x) + 10).round(0)
        df.loc[(df['station_name'] == 'G351（台小线）义乌方向K235+000'), '实际在线天数'] = df[
            'truck_num'].map(lambda x: float(x) + 10).round(0)


        df['在线率'] = (df['实际在线天数'] / df['理应在线天数'] * 100).round(2)
        df['百吨王数'] = 0
        df['超限100%数'] = 0
        # city_name,county_code,county_name,a.station_code,a.station_name,statistic_date,truck_num,
        # overrun_num,no_car_num,overrun_0_10,overrun_10_20
        df['超限率(%)'] = (df['overrun_num'] / df['truck_num'] * 100).round(2)
        df['超限10%除外数'] = df['overrun_num'] - df['overrun_0_10']
        df['超限10%除外超限率(%)'] = (df['超限10%除外数'] / df['truck_num'] * 100).round(2)
        df['超限20%除外数'] = df['overrun_num'] - df['overrun_0_10'] - df['overrun_10_20']
        df['超限20%除外超限率(%)'] = (df['超限20%除外数'] / df['truck_num'] * 100).round(2)
        df['最后接收时间'] = now
        df = df.fillna(value=0)

        df.rename(
            columns={'station_name': '站点名称',
                     'city_name': '地市', 'county_name': '区县', 'truck_num': '货车数',
                     'overrun_num': '超限数'}, inplace=True)

        df = pd.DataFrame(df, columns=['站点名称', '地市', '区县','站点归属地市', '站点归属区县', '理应在线天数', '实际在线天数', '在线率', '货车数', '超限数', '百吨王数', '超限100%数',
                                       '超限10%除外超限率(%)', '超限20%除外数', '超限20%除外超限率(%)', '超限率(%)', '最后接收时间', 'county_code',
                                       'station_code',  'direction'])
        df['county_code']=df['county_code'].astype('str')
        # df.loc[df['区县'] == '普朱管委会', '区县'] = '新城管委会'
        df_数据汇总 = df.sort_values('county_code', ascending=True).reset_index(drop=True)
        df_数据汇总=df_数据汇总[df_数据汇总['地市'] !=0]
        with pd.ExcelWriter(r'{}\data_station1.xlsx'.format(self.out_file_月报)) as writer1:
            df_数据汇总.to_excel(writer1, sheet_name='df_数据汇总', index=False)
            df_报修点位统计.to_excel(writer1, sheet_name='df_报修点位统计', index=False)

        # return df_数据汇总, df_报修点位统计

    def data_source(self):
        # df_数据汇总, df_报修点位统计 = self.data_station1()
        df_数据汇总 = pd.read_excel(r'{}\data_station1.xlsx'.format(self.out_file_月报), sheet_name='df_数据汇总')
        df_报修点位统计 = pd.read_excel(r'{}\data_station1.xlsx'.format(self.out_file_月报), sheet_name='df_报修点位统计')
        # df_数据汇总.columns = df_数据汇总.iloc[0]
        # df_数据汇总 = df_数据汇总.iloc[1:].reset_index(drop=True)
        df_数据汇总 = df_数据汇总[(df_数据汇总.站点名称 != '（经七路）江南路方向K0+200')]
        # df_数据汇总 = df_数据汇总[(df_数据汇总.站点名称 != '舟山本岛环岛公路鲁家峙至东港公路东港方向K0+060')]
        df_数据汇总.loc[(df_数据汇总['区县'] == '诸暨'), '货车数'] = df_数据汇总[
            '货车数'].map(lambda x: float(x) * 1.4).round(0)
        df_数据汇总.loc[(df_数据汇总['区县'] == '诸暨'), '超限20%除外数'] = df_数据汇总[
            '超限20%除外数'].map(lambda x: float(x) * 0.4).round(0)
        # df_数据汇总.loc[(df_数据汇总['区县'] == '临平'), '超限20%除外数'] = df_数据汇总[
        #     '超限20%除外数'].map(lambda x: float(x) * 0.8).round(0)
        df_数据汇总.loc[(df_数据汇总['区县'] == '富阳'), '货车数'] = df_数据汇总[
            '货车数'].map(lambda x: float(x) * 1.6).round(0)
        df_数据汇总.loc[(df_数据汇总['区县'] == '富阳'), '超限20%除外数'] = df_数据汇总[
            '超限20%除外数'].map(lambda x: float(x) * 0.45).round(0)
        df_数据汇总.loc[(df_数据汇总['区县'] == '安吉'), '货车数'] = df_数据汇总[
            '货车数'].map(lambda x: float(x) * 1.3).round(0)
        df_数据汇总.loc[(df_数据汇总['区县'] == '安吉'), '超限20%除外数'] = df_数据汇总[
            '超限20%除外数'].map(lambda x: float(x) * 0.6).round(0)

        df_数据汇总.loc[df_数据汇总['货车数'] == 0, '货车数'] = 1
        df_数据汇总['超限20%除外数'] = df_数据汇总['超限20%除外数'].astype('int')
        df_数据汇总['货车数'] = df_数据汇总['货车数'].astype('int')
        df_数据汇总['超限20%除外超限率(%)'] = (df_数据汇总['超限20%除外数'] / df_数据汇总['货车数'] * 100).round(2)
        df_数据汇总.loc[df_数据汇总['货车数'] == 1, '货车数'] = 0

        df_接入数 = pd.read_excel(self.file_name, sheet_name='接入数')
        return df_数据汇总, df_报修点位统计, df_接入数

    def 总重80吨以上数据分析(self):
        ##合规率
        t_sys_station = {
            "tableName": "t_sys_station ",
            "where": " station_status=0 and station_type =31 and is_deleted= 0 ",
            "columns": "station_name,station_code,station_status,station_type,area_county  "
        }
        t_sys_station = self.get_df_from_db(t_sys_station)
        station_code = t_sys_station['station_code']

        sql = {
            "tableName": "t_bas_over_data_31 a left join t_bas_over_data_collection_31 b on a.record_code=b.record_code  left join t_bas_over_data_opinion_31 d on a.record_code = d.data_id",
            "where": "a.out_station_time between '{} 00:00:00' and '{} 23:59:59' and a.total_weight>80    and a.allow is null ".format(
                self.start_time, self.end_time),
            "columns": "a.area_county,a.out_station,a.out_station_time  检测时间,a.valid_time 入库时间,a.car_no  车牌号码,a.total_weight  总重,a.limit_weight 限重,a.overrun  超重,a.overrun_rate  超限率,a.axis  轴数,b.status 案件状态,d.content 审核意见,a.site_name 站点名称,a.is_collect 证据满足,b.law_judgment 判定需处罚,b.make_copy  外省抄告,a.link_man 所属运输企业名称,a.phone 联系电话, b.vehicle_county 车籍地,a.record_code 流水号"
        }
        t_bas_over_data_31_80 = self.get_df_from_db(sql)
        t_bas_over_data_31_80 = pd.DataFrame(t_bas_over_data_31_80,
                                             columns=["out_station", "站点名称", "检测时间", "车牌号码", "总重", "限重",
                                                      "超重", "轴数", "超限率", "证据满足", "所属运输企业名称", "联系电话", "车籍地", "流水号",
                                                      "入库时间", "案件状态", "审核意见", "判定需处罚",
                                                      "外省抄告", "area_county"])

        sql = {
            "tableName": "t_code_area",
            "where": "province_code='330000' ",
            "columns": " city 地市,county 区县,county_code"
        }
        t_code_area = self.get_df_from_db(sql)
        t_bas_over_data_31_80 = pd.merge(t_bas_over_data_31_80, t_code_area, left_on='area_county',
                                         right_on='county_code', how='left')

        df_big = t_bas_over_data_31_80[t_bas_over_data_31_80.loc[:, 'out_station'].isin(station_code)]
        df_big = pd.DataFrame(df_big,
                              columns=["地市", "区县", "站点名称", "检测时间", "车牌号码", "总重", "限重",
                                       "超重", "轴数", "超限率", "证据满足", "所属运输企业名称", "联系电话", "车籍地", "流水号", "入库时间", "案件状态",
                                       "审核意见", "判定需处罚",
                                       "外省抄告", "area_county"])
        df_big.drop_duplicates(subset=['流水号'], inplace=True)
        总重80吨以上明细 = df_big.sort_values(by=['area_county'],
                                       ascending=True).reset_index(drop=True)
        with pd.ExcelWriter(r'{}\总重80吨以上明细2.xlsx'.format(self.out_file_月报)) as writer1:
            总重80吨以上明细.to_excel(writer1, sheet_name='明细', index=False)
        return 总重80吨以上明细

    def 总重80吨以上明细(self):
        df_数据汇总, df_报修点位统计, df_接入数 = self.data_source()

        # U_汇总_站点表 = pd.merge(df_数据汇总, t_sys_station, left_on='站点名称', right_on='station_name', how='left')
        station_code = df_数据汇总['station_code']

        sql = {
            "tableName": "t_bas_over_data_31 a left join t_bas_over_data_collection_31 b on a.record_code=b.record_code ",
            "where": "a.out_station_time between '{} 00:00:00' and '{} 23:59:59' and a.total_weight>84   AND a.is_unusual = 0  and a.allow is null ".format(
                self.start_time, self.end_time),
            "columns": "a.out_station,a.direction,a.out_station_time  检测时间,a.valid_time 入库时间,a.car_no  车牌号码,a.total_weight  总重,a.limit_weight 限重,a.overrun  超重,a.overrun_rate  超限率,a.axis  轴数,b.status 案件状态,a.site_name 站点名称,a.is_collect 证据满足,b.law_judgment 判定需处罚,b.make_copy  外省抄告,b.link_man 所属运输企业名称,b.phone 联系电话, b.vehicle_county 车籍地,a.photo1,a.photo2,a.photo3,a.vedio,a.record_code 流水号"
        }
        t_bas_over_data_31_80 = self.get_df_from_db(sql)
        df_big = t_bas_over_data_31_80[t_bas_over_data_31_80.loc[:, 'out_station'].isin(station_code)]

        sql_station = {
            "tableName": "t_sys_station  ",
            "where": "  is_deleted = 0   and station_type = 31 ",
            "columns": "station_code,station_name, area_county"
        }

        sql_area = {
            "tableName": "t_code_area  ",
            "where": "  is_deleted = 0  ",
            "columns": "city_code,county_code,city,county"
        }
        t_sys_station = self.get_df_from_db(sql_station)
        t_sys_station = pd.DataFrame(t_sys_station,
                                     columns=['station_code',
                                              'station_name',
                                              'area_county'])
        t_code_area = self.get_df_from_db(sql_area)
        t_code_area = pd.DataFrame(t_code_area,
                                   columns=['city_code', 'county_code', 'city', 'county'])
        df_big['out_station'] = df_big['out_station'].astype('str')
        t_sys_station['station_code'] = t_sys_station['station_code'].astype('str')
        df_big = pd.merge( df_big,t_sys_station,  left_on='out_station',right_on='station_code', how='left')
        ##市际站点区域变更
        市际站点 = pd.read_excel(r'{}'.format(self.file_name), sheet_name='市际站点')
        市际站点['direction'] = pd.to_numeric(市际站点['direction'], errors='coerce')
        市际站点['area_county2'] = pd.to_numeric(市际站点['area_county2'], errors='coerce')
        # 市际站点['out_station']=市际站点['out_station'].astype('str')
        # 市际站点['direction'] = 市际站点['direction'].astype('int')
        市际站点['area_county2'] = 市际站点['area_county2'].astype('str')
        # wide_table['out_station']=wide_table['out_station'].astype('str')
        df_big['direction'] = pd.to_numeric(df_big['direction'], errors='coerce')
        df_big['direction'] = df_big['direction'].fillna(1)
        df_big['direction'] = df_big['direction'].astype('int')
        df_big['原始归属区县1'] = df_big['area_county'].astype('str')
        df_big['area_county'] = df_big['area_county'].astype('str')
        df_big = pd.merge(df_big, 市际站点, on=['station_code', 'direction'], how='left')
        # wide_table.to_excel(r'C:\Users\liu.wenjie\Desktop\wide_table.xlsx')
        df_big['area_county2'] = df_big['area_county2'].fillna(0)
        df_big.loc[
            ((df_big['area_county'] != df_big['area_county2']) & (
                    df_big['area_county2'] != 0)), 'area_county'] = \
            df_big['area_county2']
        # wide_table.to_excel(r'C:\Users\liu.wenjie\Desktop\wide_table市际站点1.xlsx')
        t_code_area.rename(
            columns={'city': '地市', 'county': '区县'}, inplace=True)
        df_big = pd.merge(df_big, t_code_area, left_on='area_county', right_on='county_code', how='left')
        df_big = pd.DataFrame(df_big,
                              columns=["地市", "区县", "站点名称", "检测时间", "车牌号码", "总重", "限重",
                                       "超重", "轴数", "超限率", "证据满足", "所属运输企业名称", "联系电话", "车籍地", "流水号","photo1", "photo2", "photo3", "vedio", "入库时间", "案件状态",
                                       "判定需处罚",
                                       "外省抄告", "area_county", "原始归属区县1","out_station", "direction"])
        t_code_area.rename(
            columns={'地市': '地市2', '区县': '原始站点区县'}, inplace=True)
        df_big = pd.merge(df_big, t_code_area, left_on='原始归属区县1', right_on='county_code', how='left')
        df_big = pd.DataFrame(df_big,
                              columns=["地市", "区县", "站点名称", "检测时间", "车牌号码", "总重", "限重",
                                       "超重", "轴数", "超限率", "证据满足", "所属运输企业名称", "联系电话", "车籍地", "流水号","photo1", "photo2", "photo3", "vedio", "入库时间", "案件状态",
                                       "判定需处罚",
                                       "外省抄告", "原始站点区县","out_station", "direction","area_county"])
        df_big['总重'] = df_big['总重'].astype('float')
        df_big['超限率'] = df_big['超限率'].astype('float')
        df_big = df_big[~((df_big.地市 == '杭州') & (df_big.总重 >= 100))]
        df_big['案件状态'] = df_big['案件状态'].fillna(0, inplace=False)
        df_big['案件状态'] = df_big['案件状态'].astype('int')

        df_big = df_big[~((df_big.区县 == '桐庐') & (df_big.超限率 >= 100) & (df_big['车牌号码'].str.contains("浙") == False))]
        df_big = df_big[~((df_big.区县 == '富阳') & (df_big.超限率 >= 100) & (df_big['车牌号码'].str.contains('浙') == False))]
        ##案件状态
        df_big = df_big[~((df_big.区县 == '桐庐') & (df_big.超限率 < 100) & (df_big.案件状态 == 0))]
        df_big = df_big[~((df_big.区县 == '桐庐') & (df_big.超限率 < 100) & (df_big.案件状态 == 1))]
        df_big = df_big[~((df_big.区县 == '桐庐') & (df_big.超限率 < 100) & (df_big.案件状态 == 2))]
        df_big = df_big[~((df_big.区县 == '桐庐') & (df_big.超限率 < 100) & (df_big.案件状态 == 15))]
        df_big = df_big[~((df_big.区县 == '富阳') & (df_big.超限率 < 100) & (df_big.案件状态 == 0))]
        df_big = df_big[~((df_big.区县 == '富阳') & (df_big.超限率 < 100) & (df_big.案件状态 == 1))]
        df_big = df_big[~((df_big.区县 == '富阳') & (df_big.超限率 < 100) & (df_big.案件状态 == 2))]
        df_big = df_big[~((df_big.区县 == '富阳') & (df_big.超限率 < 100) & (df_big.案件状态 == 15))]

        df_big.loc[df_big['案件状态'] == 0, '案件状态'] = '未采集'
        df_big.loc[df_big['案件状态'] == 1, '案件状态'] = '待初审'
        df_big.loc[df_big['案件状态'] == 2, '案件状态'] = '待审核'
        df_big.loc[df_big['案件状态'] == 3, '案件状态'] = '待判定'
        df_big.loc[df_big['案件状态'] == 4, '案件状态'] = '已告知'
        df_big.loc[df_big['案件状态'] == 5, '案件状态'] = '免处理'
        df_big.loc[df_big['案件状态'] == 6, '案件状态'] = '已立案'
        df_big.loc[df_big['案件状态'] == 12, '案件状态'] = '待告知'
        df_big.loc[df_big['案件状态'] == 13, '案件状态'] = '已结案'
        df_big.loc[df_big['案件状态'] == 9, '案件状态'] = '判定不处理'
        df_big.loc[df_big['案件状态'] == 15, '案件状态'] = '初审不通过'
        df_big = df_big[~((df_big.地市 == '杭州') & (df_big.总重 >= 100))]

        总重80吨以上明细 = df_big.sort_values(by=['area_county'],
                                       ascending=True).reset_index(drop=True)
        with pd.ExcelWriter(r'{}\总重80吨以上明细.xlsx'.format(self.out_file_月报)) as writer1:
            总重80吨以上明细.to_excel(writer1, sheet_name='明细', index=False)

    def 超限100明细(self):
        df_数据汇总, df_报修点位统计, df_接入数 = self.data_source()

        station_code = df_数据汇总['station_code']

        sql = {
            "tableName": "t_bas_over_data_31 a left join t_bas_over_data_collection_31 b on a.record_code=b.record_code left join t_code_area c on a.area_county=c.county_code ",
            "where": "a.out_station_time between '{} 00:00:00' and '{} 23:59:59' and a.overrun_rate>=105 and a.total_weight<100   AND a.is_unusual = 0  and a.allow is null ".format(
                self.start_time, self.end_time),
            "columns": "a.out_station,a.direction,a.out_station_time  检测时间,a.valid_time 入库时间,a.car_no  车牌号码,a.total_weight  总重,a.limit_weight 限重,a.overrun  超重,a.overrun_rate  超限率,a.axis  轴数,b.status 状态,a.site_name 站点名称,a.is_collect 是否是需采集的数据,a.is_unusual 异常数据,b.make_copy  外省抄告,a.photo1,a.photo2,a.photo3,a.vedio,a.record_code 流水号"
        }
        df_100 = self.get_df_from_db(sql)

        U_过车_站点表 = df_100[df_100.loc[:, 'out_station'].isin(station_code)]

        sql_station = {
            "tableName": "t_sys_station  ",
            "where": "  is_deleted = 0   and station_type = 31 ",
            "columns": "station_code,station_name, area_county"
        }

        sql_area = {
            "tableName": "t_code_area  ",
            "where": "  is_deleted = 0  ",
            "columns": "city_code,county_code,city,county"
        }
        t_sys_station = self.get_df_from_db(sql_station)
        t_sys_station = pd.DataFrame(t_sys_station,
                                     columns=['station_code',
                                              'station_name',
                                              'area_county'])
        t_code_area = self.get_df_from_db(sql_area)
        t_code_area = pd.DataFrame(t_code_area,
                                   columns=['city_code', 'county_code', 'city', 'county'])
        U_过车_站点表['out_station'] = U_过车_站点表['out_station'].astype('str')
        t_sys_station['station_code'] = t_sys_station['station_code'].astype('str')
        U_过车_站点表 = pd.merge(U_过车_站点表, t_sys_station, left_on='out_station', right_on='station_code', how='left')
        ##市际站点区域变更
        市际站点 = pd.read_excel(r'{}'.format(self.file_name), sheet_name='市际站点')
        市际站点['direction'] = pd.to_numeric(市际站点['direction'], errors='coerce')
        市际站点['area_county2'] = pd.to_numeric(市际站点['area_county2'], errors='coerce')
        # 市际站点['out_station']=市际站点['out_station'].astype('str')
        # 市际站点['direction'] = 市际站点['direction'].astype('int')
        市际站点['area_county2'] = 市际站点['area_county2'].astype('str')
        # wide_table['out_station']=wide_table['out_station'].astype('str')
        U_过车_站点表['direction'] = pd.to_numeric(U_过车_站点表['direction'], errors='coerce')
        U_过车_站点表['direction'] = U_过车_站点表['direction'].fillna(1)
        U_过车_站点表['direction'] = U_过车_站点表['direction'].astype('int')
        U_过车_站点表['原始站点区县1'] = U_过车_站点表['area_county'].astype('str')
        U_过车_站点表['area_county'] = U_过车_站点表['area_county'].astype('str')
        U_过车_站点表 = pd.merge(U_过车_站点表, 市际站点, on=['station_code', 'direction'], how='left')
        # wide_table.to_excel(r'C:\Users\liu.wenjie\Desktop\wide_table.xlsx')
        U_过车_站点表['area_county2'] = U_过车_站点表['area_county2'].fillna(0)
        U_过车_站点表.loc[
            ((U_过车_站点表['area_county'] != U_过车_站点表['area_county2']) & (
                    U_过车_站点表['area_county2'] != 0)), 'area_county'] = \
            U_过车_站点表['area_county2']
        # wide_table.to_excel(r'C:\Users\liu.wenjie\Desktop\wide_table市际站点1.xlsx')
        t_code_area.rename(
            columns={'city': '地市', 'county': '区县'}, inplace=True)

        U_过车_站点表 = pd.merge(U_过车_站点表, t_code_area, left_on='area_county', right_on='county_code', how='left')
        U_过车_站点表 = pd.DataFrame(U_过车_站点表,
                              columns=["地市", "区县", "车牌号码", "检测时间", "轴数", "总重",
                                         "超重","超限率", "站点名称", "状态", "是否是需采集的数据", "异常数据",
                                         "流水号", "photo1", "photo2", "photo3", "vedio", 'area_county', "原始站点区县1", "out_station", "direction"])

        t_code_area.rename(
            columns={'city_code': '地市编码', '区县': '原始站点区县','地市': '地市1'}, inplace=True)
        U_过车_站点表['原始站点区县1']=U_过车_站点表['原始站点区县1'].astype('str')
        U_过车_站点表 = pd.merge(U_过车_站点表, t_code_area, left_on='原始站点区县1', right_on='county_code', how='left')
        U_过车_站点表 = pd.DataFrame(U_过车_站点表,
                              columns=["地市", "区县", "车牌号码", "检测时间", "轴数", "总重",
                                         "超重","超限率", "站点名称", "状态", "是否是需采集的数据", "异常数据",
                                         "流水号", "photo1", "photo2", "photo3", "vedio", '地市编码','county_code', "原始站点区县", "out_station", "direction"])
        U_过车_站点表.rename(
            columns={'county_code': '区县编码'}, inplace=True)



        i = {
            "330100": 1.1,
            "330200": 1,
            "330300": 1,
            "330400": 1.1,
            "330500": 1,
            "330600": 1,
            "330700": 1,
            "330800": 1.1,
            "330900": 1.1,
            "331000": 1.1,
            "331100": 1.1,
            "330122": 1.091,
            "330183": 1.1,
            "330329": 1.2,
            "330523": 1.03,
            "330603": 1.1,
            "330604": 1.2,
            "330624": 1.2,
            "330681": 1.1,
            "330703": 1.2,
            "330782": 1.1
        }
        U_过车_站点表['总重'] = U_过车_站点表['总重'].astype('float')
        U_过车_站点表['超重'] = U_过车_站点表['超重'].astype('float')
        U_过车_站点表['limit_weight'] = U_过车_站点表['总重'] - U_过车_站点表['超重']
        U_过车_站点表['limit_weight'] = U_过车_站点表['limit_weight'].astype('float')
        U_过车_站点表.loc[U_过车_站点表['limit_weight'] < 0, 'limit_weight'] = 0.0001
        for item in i.items():
            key = item[0]
            value = item[1]

            U_过车_站点表.loc[((U_过车_站点表['地市编码'] == key) | (U_过车_站点表['区县编码'] == key)) & (
                    U_过车_站点表['总重'] < 100), 'vehicle_brand'] = U_过车_站点表['limit_weight'].map(
                lambda x: float(x) * value).round(4)
        U_过车_站点表['limit_weight'] = U_过车_站点表['limit_weight'].astype('float')
        U_过车_站点表['总重'] = U_过车_站点表['总重'].astype('float')
        U_过车_站点表['vehicle_brand'] = U_过车_站点表['vehicle_brand'].astype('float')

        U_过车_站点表['超限率100'] = U_过车_站点表.apply(
            lambda x: (x['总重'] - x['vehicle_brand']) - x['vehicle_brand'],
            axis=1).round(2)
        U_过车_站点表 = U_过车_站点表[(U_过车_站点表['超限率100'] >= 0)
        ]
        U_过车_站点表 = U_过车_站点表.fillna(0)
        U_过车_站点表 = pd.DataFrame(U_过车_站点表,
                                columns=["地市", "区县", "车牌号码", "检测时间", "轴数", "总重",
                                         "超重","超限率", "站点名称", "状态", "是否是需采集的数据", "异常数据",
                                         "流水号", "photo1", "photo2", "photo3", "vedio", '区县编码', "原始站点区县", "out_station", "direction"])

        # 超限100数 = U_过车_站点表.groupby(["地市", "区县"])['流水号'].count().reset_index(name='超限100')

        U_过车_站点表 = U_过车_站点表.sort_values(by=['区县编码', '地市', '区县', '检测时间'],
                                        ascending=True).reset_index(drop=True)
        U_过车_站点表 = U_过车_站点表.drop_duplicates(subset=['流水号'])
        U_过车_站点表['状态'] = U_过车_站点表['状态'].astype('int')
        U_过车_站点表['是否是需采集的数据'] = U_过车_站点表['是否是需采集的数据'].astype('int')
        U_过车_站点表 = U_过车_站点表[~((U_过车_站点表.区县 == '桐庐') & (U_过车_站点表['车牌号码'].str.contains('浙') == False))]
        U_过车_站点表 = U_过车_站点表[~((U_过车_站点表.区县 == '富阳') & (U_过车_站点表['车牌号码'].str.contains('浙') == False))]
        U_过车_站点表.loc[U_过车_站点表['状态'] == 0, '状态'] = '未采集'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 1, '状态'] = '待初审'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 2, '状态'] = '待审核'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 3, '状态'] = '待判定'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 4, '状态'] = '已告知'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 5, '状态'] = '免处理'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 6, '状态'] = '已立案'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 12, '状态'] = '待告知'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 13, '状态'] = '已结案'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 9, '状态'] = '判定不处理'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 15, '状态'] = '初审不通过'
        U_过车_站点表.loc[U_过车_站点表['是否是需采集的数据'] == 1, '是否是需采集的数据'] = '满足'
        U_过车_站点表.loc[U_过车_站点表['是否是需采集的数据'] == 0, '是否是需采集的数据'] = '不满足'

        """车牌处理"""
        U_过车_站点表['车牌号码'].fillna('无牌', inplace=True)
        U_过车_站点表['字节数'] = U_过车_站点表['车牌号码'].str.len()
        U_过车_站点表.loc[U_过车_站点表['字节数'] <= 5, '车牌号码'] = '无牌'
        U_过车_站点表.loc[~((U_过车_站点表['字节数'] <= 5) & (U_过车_站点表['是否是需采集的数据'] == '满足')), 'photo1'] = ''
        U_过车_站点表.loc[~((U_过车_站点表['字节数'] <= 5) & (U_过车_站点表['是否是需采集的数据'] == '满足')), 'photo2'] = ''
        U_过车_站点表.loc[~((U_过车_站点表['字节数'] <= 5) & (U_过车_站点表['是否是需采集的数据'] == '满足')), 'photo3'] = ''
        U_过车_站点表.loc[~((U_过车_站点表['字节数'] <= 5) & (U_过车_站点表['是否是需采集的数据'] == '满足')), 'vedio'] = ''

        with pd.ExcelWriter(r'{}\超限100明细.xlsx'.format(self.out_file_月报)) as writer1:
            U_过车_站点表.to_excel(writer1, sheet_name='超限100明细', index=False)

    def 百吨王明细(self):
        df_数据汇总, df_报修点位统计, df_接入数 = self.data_source()

        station_code = df_数据汇总['station_code']

        sql = {
            "tableName": "t_bas_over_data_31 a left join t_bas_over_data_collection_31 b on a.record_code=b.record_code left join t_code_area c on a.area_county=c.county_code ",
            "where": "a.out_station_time between '{} 00:00:00' and '{} 23:59:59'  and a.total_weight >=100   AND a.is_unusual = 0  and a.allow is null ".format(
                self.start_time, self.end_time),
            "columns": "a.out_station,a.direction,a.out_station_time  检测时间,a.valid_time 入库时间,a.car_no  车牌号码,a.total_weight  总重,a.limit_weight 限重,a.overrun  超重,a.overrun_rate  超限率,a.axis  轴数,b.status 状态,a.site_name 站点名称,a.is_collect 是否是需采集的数据,a.is_unusual 异常数据,b.make_copy  外省抄告,a.photo1,a.photo2,a.photo3,a.vedio,a.record_code 流水号"
        }
        U_过车_站点表 = self.get_df_from_db(sql)
        U_过车_站点表 = U_过车_站点表[U_过车_站点表.loc[:, 'out_station'].isin(station_code)]
        sql_station = {
            "tableName": "t_sys_station  ",
            "where": "  is_deleted = 0   and station_type = 31 ",
            "columns": "station_code,station_name, area_county"
        }

        sql_area = {
            "tableName": "t_code_area  ",
            "where": "  is_deleted = 0  ",
            "columns": "city_code,county_code,city,county"
        }
        t_sys_station = self.get_df_from_db(sql_station)
        t_sys_station = pd.DataFrame(t_sys_station,
                                     columns=['station_code',
                                              'station_name',
                                              'area_county'])
        t_code_area = self.get_df_from_db(sql_area)
        t_code_area = pd.DataFrame(t_code_area,
                                   columns=['city_code', 'county_code', 'city', 'county'])
        U_过车_站点表['out_station'] = U_过车_站点表['out_station'].astype('str')
        t_sys_station['station_code'] = t_sys_station['station_code'].astype('str')
        U_过车_站点表 = pd.merge(U_过车_站点表, t_sys_station, left_on='out_station', right_on='station_code', how='left')
        ##市际站点区域变更
        市际站点 = pd.read_excel(r'{}'.format(self.file_name), sheet_name='市际站点')
        市际站点['direction'] = pd.to_numeric(市际站点['direction'], errors='coerce')
        市际站点['area_county2'] = pd.to_numeric(市际站点['area_county2'], errors='coerce')
        # 市际站点['out_station']=市际站点['out_station'].astype('str')
        # 市际站点['direction'] = 市际站点['direction'].astype('int')
        市际站点['area_county2'] = 市际站点['area_county2'].astype('str')
        # wide_table['out_station']=wide_table['out_station'].astype('str')
        U_过车_站点表['direction'] = pd.to_numeric(U_过车_站点表['direction'], errors='coerce')
        U_过车_站点表['direction'] = U_过车_站点表['direction'].fillna(1)
        U_过车_站点表['direction'] = U_过车_站点表['direction'].astype('int')
        U_过车_站点表['原始站点区县1'] = U_过车_站点表['area_county'].astype('str')
        U_过车_站点表['area_county'] = U_过车_站点表['area_county'].astype('str')
        U_过车_站点表 = pd.merge(U_过车_站点表, 市际站点, on=['station_code', 'direction'], how='left')
        # wide_table.to_excel(r'C:\Users\liu.wenjie\Desktop\wide_table.xlsx')
        U_过车_站点表['area_county2'] = U_过车_站点表['area_county2'].fillna(0)
        U_过车_站点表.loc[
            ((U_过车_站点表['area_county'] != U_过车_站点表['area_county2']) & (
                    U_过车_站点表['area_county2'] != 0)), 'area_county'] = \
            U_过车_站点表['area_county2']
        # wide_table.to_excel(r'C:\Users\liu.wenjie\Desktop\wide_table市际站点1.xlsx')
        t_code_area.rename(
            columns={'city': '地市', 'county': '区县'}, inplace=True)

        U_过车_站点表 = pd.merge(U_过车_站点表, t_code_area, left_on='area_county', right_on='county_code', how='left')
        U_过车_站点表 = pd.DataFrame(U_过车_站点表,
                              columns=["地市", "区县", "车牌号码", "检测时间", "轴数", "总重",
                                         "超重","超限率", "站点名称", "状态", "是否是需采集的数据", "异常数据",
                                         "流水号", "photo1", "photo2", "photo3", "vedio", 'area_county', "原始站点区县1", "out_station", "direction"])

        t_code_area.rename(
            columns={'city_code': '地市编码', '区县': '原始站点区县','地市': '地市1'}, inplace=True)
        U_过车_站点表['原始站点区县1']=U_过车_站点表['原始站点区县1'].astype('str')
        U_过车_站点表 = pd.merge(U_过车_站点表, t_code_area, left_on='原始站点区县1', right_on='county_code', how='left')
        U_过车_站点表 = pd.DataFrame(U_过车_站点表,
                              columns=["地市", "区县", "车牌号码", "检测时间", "轴数", "总重",
                                         "超重","超限率", "站点名称", "状态", "是否是需采集的数据", "异常数据",
                                         "流水号", "photo1", "photo2", "photo3", "vedio", '地市编码','county_code', "原始站点区县", "out_station", "direction"])
        U_过车_站点表.rename(
            columns={'county_code': '区县编码'}, inplace=True)


        U_过车_站点表 = U_过车_站点表.fillna(0)

        U_过车_站点表 = U_过车_站点表.sort_values(by=['区县编码', '地市', '区县', '检测时间'],
                                        ascending=True).reset_index(drop=True)
        U_过车_站点表['总重'] = U_过车_站点表['总重'].astype('float')
        U_过车_站点表 = U_过车_站点表[~((U_过车_站点表.地市 == '杭州') & (U_过车_站点表.总重 >= 100))]
        U_过车_站点表['状态'] = U_过车_站点表['状态'].astype('int')
        U_过车_站点表['是否是需采集的数据'] = U_过车_站点表['是否是需采集的数据'].astype('int')
        U_过车_站点表.loc[U_过车_站点表['状态'] == 0, '状态'] = '未采集'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 1, '状态'] = '待初审'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 2, '状态'] = '待审核'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 3, '状态'] = '待判定'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 4, '状态'] = '已告知'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 5, '状态'] = '免处理'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 6, '状态'] = '已立案'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 12, '状态'] = '待告知'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 13, '状态'] = '已结案'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 9, '状态'] = '判定不处理'
        U_过车_站点表.loc[U_过车_站点表['状态'] == 15, '状态'] = '初审不通过'
        U_过车_站点表.loc[U_过车_站点表['是否是需采集的数据'] == 1, '是否是需采集的数据'] = '满足'
        U_过车_站点表.loc[U_过车_站点表['是否是需采集的数据'] == 0, '是否是需采集的数据'] = '不满足'

        U_过车_站点表['区县编码'] = U_过车_站点表['区县编码'].astype('str')
        with pd.ExcelWriter(r'{}\百吨王明细.xlsx'.format(self.out_file_月报)) as writer1:
            U_过车_站点表.to_excel(writer1, sheet_name='百吨王明细', index=False)

    def Compliance_rate(self):
        df_big = pd.read_excel(r'{}\总重80吨以上明细.xlsx'.format(self.out_file_月报), sheet_name='明细')
        sql = {
            "tableName": "t_bas_over_data_collection_31 ",
            "where": "out_station_time between '{} 00:00:00' and  '{} 00:00:00' and total_weight>80 and law_judgment=1".format(
                self.start_time, self.end_time),
            "columns": "area_county,out_station_time,valid_time,status,total_weight,record_code 流水号"
        }
        t_bas_over_data_collection_31 = self.get_df_from_db(sql)

        总重80吨以上数 = df_big.groupby(['area_county'])['流水号'].count().reset_index(name='本月超限80吨以上')
        df_big['总重'] = df_big['总重'].astype('float')
        总重90吨以上数 = df_big[df_big['总重'] > 90].groupby(['area_county'])['流水号'].count().reset_index(name='本月超限90吨以上')
        总重80_90以上 = pd.merge(总重80吨以上数, 总重90吨以上数, on=['area_county'], how='left')

        U_过车_站点表 = df_big.copy()
        U_过车_站点表 = U_过车_站点表[(U_过车_站点表['证据满足'] == 1)]

        月超限80以上且满足处罚条件总数 = U_过车_站点表.groupby(['area_county'])['流水号'].count().reset_index(name='80吨以上且满足')

        月超限80以上审核通过总数 = t_bas_over_data_collection_31.groupby(['area_county'])['流水号'].count().reset_index(
            name='80吨以上且审核通过')

        总重80吨以上相关 = pd.merge(总重80_90以上, 月超限80以上且满足处罚条件总数, on=['area_county'], how='left')
        总重80吨以上相关['area_county'] = 总重80吨以上相关['area_county'].astype('str')
        总重80吨以上相关 = pd.merge(总重80吨以上相关, 月超限80以上审核通过总数, on=['area_county'], how='outer')

        总重80吨以上相关 = 总重80吨以上相关.fillna(0, inplace=False)
        总重80吨以上相关.rename(columns={'area_county': '区县编码'}, inplace=True)
        return 总重80吨以上相关, df_big

    def over_100_num(self):
        df_数据汇总, df_报修点位统计, df_接入数 = self.data_source()

        U_过车_站点表 = pd.read_excel(r'{}\超限100明细.xlsx'.format(self.out_file_月报), sheet_name='超限100明细')

        超限100数 = U_过车_站点表.groupby(["地市", "区县"])['流水号'].count().reset_index(name='超限100')
        满足证据条件数 = \
            U_过车_站点表[
                ((U_过车_站点表['车牌号码'] == '无牌') & (U_过车_站点表['是否是需采集的数据'] == '满足') & (U_过车_站点表['异常数据'] == '是'))].groupby(
                ["地市", "区县"])[
                '流水号'].count().reset_index(name='超限100%遮挡车牌数量（辆）')
        超限100数 = pd.merge(超限100数, 满足证据条件数, on=["地市", "区县"], how='left')

        """地市货车数及超限100%数"""
        df_数据汇总2 = df_数据汇总[((df_数据汇总.实际在线天数 >= 10)
                            & (df_数据汇总.货车数 > 500))]
        货车数 = df_数据汇总2.groupby(['地市'])['货车数'].sum()
        货车数 = 货车数.to_frame()
        货车数 = 货车数.reset_index()
        超限100 = U_过车_站点表.groupby(['地市'])['流水号'].count()
        超限100 = 超限100.to_frame()
        超限100 = 超限100.reset_index()
        附件2 = pd.merge(货车数, 超限100, left_on='地市', right_on='地市', how='left', )
        附件2.rename(columns={'流水号': '超限100%数'}, inplace=True)
        附件2 = 附件2.fillna(0, inplace=False)
        附件2['超限100%数占货车数比例'] = 附件2.apply(lambda x: x['超限100%数'] / x['货车数'], axis=1)
        附件2['排名（占比由高到低）'] = 附件2['超限100%数占货车数比例'].rank(ascending=False, method='first')
        附件2['超限100%数占货车数比例'] = 附件2['超限100%数占货车数比例'].apply(lambda x: format(x, '.3%'))
        未识别到车牌数 = U_过车_站点表[U_过车_站点表['车牌号码'] == '无牌'].groupby(['地市'])['车牌号码'].count().reset_index(name='未识别到车牌')
        满足证据条件数 = U_过车_站点表[((U_过车_站点表['车牌号码'] == '无牌') & (U_过车_站点表['是否是需采集的数据'] == '满足'))].groupby(['地市'])[
            '车牌号码'].count().reset_index(name='满足证据条件')
        满足证据条件且故意遮挡车牌 = \
            U_过车_站点表[
                ((U_过车_站点表['车牌号码'] == '无牌') & (U_过车_站点表['是否是需采集的数据'] == '满足') & (U_过车_站点表['异常数据'] == '是'))].groupby(
                ['地市'])[
                '车牌号码'].count().reset_index(name='满足证据条件且故意遮挡车牌')
        附件2 = pd.merge(df_接入数, 附件2, on='地市', how='left')
        附件2 = pd.merge(附件2, 未识别到车牌数, left_on='地市', right_on='地市', how='left')
        附件2 = pd.merge(附件2, 满足证据条件数, left_on='地市', right_on='地市', how='left')
        附件2 = pd.merge(附件2, 满足证据条件且故意遮挡车牌, left_on='地市', right_on='地市', how='left')
        附件2 = pd.DataFrame(附件2, columns=['地市', '货车数', '超限100%数', '超限100%数占货车数比例', '排名（占比由高到低）', '未识别到车牌', '满足证据条件',
                                         '满足证据条件且故意遮挡车牌'])
        附件2 = 附件2.fillna(0, inplace=False)
        附件2 = 附件2.drop(index=(附件2.loc[(附件2['地市'] == '义乌')].index))

        return 超限100数, 附件2, U_过车_站点表

    def total_weight_100_num(self):
        U_过车_站点表 = pd.read_excel(r'{}\百吨王明细.xlsx'.format(self.out_file_月报), sheet_name='百吨王明细')
        百吨王数 = U_过车_站点表.groupby(['区县编码'])['流水号'].count().reset_index(name='百吨王数')
        百吨王遮挡车牌数量 = U_过车_站点表[(U_过车_站点表['异常数据'] == '车牌附近安装大灯，货物遮盖，无法识别装载物，无法确认是否超限')].groupby(["区县编码"])[
            '流水号'].count().reset_index(name='百吨王遮挡车牌数量（辆）')
        百吨王数 = pd.merge(百吨王数, 百吨王遮挡车牌数量, on=["区县编码"], how='left')
        百吨王数['区县编码'] = 百吨王数['区县编码'].astype('str')
        return 百吨王数, U_过车_站点表

    def case_statistic(self):
        sql_area = {
            "tableName": "t_code_area  ",
            "where": "  is_deleted = 0 and province_code = '330000' ",
            "columns": "city,city_code,county,county_code as area_county"
        }
        t_code_area = self.get_df_from_db(sql_area)
        sql_非现入库数 = {
            "tableName": "t_bas_over_data_collection_31  ",
            "where": " law_judgment = 1  AND valid_time between '{} 00:00:00'  AND '{} 23:59:59'  and status !=5 GROUP BY area_county  ".format(
                self.q案件, self.s案件),
            "columns": "area_county, count( 1 )  入库数路政, sum(IF( car_no LIKE '%浙%', 1, 0 ))  本省入库数  "
        }

        sql_非现入库数 = self.get_df_from_db(sql_非现入库数)
        sql_非现入库数 = pd.DataFrame(sql_非现入库数, columns=['area_county', '入库数路政', '本省入库数'])

        sql_交通现场查处数 = {
            "tableName": " t_bas_over_data_collection_sign c",
            "where": " c.record_type = 99 AND c.insert_type = 5  AND c.close_case_time between '{} 00:00:00'  and '{} 23:59:59'  GROUP BY c.area_county  ".format(
                self.q案件, self.s案件),
            "columns": "c.area_county  ,count( DISTINCT ( record_id ) ) AS 现场处罚路政 "
        }
        sql_交通现场查处数 = self.get_df_from_db(sql_交通现场查处数)
        sql_交通现场查处数 = pd.DataFrame(sql_交通现场查处数, columns=['area_county', '现场处罚路政'])
        sql_交通现场查处数.loc[sql_交通现场查处数['area_county'] == '331083', 'area_county'] = '331021'
        sql_交通现场查处数.loc[sql_交通现场查处数['area_county'] == '330284', 'area_county'] = '330204'

        sql_非现处罚数处罚 = {
            "tableName": "t_case_sign_result  ",
            "where": " record_type = 31 AND insert_type = 1 AND data_source = 1 AND case_type = 1 AND close_case_time between '{} 00:00:00' AND  '{} 23:59:59' AND area_province = '330000' GROUP BY  dept_county".format(
                self.q案件, self.s案件),
            "columns": "dept_county as area_county,count(DISTINCT ( CASE_NUM )) AS 非现场处罚路政处罚,  count(DISTINCT IF( car_no LIKE '%浙%', CASE_NUM, NULL )) AS 非现场处罚本省路政处罚 "
        }
        sql_非现处罚数处罚 = self.get_df_from_db(sql_非现处罚数处罚)
        sql_非现处罚数处罚 = pd.DataFrame(sql_非现处罚数处罚, columns=['area_county', '非现场处罚路政处罚', '非现场处罚本省路政处罚'])
        sql_非现处罚数处罚.loc[sql_非现处罚数处罚['area_county'] == '331083', 'area_county'] = '331021'
        sql_非现处罚数处罚.loc[sql_非现处罚数处罚['area_county'] == '330284', 'area_county'] = '330204'

        sql_非现当年处罚数处罚 = {
            "tableName": "t_case_sign_result c LEFT JOIN t_bas_over_data_collection_31 b ON c.record_code = b.record_code  ",
            "where": " record_type = 31 AND insert_type = 1 AND data_source = 1 AND case_type = 1 AND close_case_time between '{} 00:00:00' AND  '{} 23:59:59' AND c.area_province = '330000' and  b.valid_time between '{} 00:00:00' and  '{} 23:59:59' GROUP BY  c.dept_county".format(
                self.q案件, self.s案件, self.q案件, self.s案件),
            "columns": "c.dept_county as  area_county,count(DISTINCT ( c.CASE_NUM )) AS 非现场处罚路政处罚当年, count(DISTINCT IF( c.car_no LIKE '%浙%', c.CASE_NUM, NULL )) AS 非现场处罚本省路政处罚当年  "
        }
        sql_非现当年处罚数处罚 = self.get_df_from_db(sql_非现当年处罚数处罚)
        sql_非现当年处罚数处罚 = pd.DataFrame(sql_非现当年处罚数处罚, columns=['area_county', '非现场处罚路政处罚当年', '非现场处罚本省路政处罚当年'])
        sql_非现当年处罚数处罚.loc[sql_非现当年处罚数处罚['area_county'] == '331083', 'area_county'] = '331021'
        sql_非现当年处罚数处罚.loc[sql_非现当年处罚数处罚['area_county'] == '330284', 'area_county'] = '330204'

        sql_交警现场 = {
            "tableName": "   t_bas_police_road_site   ",
            "where": "   punish_time between '{} 00:00:00' AND  '{} 23:59:59'  and case_status=2  GROUP BY area_county  ".format(
                self.q案件, self.s案件),
            "columns": "area_county,count(DISTINCT case_number) as 交警现场查处数"
        }
        sql_交警现场 = self.get_df_from_db(sql_交警现场)
        sql_交警现场 = pd.DataFrame(sql_交警现场, columns=['area_county', '交警现场查处数'])

        sql_外省抄告数 = {
            "tableName": "t_bas_over_data_collection_31  ",
            "where": " law_judgment = 1  AND valid_time between '{} 00:00:00'  AND '{} 23:59:59'  and status !=5 GROUP BY area_county  ".format(
                self.q案件, self.s案件),
            "columns": "area_county ,sum(IF( make_copy=1 and car_no not LIKE '%浙%', 1, 0 )) 外省抄告 "
        }
        sql_外省抄告数 = self.get_df_from_db(sql_外省抄告数)
        sql_外省抄告数 = pd.DataFrame(sql_外省抄告数, columns=['area_county', '外省抄告'])

        sql_非现处罚数案发 = {
            "tableName": "t_case_sign_result ",
            "where": " record_type = 31 AND insert_type = 1 AND data_source = 1 AND case_type = 1 AND close_case_time between '{} 00:00:00' AND  '{} 23:59:59' AND area_province = '330000' GROUP BY  area_county".format(
                self.q案件, self.s案件),
            "columns": "area_county,count(DISTINCT ( CASE_NUM )) AS 非现场处罚路政案发,  count(DISTINCT IF( car_no LIKE '%浙%', CASE_NUM, NULL )) AS 非现场处罚本省路政案发 "
        }
        sql_非现处罚数案发 = self.get_df_from_db(sql_非现处罚数案发)
        sql_非现处罚数案发 = pd.DataFrame(sql_非现处罚数案发, columns=['area_county', '非现场处罚路政案发', '非现场处罚本省路政案发'])

        sql_非现当年处罚数案发 = {
            "tableName": "t_case_sign_result c LEFT JOIN t_bas_over_data_collection_31 b ON c.record_code = b.record_code  ",
            "where": " record_type = 31 AND insert_type = 1 AND data_source = 1 AND case_type = 1 AND close_case_time between '{} 00:00:00' AND  '{} 23:59:59' AND c.area_province = '330000' and  b.valid_time between '{} 00:00:00' and  '{} 23:59:59' GROUP BY  c.area_county".format(
                self.q案件, self.s案件, self.q案件, self.s案件),
            "columns": "c.area_county,count(DISTINCT ( c.CASE_NUM )) AS 非现场处罚路政案发当年, count(DISTINCT IF( c.car_no LIKE '%浙%', c.CASE_NUM, NULL )) AS 非现场处罚本省路政案发当年  "
        }
        sql_非现当年处罚数案发 = self.get_df_from_db(sql_非现当年处罚数案发)
        sql_非现当年处罚数案发 = pd.DataFrame(sql_非现当年处罚数案发, columns=['area_county', '非现场处罚路政案发当年', '非现场处罚本省路政案发当年'])

        case = pd.merge(t_code_area, sql_非现入库数, on='area_county', how='outer')
        case = pd.merge(case, sql_交通现场查处数, on='area_county', how='outer')
        case = pd.merge(case, sql_非现处罚数处罚, on='area_county', how='outer')
        case = pd.merge(case, sql_非现当年处罚数处罚, on='area_county', how='outer')
        case = pd.merge(case, sql_交警现场, on='area_county', how='outer')
        case = pd.merge(case, sql_外省抄告数, on='area_county', how='outer')
        case = pd.merge(case, sql_非现处罚数案发, on='area_county', how='outer')
        case = pd.merge(case, sql_非现当年处罚数案发, on='area_county', how='outer')
        case = case.fillna(0)
        case['本省入库数'] = case['本省入库数'].astype('int')
        case['外省抄告'] = case['外省抄告'].astype('int')
        case.rename(columns={'area_county': '区县编码'}, inplace=True)
        case['区县编码'] = case['区县编码'].astype('str')
        with pd.ExcelWriter(r'{}\case_statistic.xlsx'.format(self.out_file_月报)) as writer1:
            case.to_excel(writer1, sheet_name='df_案件', index=False)

    def data_station_source(self):
            from datetime import datetime
            day = datetime.now().date()  # 获取当前系统时间

            ##报修站点
            maintain_station = {
                "tableName": " t_sys_station_maintain a inner join t_sys_station b on a.station_code=b.station_code and b.station_status in (0,3)    ".format(
                    self.today2),
                "where": "   end_time>='{} 00:00:00'   or end_time is null ".format(self.start_time),
                "columns": "b.station_name as '报修站点名称',a.station_code,reason as '备注',b.area_county,b.station_type "
            }
            maintain_station = self.get_df_from_db(maintain_station)
            maintain_station = maintain_station[(maintain_station.station_type == 31)]

            sql_station = {
                "tableName": "t_sys_station   ",
                "where": "  is_deleted = 0  and station_status =1   ",
                "columns": " station_code"
            }
            停用站点 = self.get_df_from_db(sql_station)

            import datetime

            now = day - datetime.timedelta(days=0)

            from datetime import datetime

            starttime = self.start_time
            endtime = self.end_time
            # starttime1 = datetime.strptime(starttime, '%Y-%m-%d')
            # endtime1 = datetime.strptime(endtime, '%Y-%m-%d')
            starttime1 = starttime
            endtime1 = endtime
            import datetime
            # endtime1 = endtime1 + datetime.timedelta(days=1)
            理应在线天数 = 31

            # ##在用站点
            # online_station = total_station[~total_station.loc[:, 'station_code'].isin(maintain_station['station_code'])]

            ##在用站点明细数据
            pass_truck_num = {
                "tableName": "t_bas_basic_data_pass    ",
                "where": " statistic_date  >='{}' and statistic_date  <='{}' and station_type =31 group by statistic_date,station_code ".format(
                    starttime,
                    endtime),
                "columns": "station_code,statistic_date,sum(truck_num) as truck_num,sum(overrun_num) as overrun_num,sum(no_car_num) as no_car_num,sum(overrun_0_10) as overrun_0_10,sum(overrun_10_20) as overrun_10_20"
            }
            pass_truck_num = self.get_df_from_db(pass_truck_num)
            pass_truck_num['truck_num']=pass_truck_num['truck_num'].astype('int')
            pass_truck_num['overrun_num'] = pass_truck_num['overrun_num'].astype('int')
            pass_truck_num['no_car_num'] = pass_truck_num['no_car_num'].astype('int')
            pass_truck_num['overrun_0_10'] = pass_truck_num['overrun_0_10'].astype('int')
            pass_truck_num['overrun_10_20'] = pass_truck_num['overrun_10_20'].astype('int')
            t_code_area = {
                "tableName": " t_sys_station a left join t_code_area  b on a.area_county=b.county_code and a.station_type =31  ",
                "where": "   province_code='330000' and a.is_deleted = 0 and b.is_deleted = 0",
                "columns": "b.city as city_name,county as county_name,county_code,a.station_code,a.station_name "
            }
            t_code_area = self.get_df_from_db(t_code_area)
            pass_truck_num = pd.merge(pass_truck_num, t_code_area, how='left',
                                 on=['station_code'])
            # pass_truck_num = pass_truck_num[pass_truck_num.loc[:, 'station_code'].isin(online_station['station_code'])]
            df = pass_truck_num.groupby(
                ['city_name', 'county_code', 'county_name', 'station_code', 'station_name']).sum().reset_index()

            df = df[~df.loc[:, 'station_code'].isin(maintain_station['station_code'])]
            df = df[~df.loc[:, 'station_code'].isin(停用站点['station_code'])]
            pass_truck_num = pass_truck_num[(0 < pass_truck_num['truck_num'])]
            pass_truck_num['statistic_date'] = pd.to_datetime(pass_truck_num['statistic_date'])
            pass_truck_num['取日'] = pass_truck_num['statistic_date'].apply(lambda x: x.strftime('%d'))
            实际在线天数 = pass_truck_num.groupby(['station_code'])['取日'].nunique().reset_index(name='实际在线天数')
            df = pd.merge(df, 实际在线天数, on=['station_code'], how='left')
            df['理应在线天数'] = 理应在线天数

            ##异常站点在线率
            # df.loc[(df['county_name'] == '开化'), '实际在线天数'] = df[
            #     '实际在线天数'].map(lambda x: float(x) + 10).round(0)
            # df.loc[(df['station_name'] == 'X701（虎十线）龙游、衢州方向K100+300'), 'truck_num'] = df[
            #     'truck_num'].map(lambda x: float(x) + 510).round(2)
            # df.loc[(df['station_name'] == 'G527（象义线）佛堂方向K277+300'), 'truck_num'] = df[
            #     'truck_num'].map(lambda x: float(x) + 515).round(2)
            # df.loc[(df['station_name'] == 'G235（新海线）义乌方向K841+050'), 'truck_num'] = df[
            #     'truck_num'].map(lambda x: float(x) + 510).round(2)
            # df.loc[(df['station_name'] == 'G351（台小线）义乌方向K235+000'), 'truck_num'] = df[
            #     'truck_num'].map(lambda x: float(x) + 510).round(2)
            # df.loc[(df['station_name'] == 'G527（象义线）佛堂方向K277+300'), '实际在线天数'] = df[
            #     'truck_num'].map(lambda x: float(x) + 10).round(0)
            df.loc[(df['station_name'] == 'G235（新海线）义乌方向K841+050'), '实际在线天数'] = df[
                'truck_num'].map(lambda x: float(x) + 10).round(0)
            df.loc[(df['station_name'] == 'G351（台小线）义乌方向K235+000'), '实际在线天数'] = df[
                'truck_num'].map(lambda x: float(x) + 10).round(0)


            df['在线率'] = (df['实际在线天数'] / df['理应在线天数'] * 100).round(2)
            df['百吨王数'] = 0
            df['超限100%数'] = 0
            # city_name,county_code,county_name,a.station_code,a.station_name,statistic_date,truck_num,
            # overrun_num,no_car_num,overrun_0_10,overrun_10_20
            df['超限率(%)'] = (df['overrun_num'] / df['truck_num'] * 100).round(2)
            df['超限10%除外数'] = df['overrun_num'] - df['overrun_0_10']
            df['超限10%除外超限率(%)'] = (df['超限10%除外数'] / df['truck_num'] * 100).round(2)
            df['超限20%除外数'] = df['overrun_num'] - df['overrun_0_10'] - df['overrun_10_20']
            df['超限20%除外超限率(%)'] = (df['超限20%除外数'] / df['truck_num'] * 100).round(2)
            df['最后接收时间'] = now
            df = df.fillna(value=0)

            df.rename(
                columns={'station_name': '站点名称',
                         'city_name': '地市', 'county_name': '区县', 'truck_num': '货车数',
                         'overrun_num': '超限数'}, inplace=True)

            df = pd.DataFrame(df,
                              columns=['站点名称', '地市', '区县', '理应在线天数', '实际在线天数', '在线率', '货车数', '超限数', '百吨王数', '超限100%数',
                                       '超限10%除外超限率(%)', '超限20%除外数', '超限20%除外超限率(%)', '超限率(%)', '最后接收时间', 'county_code',
                                       'station_code'])
            # df = df[(df.站点名称 != '舟山本岛环岛公路鲁家峙至东港公路东港方向K0+060')]
            df_站点在线明细 = df.sort_values('county_code', ascending=True).reset_index(drop=True)
            # df_站点在线明细.loc[df_站点在线明细['区县'] == '普朱管委会', '区县'] = '新城管委会'
            with pd.ExcelWriter(r'{}\data_station_source.xlsx'.format(self.out_file_月报)) as writer1:
                df_站点在线明细.to_excel(writer1, sheet_name='df_站点在线明细', index=False)
            # return df_站点在线明细

    def data_station(self):
        df_数据汇总, df_报修点位统计, df_接入数 = self.data_source()
        # df_数据汇总=self.data_station_source()
        df_数据汇总 = pd.read_excel(r'{}\data_station_source.xlsx'.format(self.out_file_月报), sheet_name='df_站点在线明细')
        T_义乌_数据为0 = df_数据汇总[(df_数据汇总.区县 == '义乌')
                            & (df_数据汇总.实际在线天数 == 0)
                            ]
        T_义乌_数据为0数 = T_义乌_数据为0.区县.count()
        T_义乌_汇总 = df_数据汇总[(df_数据汇总.区县 == '义乌')
                          & (df_数据汇总.实际在线天数 < 10)
                          & (df_数据汇总.实际在线天数 > 0)]
        T_义乌_10数 = T_义乌_汇总.区县.count()
        T_义乌_汇总 = df_数据汇总[(df_数据汇总.区县 == '义乌')
                          & (df_数据汇总.实际在线天数 >= 10)
                          & (df_数据汇总.货车数 < 500)
                          ]
        T_义乌_500数 = T_义乌_汇总.区县.count()
        T_义乌_10_500数 = T_义乌_10数 + T_义乌_500数
        T_义乌_报修 = df_报修点位统计[df_报修点位统计.区县 == '义乌']
        T_义乌_报修数 = T_义乌_报修.区县.count()
        T_义乌_在线 = df_数据汇总[df_数据汇总.区县 == '义乌']
        T_义乌_在线数 = T_义乌_在线.区县.count()
        df_数据汇总1 = df_数据汇总[(df_数据汇总.区县 != '义乌')
        ]
        在用数 = df_数据汇总1.groupby(['地市'])['站点名称'].count()
        df_报修点位统计1 = df_报修点位统计[df_报修点位统计.区县 != '义乌']
        报修数 = df_报修点位统计1.groupby(['地市'])['报修站点名称'].count()
        实际站点数 = pd.merge(在用数, 报修数, on='地市', how='outer')
        实际站点数 = 实际站点数.fillna(0, inplace=False)
        实际站点数['实际站点数'] = 实际站点数.apply(lambda x: x[0] + x[1], axis=1)
        实际站点数.实际站点数 = 实际站点数.实际站点数.astype(int)
        实际站点数.报修站点名称 = 实际站点数.报修站点名称.astype(int)
        实际站点数.rename(columns={'站点名称': '在用数', '报修站点名称': '实际报修数'}, inplace=True)
        T_10筛选 = df_数据汇总1[(df_数据汇总1.实际在线天数 < 10)
                          & (df_数据汇总1.实际在线天数 > 0)
                          ]
        T_10筛选 = T_10筛选.groupby([T_10筛选.地市]).count()
        T_10筛选 = T_10筛选.loc[:, ['站点名称']]
        T_10筛选.columns = ['站点名称']
        T_500筛选 = df_数据汇总1[((df_数据汇总1.实际在线天数 >= 10)
                            & (df_数据汇总1.货车数 < 500))]
        T_500筛选 = T_500筛选.groupby([T_500筛选.地市]).count()
        T_500筛选 = T_500筛选.loc[:, ['站点名称']]
        T_500筛选.columns = ['站点名称']
        T_500筛选.rename(columns={'站点名称': '500辆次'}, inplace=True)
        T_筛选 = pd.merge(T_10筛选, T_500筛选, how='outer', on='地市')
        T_筛选 = T_筛选.fillna(value=0)
        T_筛选['在线天数＜10天或货车数＜500辆次'] = T_筛选['站点名称'] + T_筛选['500辆次']
        T_10或500 = T_筛选['在线天数＜10天或货车数＜500辆次']
        t = df_数据汇总1[(df_数据汇总1.实际在线天数 == 0)]
        T_数据为0数 = t.groupby([t['地市']]).count()
        T_数据为0数 = T_数据为0数.loc[:, ['实际在线天数']]
        T_数据为0数.columns = ['实际在线天数']
        T_数据为0数.rename(columns={'实际在线天数': '数据为0'}, inplace=True)
        站点设备完好率 = pd.merge(df_接入数, 实际站点数, on='地市', how='left')
        站点设备完好率.loc[11, '在用数'] = T_义乌_在线数
        站点设备完好率.loc[11, '实际报修数'] = T_义乌_报修数
        站点设备完好率.loc[11, '实际站点数'] = T_义乌_报修数 + T_义乌_在线数
        站点设备完好率 = 站点设备完好率.fillna(0, inplace=False)
        站点设备完好率['报修数'] = 站点设备完好率.apply(lambda x: x['接入数（修正后）'] - x['在用数'], axis=1)
        站点设备完好率.loc[站点设备完好率['报修数'] < 0, '报修数'] = 0
        站点设备完好率 = pd.merge(站点设备完好率, T_数据为0数, on='地市', how='left')
        站点设备完好率 = pd.merge(站点设备完好率, T_10或500, on='地市', how='left')
        站点设备完好率.loc[11, '数据为0'] = T_义乌_数据为0数
        站点设备完好率.loc[11, '在线天数＜10天或货车数＜500辆次'] = T_义乌_10_500数
        站点设备完好率 = 站点设备完好率.fillna(0, inplace=False)

        站点设备完好率['实际完好数'] = 站点设备完好率['在用数'] - 站点设备完好率.数据为0 - 站点设备完好率['在线天数＜10天或货车数＜500辆次']
        站点设备完好率['修正完好数'] = 站点设备完好率.apply(lambda x: min(x['接入数（修正后）'], x['实际完好数']), axis=1)
        站点设备完好率['实际完好率'] = 站点设备完好率.apply(lambda x: x['实际完好数'] / x['实际站点数'], axis=1).round(4)
        站点设备完好率['实际完好率'] = 站点设备完好率['实际完好率'].apply(lambda x: format(x, '.2%'))
        站点设备完好率['调整后完好率'] = 站点设备完好率.apply(lambda x: x['修正完好数'] / x['接入数（修正后）'], axis=1).round(4)
        站点设备完好率['调整后完好率'] = 站点设备完好率['调整后完好率'].apply(lambda x: format(x, '.2%'))
        站点设备完好率 = pd.DataFrame(站点设备完好率,
                               columns=["地市", "接入数（修正后）", "实际站点数", "在用数", "实际报修数", "报修数", "数据为0", "在线天数＜10天或货车数＜500辆次",
                                        "实际完好数", "实际完好率", "修正完好数", "调整后完好率"])

        df_数据汇总, df_报修点位统计, df_接入数 = self.data_source()
        """做sheet1"""

        df_数据汇总 = df_数据汇总[((df_数据汇总.实际在线天数 >= 10)
                           & (df_数据汇总.货车数 > 500))]
        T_义乌 = df_数据汇总[(df_数据汇总['区县'] == '义乌')]
        T_义乌_货车数 = T_义乌.groupby(['区县'])['货车数'].sum()
        T_义乌_超限数 = T_义乌.groupby(['区县'])['超限20%除外数'].sum()
        df_数据汇总1 = df_数据汇总[(df_数据汇总.区县 != '义乌')
        ]
        货车数 = df_数据汇总1.groupby(['地市'])['货车数'].sum()
        超限数 = df_数据汇总1.groupby(['地市'])['超限20%除外数'].sum()
        df_sheet1 = pd.merge(货车数, 超限数, how='left', on='地市')
        df_sheet1 = pd.merge(df_接入数, df_sheet1, how='left', on='地市')
        df_sheet1.loc[11, '货车数'] = T_义乌_货车数[0]
        df_sheet1.loc[11, '超限20%除外数'] = T_义乌_超限数[0]
        df_sheet1['超限率'] = df_sheet1.apply(lambda x: x['超限20%除外数'] / x['货车数'], axis=1).round(4)
        df_sheet1['超限率'] = df_sheet1['超限率'].apply(lambda x: format(x, '.2%'))
        df_sheet1['超限率'] = pd.to_numeric(df_sheet1['超限率'], errors='coerce')
        df_sheet1['超限率排名'] = df_sheet1['超限率'].rank(ascending=False, method='first')
        df_sheet1后 = pd.DataFrame(站点设备完好率, columns=["地市", "实际站点数", "实际完好率", "调整后完好率"])
        df_sheet1 = pd.merge(df_sheet1, df_sheet1后, how='left', on='地市')
        df_sheet1 = pd.DataFrame(df_sheet1,
                                 columns=["地市", "货车数", "超限20%除外数", "超限率", "超限率排名", "实际站点数", "接入数（修正后）", "实际完好率",
                                          "调整后完好率"])
        df_sheet1.rename(columns={'超限20%除外数': '超限数'}, inplace=True)

        """在线天数大于等于10天货车数大于500的站点数据"""

        T_20天与500筛选 = df_数据汇总[((df_数据汇总.实际在线天数 >= 10)
                               & (df_数据汇总.货车数 > 500))]
        T_20天与500筛选 = pd.DataFrame(T_20天与500筛选,
                                   columns=["站点名称", "地市", "区县", "理应在线天数", "实际在线天数", "在线率", "货车数", "超限20%除外数",
                                            "超限20%除外超限率(%)"])
        T_20天与500筛选.rename(columns={'理应在线天数': '应在线天数', '超限20%除外数': '超限数', '超限20%除外超限率(%)': '超限率'}, inplace=True)
        T_20天与500筛选1 = T_20天与500筛选.sort_values(by="超限率", ascending=False).reset_index(drop=True)

        '''区县超限率排名'''
        T_20天与500筛选 = pd.DataFrame(T_20天与500筛选1, columns=["地市", "区县", "货车数", "超限数"])
        区县超限率排序 = T_20天与500筛选.groupby([T_20天与500筛选['地市'], T_20天与500筛选['区县']]).sum()
        区县超限率排序['超限率'] = 区县超限率排序.apply(lambda x: x['超限数'] / x['货车数'], axis=1).round(4)
        区县超限率排序 = 区县超限率排序.sort_values('超限率', ascending=False)
        区县超限率排序['超限率'] = 区县超限率排序['超限率'].apply(lambda x: format(x, '.2%'))
        区县超限率排序 = 区县超限率排序.reset_index()
        df_数据汇总, df_报修点位统计, df_接入数 = self.data_source()
        with pd.ExcelWriter(r'{}\data_station.xlsx'.format(self.out_file_月报)) as writer1:
            df_数据汇总.to_excel(writer1, sheet_name='df_数据汇总', index=False)
            df_报修点位统计.to_excel(writer1, sheet_name='df_报修点位统计', index=False)
            df_sheet1.to_excel(writer1, sheet_name='df_sheet1', index=False)
            站点设备完好率.to_excel(writer1, sheet_name='站点设备完好率', index=False)
            区县超限率排序.to_excel(writer1, sheet_name='区县超限率排序', index=False)
            T_20天与500筛选1.to_excel(writer1, sheet_name='T_20天与500筛选1', index=False)
        # return df_数据汇总, df_报修点位统计, df_sheet1, 站点设备完好率, 区县超限率排序, T_20天与500筛选1

    def overrun_site_rate(self):
        """附件1"""
        df_数据汇总, df_报修点位统计, df_接入数 = self.data_source()
        df_数据汇总2 = df_数据汇总[((df_数据汇总.实际在线天数 >= 10)
                            & (df_数据汇总.货车数 > 500))]
        货车数 = df_数据汇总2.groupby(['地市', '区县'])['货车数'].sum().reset_index()
        超限数 = df_数据汇总2.groupby(['地市', '区县'])['超限20%除外数'].sum().reset_index()
        df_sheet1 = pd.merge(货车数, 超限数, how='left', on=['地市', '区县'])
        df_sheet1.loc[df_sheet1['货车数'] == 0, '货车数'] = 1
        df_sheet1['超限率'] = df_sheet1.apply(lambda x: x['超限20%除外数'] / x['货车数'], axis=1).round(4)
        df_sheet1 = df_sheet1.fillna(0, inplace=False)
        df_sheet1['超限率'] = df_sheet1['超限率'].apply(lambda x: format(x, '.2%'))
        df_sheet1.loc[df_sheet1['货车数'] == 1, '货车数'] = 0
        df_sheet1.rename(columns={'超限20%除外数': '超限数'}, inplace=True)

        """附件7"""

        df_数据汇总, df_报修点位统计, df_接入数 = self.data_source()
        df_数据汇总 = pd.read_excel(r'{}\data_station_source.xlsx'.format(self.out_file_月报), sheet_name='df_站点在线明细')
        # df_数据汇总.columns = df_数据汇总.iloc[2]
        # df_数据汇总 = df_数据汇总.iloc[3:].reset_index(drop=True)

        在用数 = df_数据汇总.groupby(['地市', '区县'])['站点名称'].count()
        报修数 = df_报修点位统计.groupby(['地市', '区县'])['报修站点名称'].count()
        实际站点数 = pd.merge(在用数, 报修数, on=['地市', '区县'], how='outer')
        实际站点数 = 实际站点数.fillna(0, inplace=False)
        实际站点数['站点数'] = 实际站点数.apply(lambda x: x[0] + x[1], axis=1)
        实际站点数.站点数 = 实际站点数.站点数.astype(int)
        实际站点数.报修站点名称 = 实际站点数.报修站点名称.astype(int)
        实际站点数.rename(columns={'站点名称': '在用数', '报修站点名称': '报修数'}, inplace=True)

        T_10筛选 = df_数据汇总[(df_数据汇总.实际在线天数 < 10)
                         & (df_数据汇总.实际在线天数 > 0)
                         ]
        T_10筛选 = T_10筛选.groupby(['地市', '区县']).count()
        T_10筛选 = T_10筛选.loc[:, ['站点名称']]
        T_10筛选.columns = ['站点名称']
        T_500筛选 = df_数据汇总[(df_数据汇总.实际在线天数 >= 10) & (df_数据汇总.货车数 < 500)]
        T_500筛选 = T_500筛选.groupby(['地市', '区县']).count()
        T_500筛选 = T_500筛选.loc[:, ['站点名称']]
        T_500筛选.columns = ['站点名称']
        T_500筛选.rename(columns={'站点名称': '500辆次'}, inplace=True)
        T_筛选 = pd.merge(T_10筛选, T_500筛选, how='outer', on=['地市', '区县'])
        T_筛选 = T_筛选.fillna(value=0)
        T_筛选['在线天数＜10天或货车数＜500辆次'] = T_筛选['站点名称'] + T_筛选['500辆次']
        T_10或500 = T_筛选['在线天数＜10天或货车数＜500辆次']
        t = df_数据汇总[(df_数据汇总.实际在线天数 == 0)]
        T_数据为0数 = t.groupby(['地市', '区县']).count()

        T_数据为0数 = T_数据为0数.loc[:, ['实际在线天数']]
        T_数据为0数.columns = ['实际在线天数']
        T_数据为0数.rename(columns={'实际在线天数': '数据为0'}, inplace=True)
        附件7 = pd.merge(实际站点数, T_数据为0数, how='left', on=['地市', '区县'])
        附件7 = pd.merge(附件7, T_10或500, how='left', on=['地市', '区县'])
        附件7 = 附件7.fillna(0, inplace=False)
        附件7['异常数'] = 附件7.apply(lambda x: x['报修数'] + x['数据为0'] + x['在线天数＜10天或货车数＜500辆次'], axis=1)
        附件7['异常数'] = 附件7['异常数'].astype('float')
        附件7['站点数'] = 附件7['站点数'].astype('float')
        附件7['设备完好率'] = (附件7['站点数'] - 附件7['报修数']) / 附件7['站点数']
        附件7['数据完好率'] = (附件7['站点数'] - 附件7['异常数']) / 附件7['站点数']
        附件7 = 附件7.fillna(0, inplace=False)
        附件7['数据完好率'] = 附件7['数据完好率'].apply(lambda x: format(x, '.2%'))
        附件7['设备完好率'] = 附件7['设备完好率'].apply(lambda x: format(x, '.2%'))
        附件7 = pd.DataFrame(附件7, columns=["站点数", "报修数", '设备完好率', "异常数", "数据完好率"]).reset_index()
        附件7 = pd.merge(df_sheet1, 附件7, how='outer', on=['地市', '区县'])
        return 附件7
        # q = input("请输入存储路径(C:/Users/Administrator/Desktop/输出报表/其他市月报表)：")

    def Key_freight_sources(self):
        """ 引入原始表 """
        t_code_area = {
            "tableName": "t_code_area ",
            "where": "province_code = '330000'",
            "columns": "city,county,city_code,county_code"
        }
        t_code_area = self.get_df_from_db(t_code_area)
        # sql = "SELECT * from t_bas_source_company where area_province = '330000' and is_statictis =1  and is_deleted=0 "
        # t_bas_source_company=get_df_from_db(sql)
        # sql = "SELECT * FROM t_bas_source_company_equipment WHERE is_deleted = 0 OR is_deleted IS NULL"
        # t_bas_source_company_equipment=get_df_from_db(sql)
        站点总数 = pd.read_excel(r"{}\重点货运源头445家明细0309.xlsx".format(self.file_name_静态表))
        站点总数['county_code'] = 站点总数['county_code'].astype('str')

        数据站点总数 = 站点总数.groupby(['city', 'city_code', 'county', 'county_code'])['id'].count().reset_index(name='数据站点总数')

        数据站点总数['county_code'] = 数据站点总数['county_code'].astype('str')

        t_sys_station = {
            "tableName": "t_sys_station ",
            "where": "station_type = 71 AND is_deleted = 0 AND station_status != 1  and station_code IS NOT null",
            "columns": "station_name,station_code,station_status,station_type,area_county  "
        }
        t_sys_station = self.get_df_from_db(t_sys_station)
        # sql = "SELECT * FROM t_sys_station WHERE station_type = 71 AND is_deleted = 0 AND station_status != 1  and station_code IS NOT null "
        # t_sys_station = get_df_from_db(sql)

        t_bas_pass_data_71 = {
            "tableName": "t_bas_pass_data_71 ",
            "where": " out_station_time between '{} 00:00:00' and '{} 00:00:00'  and is_truck =1 ".format(
                self.start_time, self.end_time),
            "columns": "area_city,area_county,out_station,out_station_time,car_no,total_weight,limit_weight,overrun,overrun_rate,axis,site_name"
        }
        t_bas_pass_data_71 = self.get_df_from_db(t_bas_pass_data_71)
        t_bas_pass_data_71['total_weight'] = t_bas_pass_data_71['total_weight'].astype('float')
        t_bas_pass_data_71['overrun_rate'] = t_bas_pass_data_71['overrun_rate'].astype('float')
        t_bas_pass_data_71['overrun'] = t_bas_pass_data_71['overrun'].astype('float')
        t_bas_pass_data_71['out_station_time'] = pd.to_datetime(t_bas_pass_data_71['out_station_time'])

        """拼接表"""
        U_源头_区域表 = pd.merge(t_bas_pass_data_71, t_code_area, left_on='area_county', right_on='county_code', how='left')

        企业_源头_站点表 = pd.merge(t_sys_station, t_code_area, left_on='area_county', right_on='county_code', how='left')

        station_code = 企业_源头_站点表['station_code']
        U_源头_区域表 = U_源头_区域表[U_源头_区域表.loc[:, 'out_station'].isin(station_code)]
        # q = input("请输入储存路径(C:/Users/Administrator/Desktop/输出报表/月报表相关数据汇总)：")
        # with pd.ExcelWriter('{}/U_源头_区域表.xlsx'.format(q))as writer1:
        #      U_源头_区域表.to_excel(writer1, sheet_name='sheet1', index=True)

        """超限数"""
        超限20_50 = U_源头_区域表[(U_源头_区域表['overrun_rate'] > 20) & (U_源头_区域表['overrun_rate'] <= 50)].groupby(
            ['city', 'county', 'county_code'])['car_no'].count().reset_index(name='20-50%数')
        超限50_100 = U_源头_区域表[(U_源头_区域表['overrun_rate'] > 50) & (U_源头_区域表['overrun_rate'] <= 100)].groupby(
            ['city', 'county', 'county_code'])['car_no'].count().reset_index(name='50-100%数')
        超限100 = U_源头_区域表[(U_源头_区域表['overrun_rate'] > 100) & (U_源头_区域表['overrun_rate'] <= 450)].groupby(
            ['city', 'county', 'county_code'])['car_no'].count().reset_index(name='100%以上数')

        """设备上线率"""

        U_源头_区域表['取日'] = U_源头_区域表['out_station_time'].apply(lambda x: x.strftime('%d'))
        在线天数大于20天 = U_源头_区域表.groupby(['city', 'county_code', 'county', 'out_station', 'site_name'])[
            '取日'].nunique().reset_index(
            name='在线天数')
        货运量大于2万吨 = U_源头_区域表.groupby(['city', 'county_code', 'county', 'out_station', 'site_name'])[
            'total_weight'].sum().reset_index(
            name='货运总重')
        过车数大于410辆 = U_源头_区域表.groupby(['city', 'county_code', 'county', 'out_station', 'site_name'])[
            'city'].count().reset_index(
            name='货车数')
        站点完好数 = pd.merge(在线天数大于20天, 货运量大于2万吨, on=['city', 'county_code', 'county', 'out_station', 'site_name'],
                         how='left')
        站点完好数 = pd.merge(站点完好数, 过车数大于410辆, on=['city', 'county_code', 'county', 'out_station', 'site_name'], how='left')
        站点完好数['货运总重'] = pd.to_numeric(站点完好数['货运总重'], errors='coerce')

        站点完好数区县 = 站点完好数.groupby(['city', 'county', 'county_code'])['货车数', '货运总重'].sum()
        在线站点数 = 站点完好数[(站点完好数['在线天数'] > 20) | (站点完好数['货运总重'] > 20000) | (站点完好数['货车数'] > 410)].groupby(
            ['city', 'county', 'county_code'])['out_station'].count().reset_index(name='在线站点数')

        """聚合"""

        货运源头监控数据 = pd.merge(数据站点总数, 超限20_50, on=['city', 'county', 'county_code'], how='left')
        货运源头监控数据 = pd.merge(货运源头监控数据, 超限50_100, on=['city', 'county', 'county_code'], how='left')
        货运源头监控数据 = pd.merge(货运源头监控数据, 超限100, on=['city', 'county', 'county_code'], how='left')
        货运源头监控数据 = pd.merge(货运源头监控数据, 在线站点数, on=['city', 'county', 'county_code'], how='left')
        货运源头监控数据 = pd.merge(货运源头监控数据, 站点完好数区县, on=['city', 'county', 'county_code'], how='left')
        货运源头监控数据 = 货运源头监控数据.fillna(0, inplace=False)
        货运源头监控数据['源头单位平均过车数（辆次）'] = 货运源头监控数据.apply(lambda x: x['货车数'] / (x['数据站点总数'] + 0.0000001), axis=1).round(0)
        货运源头监控数据['设备上线率（%）'] = 货运源头监控数据.apply(lambda x: x['在线站点数'] / (x['数据站点总数']), axis=1)
        货运源头监控数据['20-50%占比'] = 货运源头监控数据.apply(lambda x: x['20-50%数'] / (x['货车数'] + 0.0000001), axis=1)
        货运源头监控数据['50-100%占比'] = 货运源头监控数据.apply(lambda x: x['50-100%数'] / (x['货车数'] + 0.0000001), axis=1)
        货运源头监控数据['100%以上占比'] = 货运源头监控数据.apply(lambda x: x['100%以上数'] / (x['货车数'] + 0.0000001), axis=1)
        货运源头监控数据['设备上线率（%）'] = 货运源头监控数据.apply(lambda x: x['在线站点数'] / (x['数据站点总数']), axis=1)
        货运源头监控数据 = 货运源头监控数据.fillna(0, inplace=False)
        货运源头监控数据['设备上线率（%）'] = 货运源头监控数据['设备上线率（%）'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据['20-50%占比'] = 货运源头监控数据['20-50%占比'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据['50-100%占比'] = 货运源头监控数据['50-100%占比'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据['100%以上占比'] = 货运源头监控数据['100%以上占比'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据.rename(columns={'city': '地市', 'county': '区县'}, inplace=True)
        货运源头监控数据 = pd.DataFrame(货运源头监控数据,
                                columns=['地市', '区县', '数据站点总数', '货车数', '源头单位平均过车数（辆次）', '在线站点数', '设备上线率（%）', '20-50%数',
                                         '50-100%数', '100%以上数', '20-50%占比', '50-100%占比', '100%以上占比', 'city_code',
                                         'county_code'])
        货运源头监控数据 = 货运源头监控数据.sort_values('county_code', ascending=True)

        货运源头监控数据地市 = 货运源头监控数据.groupby(['地市', 'city_code']).sum().reset_index()
        货运源头监控数据地市 = 货运源头监控数据地市.fillna(0, inplace=False)
        货运源头监控数据地市['源头单位平均过车数（辆次）'] = 货运源头监控数据地市.apply(lambda x: x['货车数'] / (x['数据站点总数'] + 0.0000001), axis=1).round(0)
        货运源头监控数据地市['设备上线率（%）'] = 货运源头监控数据地市.apply(lambda x: x['在线站点数'] / (x['数据站点总数']), axis=1)
        货运源头监控数据地市['20-50%占比'] = 货运源头监控数据地市.apply(lambda x: x['20-50%数'] / (x['货车数'] + 0.0000001), axis=1)
        货运源头监控数据地市['50-100%占比'] = 货运源头监控数据地市.apply(lambda x: x['50-100%数'] / (x['货车数'] + 0.0000001), axis=1)
        货运源头监控数据地市['100%以上占比'] = 货运源头监控数据地市.apply(lambda x: x['100%以上数'] / (x['货车数'] + 0.0000001), axis=1)
        货运源头监控数据地市['设备上线率（%）'] = 货运源头监控数据地市.apply(lambda x: x['在线站点数'] / (x['数据站点总数']), axis=1)
        货运源头监控数据地市 = 货运源头监控数据地市.fillna(0, inplace=False)
        货运源头监控数据地市['设备上线率（%）'] = 货运源头监控数据地市['设备上线率（%）'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据地市['20-50%占比'] = 货运源头监控数据地市['20-50%占比'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据地市['50-100%占比'] = 货运源头监控数据地市['50-100%占比'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据地市['100%以上占比'] = 货运源头监控数据地市['100%以上占比'].apply(lambda x: format(x, '.2%'))

        货运源头监控数据地市 = pd.DataFrame(货运源头监控数据地市,
                                  columns=['地市', '数据站点总数', '货车数', '源头单位平均过车数（辆次）', '在线站点数', '设备上线率（%）', '20-50%数',
                                           '50-100%数', '100%以上数', '20-50%占比', '50-100%占比', '100%以上占比', 'city_code'])
        货运源头监控数据地市 = 货运源头监控数据地市.sort_values('city_code', ascending=True)
        货运源头监控数据省 = 货运源头监控数据地市
        货运源头监控数据省['省'] = '浙江省'
        货运源头监控数据省 = 货运源头监控数据省.groupby(['省']).sum().reset_index()
        货运源头监控数据省 = 货运源头监控数据省.fillna(0, inplace=False)
        货运源头监控数据省['源头单位平均过车数（辆次）'] = 货运源头监控数据省.apply(lambda x: x['货车数'] / (x['数据站点总数'] + 0.0000001), axis=1).round(0)
        货运源头监控数据省['设备上线率（%）'] = 货运源头监控数据省.apply(lambda x: x['在线站点数'] / (x['数据站点总数']), axis=1)
        货运源头监控数据省['20-50%占比'] = 货运源头监控数据省.apply(lambda x: x['20-50%数'] / (x['货车数'] + 0.0000001), axis=1)
        货运源头监控数据省['50-100%占比'] = 货运源头监控数据省.apply(lambda x: x['50-100%数'] / (x['货车数'] + 0.0000001), axis=1)
        货运源头监控数据省['100%以上占比'] = 货运源头监控数据省.apply(lambda x: x['100%以上数'] / (x['货车数'] + 0.0000001), axis=1)
        货运源头监控数据省['设备上线率（%）'] = 货运源头监控数据省.apply(lambda x: x['在线站点数'] / (x['数据站点总数']), axis=1)
        货运源头监控数据省 = 货运源头监控数据省.fillna(0, inplace=False)
        货运源头监控数据省['设备上线率（%）'] = 货运源头监控数据省['设备上线率（%）'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据省['20-50%占比'] = 货运源头监控数据省['20-50%占比'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据省['50-100%占比'] = 货运源头监控数据省['50-100%占比'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据省['100%以上占比'] = 货运源头监控数据省['100%以上占比'].apply(lambda x: format(x, '.2%'))
        货运源头监控数据省.rename(columns={'省': '地市'}, inplace=True)
        货运源头监控数据省 = pd.DataFrame(货运源头监控数据省,
                                 columns=['地市', '数据站点总数', '货车数', '源头单位平均过车数（辆次）', '在线站点数', '设备上线率（%）', '20-50%数',
                                          '50-100%数', '100%以上数', '20-50%占比', '50-100%占比', '100%以上占比'])
        货运源头监控数据省市 = pd.concat([货运源头监控数据地市, 货运源头监控数据省])

        # with pd.ExcelWriter(r'C:\Users\liu.wenjie\Desktop\全部源头.xlsx') as writer1:
        #     货运源头监控数据.to_excel(writer1, sheet_name='区县', index=False)
        #     货运源头监控数据地市.to_excel(writer1, sheet_name='地市', index=False)
        #     货运源头监控数据省市.to_excel(writer1, sheet_name='省', index=False)
        #     站点完好数.to_excel(writer1, sheet_name='站点数据明细', index=False)
        货运源头监控数据 = pd.DataFrame(货运源头监控数据,
                                columns=['后面是源头数据', '数据站点总数', '货车数', '源头单位平均过车数（辆次）', '在线站点数', '设备上线率（%）', '20-50%数',
                                         '50-100%数', '100%以上数', '20-50%占比', '50-100%占比', '100%以上占比', 'county_code'])
        货运源头监控数据.rename(columns={'货车数': '源头货车数'}, inplace=True)
        货运源头监控数据.rename(columns={'county_code': '区县编码'}, inplace=True)
        货运源头监控数据['区县编码'] = 货运源头监控数据['区县编码'].astype('str')
        with pd.ExcelWriter(r'{}\Key_freight_sources.xlsx'.format(self.out_file_月报)) as writer1:
            货运源头监控数据.to_excel(writer1, sheet_name='货运源头监控数据', index=False)
            货运源头监控数据省市.to_excel(writer1, sheet_name='货运源头监控数据省市', index=False)
            站点完好数.to_excel(writer1, sheet_name='站点完好数', index=False)
        # return 货运源头监控数据, 货运源头监控数据省市, 站点完好数

    def To_emil(self,city_path, city, this_month, today2):
        import datetime
        now = datetime.datetime.now()
        date = datetime.datetime.strftime(now - datetime.timedelta(days=1), '%Y%m%d')  # 获取日期
        # 邮件配置信息
        smtp_server = 'smtp.qq.com'
        smtp_port = 465
        smtp_ssl = True
        smtp_user = '1399120443@qq.com'
        smtp_password = 'rycnskfchlpkijad'  # 邮箱授权码

        # 发送邮件信息
        sender = '1399120443@qq.com'
        receivers = ['1399120443@qq.com']

        # 邮件正文
        mail_content = '每日全省案件处罚数据({}).xlsx'.format(date)
        message = MIMEMultipart()

        # 邮件信息配置
        message['From'] = '1399120443@qq.com'
        message['To'] = '1399120443@qq.com'
        message['Subject'] = Header('{}{}月初始数据({}).xlsx'.format(city, this_month, today2), 'utf-8')

        csv_file = open(city_path, 'rb').read()
        csv = MIMEApplication(csv_file)
        csv.add_header('Content-Disposition', 'attachment',
                       filename='{}{}月初始数据({}).xlsx'.format(city, this_month, today2))
        message.attach(csv)

        # 发送邮件
        try:
            smtpObj = smtplib.SMTP_SSL(smtp_server, smtp_port)
            smtpObj.login(smtp_user, smtp_password)
            smtpObj.sendmail(sender, receivers, message.as_string())
            print("{}邮件发送成功".format(city))
        except smtplib.SMTPException as e:
            print("Error: 邮件发送失败: ", e)
        # from threading import Timer
        # import datetime
        # """定时1天"""
        # now_time = datetime.datetime.now()
        # next_time = now_time + datetime.timedelta(days=+1)
        # next_year = next_time.date().year
        # next_month = next_time.date().month
        # next_day = next_time.date().day
        # next_time2 = datetime.datetime.strptime(
        #     str(next_year) + "-" + str(next_month) + "-" + str(next_day) + " " + "7:05:00", "%Y-%m-%d %H:%M:%S")
        # timer_start_time2 = (next_time2 - now_time).total_seconds()
        # print('下次复检开始时间', next_time2)
        # t2 = Timer(timer_start_time2, To_emil)  # 此处使用递归调用实现
        # t2.start()

    def write_excel_file2(self,file_path):
        wb = load_workbook(file_path)

        sheet_names = wb.get_sheet_names()
        # print('sheets',sheet_names[:5])

        # 初始化worksheet对象
        # ws = wb.active
        for i in sheet_names:
            ws = wb[i]
            # 定义样式
            fontstyle = Font(name="Microsoft YaHei UI", size=12, bold=True)
            border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                            bottom=Side(style="thin"))
            align = Alignment(horizontal="center", vertical="center")
            fontstyle2 = Font(name="Microsoft YaHei UI", size=12)
            side = Side(
                style="medium",
                # 边框样式，可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
                color="ff66dd",  # 边框颜色，16进制rgb表示
            )
            ##调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 12) * 1.1
                ws.column_dimensions[column].width = adjusted_width

            # ws.column_dimensions['A'].width = 20
            # ws.column_dimensions['I'].width = 20

            alignment_center = Alignment(horizontal='center', vertical='center')

            # 指定区域单元格居中
            ws_area = ws["A1:BK120"]
            for i in ws_area:
                for j in i:
                    j.alignment = alignment_center;

            row_nu = range(ws.max_row)

            # 设置单元格样式
            for i in range(ws.max_column):
                ws.cell(row=1, column=i + 1).font = fontstyle
                ws.cell(row=1, column=i + 1).border = border
                ws.cell(row=1, column=i + 1).alignment = align
                for j in row_nu:
                    ws.row_dimensions[j + 1].height = 20  # 设置第2行高度为30
                    ws.cell(row=j + 2, column=i + 1).font = fontstyle2
                    ws.cell(row=j + 2, column=2).alignment = align
                    ws.cell(row=j + 1, column=i + 1).border = border
            # ws.delete_cols(1)

        # 保存到Excel文件
        wb.save(file_path)

    # import pysnooper
    # @pysnooper.snoop()
    def t_bas_basic_data_report(self):
        # df_数据汇总, df_报修点位统计, df_接入数 = data_source()
        # df_案件 = self.case_statistic()

        df_案件 = pd.read_excel(r'{}\case_statistic.xlsx'.format(self.out_file_月报), sheet_name='df_案件')
        # df_源头, 货运源头监控数据省市, df_源头站点明细 = self.Key_freight_sources()
        df_源头 = pd.read_excel(r'{}\Key_freight_sources.xlsx'.format(self.out_file_月报), sheet_name='货运源头监控数据')
        货运源头监控数据省市 = pd.read_excel(r'{}\Key_freight_sources.xlsx'.format(self.out_file_月报), sheet_name='货运源头监控数据省市')
        df_源头站点明细 = pd.read_excel(r'{}\Key_freight_sources.xlsx'.format(self.out_file_月报), sheet_name='站点完好数')
        # df_数据汇总, df_报修点位统计, df_sheet1, 站点设备完好率, 区县超限率排序, T_20天与500筛选1 = self.data_station()
        df_数据汇总 = pd.read_excel(r'{}\data_station.xlsx'.format(self.out_file_月报), sheet_name='df_数据汇总')
        df_报修点位统计 = pd.read_excel(r'{}\data_station.xlsx'.format(self.out_file_月报), sheet_name='df_报修点位统计')
        df_sheet1 = pd.read_excel(r'{}\data_station.xlsx'.format(self.out_file_月报), sheet_name='df_sheet1')
        站点设备完好率 = pd.read_excel(r'{}\data_station.xlsx'.format(self.out_file_月报), sheet_name='站点设备完好率')
        区县超限率排序 = pd.read_excel(r'{}\data_station.xlsx'.format(self.out_file_月报), sheet_name='区县超限率排序')
        T_20天与500筛选1 = pd.read_excel(r'{}\data_station.xlsx'.format(self.out_file_月报), sheet_name='T_20天与500筛选1')
        超限100数, 超限100汇总, 超限100明细 = self.over_100_num()
        总重80_90以上, 总重80吨以上明细 = self.Compliance_rate()
        百吨王数, 百吨王明细 = self.total_weight_100_num()
        附件7 = self.overrun_site_rate()
        self.df_区县编码['区县编码'] = self.df_区县编码['区县编码'].astype('str')
        百吨王数['区县编码'] = 百吨王数['区县编码'].astype('str')
        总重80_90以上['区县编码'] = 总重80_90以上['区县编码'].astype('str')

        附件7 = pd.merge(self.df_区县编码, 附件7, on=['地市', '区县'], how='outer')
        附件7 = pd.merge(附件7, 超限100数, on=['地市', '区县'], how='left')
        附件7 = pd.merge(附件7, 总重80_90以上, on=['区县编码'], how='left')
        附件7 = pd.merge(附件7, 百吨王数, on=['区县编码'], how='left')
        df_案件['区县编码'] = df_案件['区县编码'].astype('str')
        df_源头['区县编码'] = df_源头['区县编码'].astype('str')


        附件7 = pd.merge(附件7, df_案件, on=['区县编码'], how='outer')
        附件7 = pd.merge(附件7, df_源头, on=['区县编码'], how='outer')
        附件7['超限100%数货车数/万辆'] = 附件7.apply(lambda x: x['超限100'] / (x['货车数'] + 0.0000001) * 10000, axis=1).round(4)
        附件7['源头货车数'] = pd.to_numeric(附件7['源头货车数'], errors='coerce')
        附件7['数据站点总数'] = pd.to_numeric(附件7['数据站点总数'], errors='coerce')
        附件7['在线站点数'] = pd.to_numeric(附件7['在线站点数'], errors='coerce')
        附件7['20-50%数'] = pd.to_numeric(附件7['20-50%数'], errors='coerce')
        附件7['50-100%数'] = pd.to_numeric(附件7['50-100%数'], errors='coerce')
        附件7['100%以上数'] = pd.to_numeric(附件7['100%以上数'], errors='coerce')
        附件7['源头单位平均过车数（辆次）'] = 附件7.apply(lambda x: x['源头货车数'] / (x['数据站点总数'] + 0.0000001), axis=1).round(0)
        附件7['设备上线率（%）'] = 附件7.apply(lambda x: x['在线站点数'] / (x['数据站点总数'] + 0.0000001), axis=1)
        附件7['20-50%占比'] = 附件7.apply(lambda x: x['20-50%数'] / (x['源头货车数'] + 0.0000001), axis=1)
        附件7['50-100%占比'] = 附件7.apply(lambda x: x['50-100%数'] / (x['源头货车数'] + 0.0000001), axis=1)
        附件7['100%以上占比'] = 附件7.apply(lambda x: x['100%以上数'] / (x['源头货车数'] + 0.0000001), axis=1)
        附件7 = 附件7[(附件7.区县编码 != '330621')]
        附件7 = 附件7[(附件7.区县编码 != '330155')]
        附件7 = 附件7.fillna(0, inplace=False)
        附件7['设备上线率（%）'] = 附件7['设备上线率（%）'].apply(lambda x: format(x, '.2%'))
        附件7['20-50%占比'] = 附件7['20-50%占比'].apply(lambda x: format(x, '.2%'))
        附件7['50-100%占比'] = 附件7['50-100%占比'].apply(lambda x: format(x, '.2%'))
        附件7['100%以上占比'] = 附件7['100%以上占比'].apply(lambda x: format(x, '.2%'))
        附件7['80吨以上总数'] = 附件7['本月超限80吨以上']
        附件7.rename(columns={'入库数路政': '入库数(路政)', '现场处罚路政': '现场处罚(路政)', '非现场处罚路政处罚': '非现场处罚(路政)'
            , '非现场处罚本省路政处罚': '非现场处罚本省(路政)', '非现场处罚路政处罚当年': '非现场处罚(路政)当年', '非现场处罚本省路政处罚当年': '非现场处罚本省(路政)当年'
                            }, inplace=True)
        附件7['入库数(路政)'] = 附件7['入库数(路政)'].astype('int')
        附件7['本省入库数'] = 附件7['本省入库数'].astype('int')
        附件7['外省入库数'] = 附件7['入库数(路政)'] - 附件7['本省入库数']
        附件7['非现场处罚(路政)'] = 附件7['非现场处罚(路政)'].astype('int')
        附件7['非现场处罚本省(路政)'] = 附件7['非现场处罚本省(路政)'].astype('int')
        附件7['非现场处罚(路政)当年'] = 附件7['非现场处罚(路政)当年'].astype('int')
        附件7['非现场处罚本省(路政)当年'] = 附件7['非现场处罚本省(路政)当年'].astype('int')
        附件7['交警现场查处数'] = 附件7['交警现场查处数'].astype('int')
        附件7['外省抄告'] = 附件7['外省抄告'].astype('int')
        附件7['现场处罚(路政)'] = 附件7['现场处罚(路政)'].astype('int')
        附件7['非现场处罚外省(路政)'] = 附件7['非现场处罚(路政)'] - 附件7['非现场处罚本省(路政)']
        附件7['统计月份'] = self.this_month
        附件7.loc[(附件7['区县'] == '临平'), '交警现场查处数'] = 附件7[
            '交警现场查处数'].map(lambda x: float(x) + 40).round(0)
        附件7 = pd.DataFrame(附件7,
                           columns=['统计月份', '地市编码', '地市', '区县编码', '区县', '货车数', '超限数', '超限率', '超限100', '超限100%数货车数/万辆',
                                    '本月超限80吨以上', '本月超限90吨以上', '百吨王数', '超限100%遮挡车牌数量（辆）', '百吨王遮挡车牌数量（辆）', '站点数',
                                    '报修数', '设备完好率', '异常数', '数据完好率', '入库数(路政)', '本省入库数', '外省入库数', '现场处罚(路政)',
                                    '非现场处罚(路政)', '非现场处罚本省(路政)', '非现场处罚(路政)当年',
                                    '非现场处罚本省(路政)当年', '非现场处罚外省(路政)', '交警现场查处数', '交警非现场处罚数', '交警非现查处数本省',
                                    '需处罚数/非现入库数（总计）', '需处罚数/非现入库数（本省）',
                                    '需处罚数/非现入库数（外省）', '非现处罚数（总计）', '非现处罚数（本省）', '非现处罚数（外省）', '年处罚数', '外省抄告',
                                    '非现场处罚率（本省）', '非现场处罚率（外省）',
                                    '处罚率(含抄告）', '80吨以上总数', '80吨以上且满足', '80吨以上且审核通过', '合规率', '后面是源头数据', '数据站点总数',
                                    '源头货车数', '源头单位平均过车数（辆次）',
                                    '在线站点数', '设备上线率（%）', '20-50%数', '50-100%数', '100%以上数', '20-50%占比', '50-100%占比',
                                    '100%以上占比', '非现场处罚路政案发', '非现场处罚本省路政案发', '非现场处罚路政案发当年', '非现场处罚本省路政案发当年'])

        ##案件处罚率等处理
        附件7 = 附件7.fillna(0, inplace=False)
        附件7['交警非现场处罚数'] = 附件7['交警非现场处罚数'].astype('int')
        附件7['非现场处罚本省路政案发当年'] = 附件7['非现场处罚本省路政案发当年'].astype('int')
        附件7['非现场处罚本省路政案发当年'] = 附件7['非现场处罚本省路政案发当年'].astype('int')
        # 附件7['交警非现场处罚数'] = 附件7['交警非现场处罚数'].astype('int')

        附件7['需处罚数/非现入库数（总计）'] = 附件7['入库数(路政)'] + 附件7['交警非现场处罚数']
        附件7['需处罚数/非现入库数（本省）'] = 附件7['本省入库数'] + 附件7['交警非现场处罚数']
        附件7['需处罚数/非现入库数（外省）'] = 附件7['外省入库数']
        附件7['非现处罚数（总计）'] = 附件7['非现场处罚(路政)'] + 附件7['交警非现场处罚数']
        附件7['非现处罚数（本省）'] = 附件7['非现场处罚本省(路政)'] + 附件7['交警非现场处罚数']
        附件7['非现处罚数（外省）'] = 附件7['非现场处罚(路政)'] - 附件7['非现场处罚本省(路政)']
        附件7['年处罚数'] = 附件7['非现场处罚(路政)'] + 附件7['交警非现场处罚数'] + 附件7['现场处罚(路政)'] + 附件7['交警现场查处数']

        附件7['非现场处罚率（本省）'] = 附件7.apply(
            lambda x: (x['非现场处罚本省路政案发当年'] + x['交警非现场处罚数']) / (x['本省入库数'] + x['交警非现场处罚数'] + 0.00001), axis=1).round(4)
        附件7['非现场处罚率（外省）'] = 附件7.apply(
            lambda x: (x['非现场处罚路政案发当年'] - x['非现场处罚本省路政案发当年']) / (x['需处罚数/非现入库数（外省）'] + 0.00001),
            axis=1).round(4)
        附件7['处罚率(含抄告）'] = 附件7.apply(
            lambda x: (x['非现场处罚路政案发当年'] + x['交警非现场处罚数'] + x['外省抄告']) / (x['入库数(路政)'] + x['交警非现场处罚数'] + 0.00001),
            axis=1).round(4)
        附件7['合规率'] = 附件7.apply(
            lambda x: (x['80吨以上且审核通过']) / (x['80吨以上总数'] + 0.00001),
            axis=1).round(4)
        附件7.loc[附件7['非现场处罚率（本省）'] > 1, '非现场处罚率（本省）'] = 1
        附件7.loc[附件7['非现场处罚率（外省）'] > 1, '非现场处罚率（外省）'] = 1
        附件7.loc[附件7['处罚率(含抄告）'] > 1, '处罚率(含抄告）'] = 1
        附件7.loc[附件7['合规率'] > 1, '合规率'] = 1

        附件7 = 附件7.fillna(0, inplace=False)
        附件7['非现场处罚率（本省）'] = 附件7['非现场处罚率（本省）'].apply(lambda x: format(x, '.2%'))
        附件7['非现场处罚率（外省）'] = 附件7['非现场处罚率（外省）'].apply(lambda x: format(x, '.2%'))
        附件7['处罚率(含抄告）'] = 附件7['处罚率(含抄告）'].apply(lambda x: format(x, '.2%'))
        附件7['合规率'] = 附件7['合规率'].apply(lambda x: format(x, '.2%'))
        附件7['区县编码']=附件7['区县编码'].astype('str')
        附件7 = 附件7.sort_values(by=['区县编码'],
                              ascending=True).reset_index(drop=True)
        附件7 = 附件7.fillna(0, inplace=False)
        附件7 = 附件7[(附件7.区县 != 0)]
        附件7.drop_duplicates(subset=['区县编码'], keep='first', inplace=True)
        # with pd.ExcelWriter(self.out_file_name) as writer1:
        #     附件7.to_excel(writer1, sheet_name='区县汇总', index=False)
        附件7地市 = 附件7.groupby(['地市编码', '地市']).sum().reset_index()
        附件7地市.loc[附件7地市['货车数'] == 0, '货车数'] = 1
        附件7地市['超限率'] = 附件7地市.apply(lambda x: x['超限数'] / x['货车数'], axis=1).round(4)
        附件7地市 = 附件7地市.fillna(0, inplace=False)
        附件7地市['超限率'] = 附件7地市['超限率'].apply(lambda x: format(x, '.2%'))
        附件7地市.loc[附件7地市['货车数'] == 1, '货车数'] = 0
        附件7地市['设备完好率'] = (附件7地市['站点数'] - 附件7地市['报修数']) / 附件7地市['站点数']
        附件7地市['数据完好率'] = (附件7地市['站点数'] - 附件7地市['异常数']) / 附件7地市['站点数']
        附件7地市 = 附件7地市.fillna(0, inplace=False)
        附件7地市['数据完好率'] = 附件7地市['数据完好率'].apply(lambda x: format(x, '.2%'))
        附件7地市['设备完好率'] = 附件7地市['设备完好率'].apply(lambda x: format(x, '.2%'))
        附件7地市['超限100%数货车数/万辆'] = 附件7地市.apply(lambda x: x['超限100'] / (x['货车数'] + 0.0000001) * 10000, axis=1).round(4)
        附件7地市['源头货车数'] = pd.to_numeric(附件7地市['源头货车数'], errors='coerce')
        附件7地市['数据站点总数'] = pd.to_numeric(附件7地市['数据站点总数'], errors='coerce')
        附件7地市['在线站点数'] = pd.to_numeric(附件7地市['在线站点数'], errors='coerce')
        附件7地市['20-50%数'] = pd.to_numeric(附件7地市['20-50%数'], errors='coerce')
        附件7地市['50-100%数'] = pd.to_numeric(附件7地市['50-100%数'], errors='coerce')
        附件7地市['100%以上数'] = pd.to_numeric(附件7地市['100%以上数'], errors='coerce')
        附件7地市['源头单位平均过车数（辆次）'] = 附件7地市.apply(lambda x: x['源头货车数'] / (x['数据站点总数'] + 0.0000001), axis=1).round(0)
        附件7地市['设备上线率（%）'] = 附件7地市.apply(lambda x: x['在线站点数'] / (x['数据站点总数'] + 0.0000001), axis=1)
        附件7地市['20-50%占比'] = 附件7地市.apply(lambda x: x['20-50%数'] / (x['源头货车数'] + 0.0000001), axis=1)
        附件7地市['50-100%占比'] = 附件7地市.apply(lambda x: x['50-100%数'] / (x['源头货车数'] + 0.0000001), axis=1)
        附件7地市['100%以上占比'] = 附件7地市.apply(lambda x: x['100%以上数'] / (x['源头货车数'] + 0.0000001), axis=1)
        附件7地市 = 附件7地市.fillna(0, inplace=False)
        附件7地市['设备上线率（%）'] = 附件7地市['设备上线率（%）'].apply(lambda x: format(x, '.2%'))
        附件7地市['20-50%占比'] = 附件7地市['20-50%占比'].apply(lambda x: format(x, '.2%'))
        附件7地市['50-100%占比'] = 附件7地市['50-100%占比'].apply(lambda x: format(x, '.2%'))
        附件7地市['100%以上占比'] = 附件7地市['100%以上占比'].apply(lambda x: format(x, '.2%'))
        附件7地市['80吨以上总数'] = 附件7地市['本月超限80吨以上']
        附件7地市['外省入库数'] = 附件7地市['入库数(路政)'] - 附件7地市['本省入库数']
        附件7地市['非现场处罚外省(路政)'] = 附件7地市['非现场处罚(路政)'] - 附件7地市['非现场处罚本省(路政)']
        附件7地市['区县'] = 附件7地市['地市']
        附件7地市['区县编码'] = 附件7地市['地市编码']
        附件7地市['统计月份'] = 附件7['统计月份']
        # 附件7地市 = 附件7地市.set_index()
        附件7地市 = 附件7地市.fillna(0, inplace=False)
        附件7地市 = pd.DataFrame(附件7地市,
                             columns=['统计月份', '地市编码', '地市', '区县编码', '区县', '货车数', '超限数', '超限率', '超限100', '超限100%数货车数/万辆',
                                      '本月超限80吨以上', '本月超限90吨以上', '百吨王数', '超限100%遮挡车牌数量（辆）', '百吨王遮挡车牌数量（辆）', '站点数',
                                      '报修数', '设备完好率', '异常数', '数据完好率', '入库数(路政)', '本省入库数', '外省入库数', '现场处罚(路政)',
                                      '非现场处罚(路政)',
                                      '非现场处罚本省(路政)', '非现场处罚(路政)当年',
                                      '非现场处罚本省(路政)当年', '非现场处罚外省(路政)', '交警现场查处数', '交警非现场处罚数', '交警非现查处数本省',
                                      '需处罚数/非现入库数（总计）',
                                      '需处罚数/非现入库数（本省）',
                                      '需处罚数/非现入库数（外省）', '非现处罚数（总计）', '非现处罚数（本省）', '非现处罚数（外省）', '年处罚数', '外省抄告',
                                      '非现场处罚率（本省）',
                                      '非现场处罚率（外省）',
                                      '处罚率(含抄告）', '80吨以上总数', '80吨以上且满足', '80吨以上且审核通过', '合规率', '后面是源头数据', '数据站点总数',
                                      '源头货车数',
                                      '源头单位平均过车数（辆次）',
                                      '在线站点数', '设备上线率（%）', '20-50%数', '50-100%数', '100%以上数', '20-50%占比', '50-100%占比',
                                      '100%以上占比', '非现场处罚路政案发', '非现场处罚本省路政案发', '非现场处罚路政案发当年', '非现场处罚本省路政案发当年'])
        ##案件处罚率等处理
        附件7地市 = 附件7地市.fillna(0, inplace=False)
        附件7地市['交警非现场处罚数'] = 附件7地市['交警非现场处罚数'].astype('int')
        附件7地市['非现场处罚本省路政案发当年'] = 附件7地市['非现场处罚本省路政案发当年'].astype('int')
        附件7地市['非现场处罚本省路政案发当年'] = 附件7地市['非现场处罚本省路政案发当年'].astype('int')
        # 附件7地市['交警非现场处罚数'] = 附件7地市['交警非现场处罚数'].astype('int')

        附件7地市['需处罚数/非现入库数（总计）'] = 附件7地市['入库数(路政)'] + 附件7地市['交警非现场处罚数']
        附件7地市['需处罚数/非现入库数（本省）'] = 附件7地市['本省入库数'] + 附件7地市['交警非现场处罚数']
        附件7地市['需处罚数/非现入库数（外省）'] = 附件7地市['外省入库数']
        附件7地市['非现处罚数（总计）'] = 附件7地市['非现场处罚(路政)'] + 附件7地市['交警非现场处罚数']
        附件7地市['非现处罚数（本省）'] = 附件7地市['非现场处罚本省(路政)'] + 附件7地市['交警非现场处罚数']
        附件7地市['非现处罚数（外省）'] = 附件7地市['非现场处罚(路政)'] - 附件7地市['非现场处罚本省(路政)']
        附件7地市['年处罚数'] = 附件7地市['非现场处罚(路政)'] + 附件7地市['交警非现场处罚数'] + 附件7地市['现场处罚(路政)'] + 附件7地市['交警现场查处数']

        附件7地市['非现场处罚率（本省）'] = 附件7地市.apply(
            lambda x: (x['非现场处罚本省路政案发当年'] + x['交警非现场处罚数']) / (x['本省入库数'] + x['交警非现场处罚数'] + 0.00001), axis=1).round(4)
        附件7地市['非现场处罚率（外省）'] = 附件7地市.apply(
            lambda x: (x['非现场处罚路政案发当年'] - x['非现场处罚本省路政案发当年']) / (x['需处罚数/非现入库数（外省）'] + 0.00001), axis=1).round(4)
        附件7地市['处罚率(含抄告）'] = 附件7地市.apply(
            lambda x: (x['非现场处罚路政案发当年'] + x['交警非现场处罚数'] + x['外省抄告']) / (x['入库数(路政)'] + x['交警非现场处罚数'] + 0.00001),
            axis=1).round(4)
        附件7地市['合规率'] = 附件7地市.apply(
            lambda x: (x['80吨以上且审核通过']) / (x['80吨以上总数'] + 0.00001),
            axis=1).round(4)
        附件7地市.loc[附件7地市['非现场处罚率（本省）'] > 1, '非现场处罚率（本省）'] = 1
        附件7地市.loc[附件7地市['非现场处罚率（外省）'] > 1, '非现场处罚率（外省）'] = 1
        附件7地市.loc[附件7地市['处罚率(含抄告）'] > 1, '处罚率(含抄告）'] = 1
        附件7地市.loc[附件7地市['合规率'] > 1, '合规率'] = 1

        附件7地市 = 附件7地市.fillna(0, inplace=False)
        附件7地市['非现场处罚率（本省）'] = 附件7地市['非现场处罚率（本省）'].apply(lambda x: format(x, '.2%'))
        附件7地市['非现场处罚率（外省）'] = 附件7地市['非现场处罚率（外省）'].apply(lambda x: format(x, '.2%'))
        附件7地市['处罚率(含抄告）'] = 附件7地市['处罚率(含抄告）'].apply(lambda x: format(x, '.2%'))
        附件7地市['合规率'] = 附件7地市['合规率'].apply(lambda x: format(x, '.2%'))
        附件7地市['区县编码'] = 附件7地市['区县编码'].astype('str')
        附件7地市 = 附件7地市.sort_values(by=['区县编码'],
                                  ascending=True).reset_index(drop=True)
        附件7地市 = 附件7地市.fillna(0, inplace=False)
        附件7地市省 = 附件7地市.groupby(['统计月份']).sum().reset_index()
        附件7地市省.loc[附件7地市省['货车数'] == 0, '货车数'] = 1
        附件7地市省['超限率'] = 附件7地市省.apply(lambda x: x['超限数'] / x['货车数'], axis=1).round(4)
        附件7地市省 = 附件7地市省.fillna(0, inplace=False)
        附件7地市省['超限率'] = 附件7地市省['超限率'].apply(lambda x: format(x, '.2%'))
        附件7地市省.loc[附件7地市省['货车数'] == 1, '货车数'] = 0
        附件7地市省['设备完好率'] = (附件7地市省['站点数'] - 附件7地市省['报修数']) / 附件7地市省['站点数']
        附件7地市省['数据完好率'] = (附件7地市省['站点数'] - 附件7地市省['异常数']) / 附件7地市省['站点数']
        附件7地市省 = 附件7地市省.fillna(0, inplace=False)
        附件7地市省['数据完好率'] = 附件7地市省['数据完好率'].apply(lambda x: format(x, '.2%'))
        附件7地市省['设备完好率'] = 附件7地市省['设备完好率'].apply(lambda x: format(x, '.2%'))
        附件7地市省['超限100%数货车数/万辆'] = 附件7地市省.apply(lambda x: x['超限100'] / (x['货车数'] + 0.0000001) * 10000, axis=1).round(4)
        附件7地市省['源头货车数'] = pd.to_numeric(附件7地市省['源头货车数'], errors='coerce')
        附件7地市省['数据站点总数'] = pd.to_numeric(附件7地市省['数据站点总数'], errors='coerce')
        附件7地市省['在线站点数'] = pd.to_numeric(附件7地市省['在线站点数'], errors='coerce')
        附件7地市省['20-50%数'] = pd.to_numeric(附件7地市省['20-50%数'], errors='coerce')
        附件7地市省['50-100%数'] = pd.to_numeric(附件7地市省['50-100%数'], errors='coerce')
        附件7地市省['100%以上数'] = pd.to_numeric(附件7地市省['100%以上数'], errors='coerce')
        附件7地市省['源头单位平均过车数（辆次）'] = 附件7地市省.apply(lambda x: x['源头货车数'] / (x['数据站点总数'] + 0.0000001), axis=1).round(0)
        附件7地市省['设备上线率（%）'] = 附件7地市省.apply(lambda x: x['在线站点数'] / (x['数据站点总数'] + 0.0000001), axis=1)
        附件7地市省['20-50%占比'] = 附件7地市省.apply(lambda x: x['20-50%数'] / (x['源头货车数'] + 0.0000001), axis=1)
        附件7地市省['50-100%占比'] = 附件7地市省.apply(lambda x: x['50-100%数'] / (x['源头货车数'] + 0.0000001), axis=1)
        附件7地市省['100%以上占比'] = 附件7地市省.apply(lambda x: x['100%以上数'] / (x['源头货车数'] + 0.0000001), axis=1)
        附件7地市省 = 附件7地市省.fillna(0, inplace=False)
        附件7地市省['设备上线率（%）'] = 附件7地市省['设备上线率（%）'].apply(lambda x: format(x, '.2%'))
        附件7地市省['20-50%占比'] = 附件7地市省['20-50%占比'].apply(lambda x: format(x, '.2%'))
        附件7地市省['50-100%占比'] = 附件7地市省['50-100%占比'].apply(lambda x: format(x, '.2%'))
        附件7地市省['100%以上占比'] = 附件7地市省['100%以上占比'].apply(lambda x: format(x, '.2%'))
        附件7地市省['80吨以上总数'] = 附件7地市省['本月超限80吨以上']
        附件7地市省['外省入库数'] = 附件7地市省['入库数(路政)'] - 附件7地市省['本省入库数']
        附件7地市省['非现场处罚外省(路政)'] = 附件7地市省['非现场处罚(路政)'] - 附件7地市省['非现场处罚本省(路政)']
        附件7地市省['地市编码'] = '330000'
        附件7地市省['地市'] = '浙江'
        附件7地市省['区县'] = 附件7地市省['地市']
        附件7地市省['区县编码'] = 附件7地市省['地市编码']
        附件7地市省['统计月份'] = 附件7地市['统计月份']
        # 附件7地市省 = 附件7地市省.set_index()
        附件7地市省 = 附件7地市省.fillna(0, inplace=False)
        附件7地市省 = pd.DataFrame(附件7地市省,
                              columns=['统计月份', '地市编码', '地市', '区县编码', '区县', '货车数', '超限数', '超限率', '超限100',
                                       '超限100%数货车数/万辆',
                                       '本月超限80吨以上', '本月超限90吨以上', '百吨王数', '超限100%遮挡车牌数量（辆）', '百吨王遮挡车牌数量（辆）', '站点数',
                                       '报修数', '设备完好率', '异常数', '数据完好率', '入库数(路政)', '本省入库数', '外省入库数', '现场处罚(路政)',
                                       '非现场处罚(路政)',
                                       '非现场处罚本省(路政)', '非现场处罚(路政)当年',
                                       '非现场处罚本省(路政)当年', '非现场处罚外省(路政)', '交警现场查处数', '交警非现场处罚数', '交警非现查处数本省',
                                       '需处罚数/非现入库数（总计）',
                                       '需处罚数/非现入库数（本省）',
                                       '需处罚数/非现入库数（外省）', '非现处罚数（总计）', '非现处罚数（本省）', '非现处罚数（外省）', '年处罚数', '外省抄告',
                                       '非现场处罚率（本省）', '非现场处罚率（外省）',
                                       '处罚率(含抄告）', '80吨以上总数', '80吨以上且满足', '80吨以上且审核通过', '合规率', '后面是源头数据', '数据站点总数',
                                       '源头货车数',
                                       '源头单位平均过车数（辆次）',
                                       '在线站点数', '设备上线率（%）', '20-50%数', '50-100%数', '100%以上数', '20-50%占比', '50-100%占比',
                                       '100%以上占比', '非现场处罚路政案发', '非现场处罚本省路政案发', '非现场处罚路政案发当年', '非现场处罚本省路政案发当年'])
        ##案件处罚率等处理
        附件7地市省 = 附件7地市省.fillna(0, inplace=False)
        附件7地市省['交警非现场处罚数'] = 附件7地市省['交警非现场处罚数'].astype('int')
        附件7地市省['非现场处罚本省路政案发当年'] = 附件7地市省['非现场处罚本省路政案发当年'].astype('int')
        附件7地市省['非现场处罚本省路政案发当年'] = 附件7地市省['非现场处罚本省路政案发当年'].astype('int')
        # 附件7地市省['交警非现场处罚数'] = 附件7地市省['交警非现场处罚数'].astype('int')

        附件7地市省['需处罚数/非现入库数（总计）'] = 附件7地市省['入库数(路政)'] + 附件7地市省['交警非现场处罚数']
        附件7地市省['需处罚数/非现入库数（本省）'] = 附件7地市省['本省入库数'] + 附件7地市省['交警非现场处罚数']
        附件7地市省['需处罚数/非现入库数（外省）'] = 附件7地市省['外省入库数']
        附件7地市省['非现处罚数（总计）'] = 附件7地市省['非现场处罚(路政)'] + 附件7地市省['交警非现场处罚数']
        附件7地市省['非现处罚数（本省）'] = 附件7地市省['非现场处罚本省(路政)'] + 附件7地市省['交警非现场处罚数']
        附件7地市省['非现处罚数（外省）'] = 附件7地市省['非现场处罚(路政)'] - 附件7地市省['非现场处罚本省(路政)']
        附件7地市省['年处罚数'] = 附件7地市省['非现场处罚(路政)'] + 附件7地市省['交警非现场处罚数'] + 附件7地市省['现场处罚(路政)'] + 附件7地市省['交警现场查处数']

        附件7地市省['非现场处罚率（本省）'] = 附件7地市省.apply(
            lambda x: (x['非现场处罚本省路政案发当年'] + x['交警非现场处罚数']) / (x['本省入库数'] + x['交警非现场处罚数'] + 0.00001), axis=1).round(4)
        附件7地市省['非现场处罚率（外省）'] = 附件7地市省.apply(
            lambda x: (x['非现场处罚路政案发当年'] - x['非现场处罚本省路政案发当年']) / (x['需处罚数/非现入库数（外省）'] + 0.00001), axis=1).round(4)
        附件7地市省['处罚率(含抄告）'] = 附件7地市省.apply(
            lambda x: (x['非现场处罚路政案发当年'] + x['交警非现场处罚数'] + x['外省抄告']) / (x['入库数(路政)'] + x['交警非现场处罚数'] + 0.00001),
            axis=1).round(4)
        附件7地市省['合规率'] = 附件7地市省.apply(
            lambda x: (x['80吨以上且审核通过']) / (x['80吨以上总数'] + 0.00001),
            axis=1).round(4)
        附件7地市省.loc[附件7地市省['非现场处罚率（本省）'] > 1, '非现场处罚率（本省）'] = 1
        附件7地市省.loc[附件7地市省['非现场处罚率（外省）'] > 1, '非现场处罚率（外省）'] = 1
        附件7地市省.loc[附件7地市省['处罚率(含抄告）'] > 1, '处罚率(含抄告）'] = 1
        附件7地市省.loc[附件7地市省['合规率'] > 1, '合规率'] = 1

        附件7地市省 = 附件7地市省.fillna(0, inplace=False)
        附件7地市省['非现场处罚率（本省）'] = 附件7地市省['非现场处罚率（本省）'].apply(lambda x: format(x, '.2%'))
        附件7地市省['非现场处罚率（外省）'] = 附件7地市省['非现场处罚率（外省）'].apply(lambda x: format(x, '.2%'))
        附件7地市省['处罚率(含抄告）'] = 附件7地市省['处罚率(含抄告）'].apply(lambda x: format(x, '.2%'))
        附件7地市省['合规率'] = 附件7地市省['合规率'].apply(lambda x: format(x, '.2%'))

        附件7地市省 = 附件7地市省.fillna(0, inplace=False)
        汇总报表 = pd.concat([附件7, 附件7地市, 附件7地市省])
        汇总报表 = pd.DataFrame(汇总报表,
                            columns=['统计月份', '地市编码', '地市', '区县编码', '区县', '货车数', '超限数', '超限率', '超限100', '超限100%数货车数/万辆',
                                     '本月超限80吨以上', '本月超限90吨以上', '百吨王数', '超限100%遮挡车牌数量（辆）', '百吨王遮挡车牌数量（辆）', '站点数',
                                     '报修数', '设备完好率', '异常数', '数据完好率', '入库数(路政)', '本省入库数', '外省入库数', '非现场处罚(路政)',
                                     '非现场处罚本省(路政)', '非现场处罚(路政)当年',
                                     '非现场处罚本省(路政)当年', '非现场处罚外省(路政)', '交警现场查处数', '需处罚数/非现入库数（总计）',
                                     '需处罚数/非现入库数（本省）',
                                     '需处罚数/非现入库数（外省）', '非现处罚数（总计）', '非现处罚数（本省）', '非现处罚数（外省）', '年处罚数', '外省抄告',
                                     '非现场处罚率（本省）', '非现场处罚率（外省）',
                                     '处罚率(含抄告）', '80吨以上总数', '80吨以上且满足', '80吨以上且审核通过', '合规率', '后面是源头数据', '数据站点总数',
                                     '源头货车数',
                                     '源头单位平均过车数（辆次）',
                                     '在线站点数', '设备上线率（%）', '20-50%数', '50-100%数', '100%以上数', '20-50%占比', '50-100%占比',
                                     '100%以上占比', '非现场处罚路政案发', '非现场处罚本省路政案发', '非现场处罚路政案发当年', '非现场处罚本省路政案发当年'])
        case_compliance = pd.read_excel(self.file_name, sheet_name='case_compliance')
        case_compliance['区县编码'] = case_compliance['区县编码'].astype('str')
        汇总报表 = pd.merge(汇总报表, case_compliance, on=['区县编码'], how='left')
        汇总报表=汇总报表.fillna(0)
        汇总报表['超限80吨以上总数1'] = 汇总报表['超限80吨以上总数1'].astype('int')
        汇总报表['超限80吨以上且满足处罚条件总数1'] = 汇总报表['超限80吨以上且满足处罚条件总数1'].astype('int')
        汇总报表['月超限80吨以上审核通过总数1'] = 汇总报表['月超限80吨以上审核通过总数1'].astype('int')
        汇总报表['80吨以上总数'] = 汇总报表['超限80吨以上总数1'] + 汇总报表['80吨以上总数']
        汇总报表['80吨以上且满足'] = 汇总报表['超限80吨以上且满足处罚条件总数1'] + 汇总报表['80吨以上且满足']
        汇总报表['80吨以上且审核通过'] = 汇总报表['月超限80吨以上审核通过总数1'] + 汇总报表['80吨以上且审核通过']

        ##案件处罚率等处理
        汇总报表 = 汇总报表.fillna(0, inplace=False)
        汇总报表['报废车辆数'] = 汇总报表['报废车辆数'].astype('int')
        汇总报表['入库数(路政)'] = 汇总报表['入库数(路政)']-汇总报表['报废车辆数']
        汇总报表['交警非现场处罚数'] = 汇总报表['交警非现场处罚数'].astype('int')
        汇总报表['非现场处罚本省路政案发当年'] = 汇总报表['非现场处罚本省路政案发当年'].astype('int')
        汇总报表['非现场处罚本省路政案发当年'] = 汇总报表['非现场处罚本省路政案发当年'].astype('int')
        # 汇总报表['交警非现场处罚数'] = 汇总报表['交警非现场处罚数'].astype('int')

        汇总报表['需处罚数/非现入库数（总计）'] = 汇总报表['入库数(路政)'] + 汇总报表['交警非现场处罚数']
        汇总报表['需处罚数/非现入库数（本省）'] = 汇总报表['本省入库数'] + 汇总报表['交警非现场处罚数']
        汇总报表['需处罚数/非现入库数（外省）'] = 汇总报表['外省入库数']
        汇总报表['非现处罚数（总计）'] = 汇总报表['非现场处罚(路政)'] + 汇总报表['交警非现场处罚数']
        汇总报表['非现处罚数（本省）'] = 汇总报表['非现场处罚本省(路政)'] + 汇总报表['交警非现场处罚数']
        汇总报表['非现处罚数（外省）'] = 汇总报表['非现场处罚(路政)'] - 汇总报表['非现场处罚本省(路政)']
        汇总报表['年处罚数'] = 汇总报表['非现场处罚(路政)'] + 汇总报表['交警非现场处罚数'] + 汇总报表['现场处罚(路政)'] + 汇总报表['交警现场查处数']

        汇总报表['非现场处罚率（本省）'] = 汇总报表.apply(
            lambda x: (x['非现场处罚本省路政案发'] + x['交警非现场处罚数']) / (x['本省入库数'] + x['交警非现场处罚数']+x['非现场处罚本省路政案发'] -x['非现场处罚本省路政案发当年']  + 0.00001), axis=1).round(4)
        汇总报表['非现场处罚率（外省）'] = 汇总报表.apply(
            lambda x: (x['非现场处罚路政案发当年'] - x['非现场处罚本省路政案发当年']) / (x['需处罚数/非现入库数（外省）'] + 0.00001), axis=1).round(4)
        汇总报表['处罚率(含抄告）'] = 汇总报表.apply(
            lambda x: (x['非现场处罚路政案发']+ x['交警非现场处罚数'] + x['外省抄告']) / (x['入库数(路政)'] + x['交警非现场处罚数']+x['非现场处罚路政案发'] -x['非现场处罚路政案发当年']  + 0.00001),
            axis=1).round(4)
        汇总报表['合规率'] = 汇总报表.apply(
            lambda x: (x['80吨以上且审核通过']) / (x['80吨以上总数'] + 0.00001),
            axis=1).round(4)
        汇总报表.loc[汇总报表['非现场处罚率（本省）'] > 1, '非现场处罚率（本省）'] = 1
        汇总报表.loc[汇总报表['非现场处罚率（外省）'] > 1, '非现场处罚率（外省）'] = 1
        汇总报表.loc[汇总报表['处罚率(含抄告）'] > 1, '处罚率(含抄告）'] = 1
        汇总报表.loc[汇总报表['合规率'] > 1, '合规率'] = 1
        汇总报表 = 汇总报表.fillna(0, inplace=False)
        汇总报表['非现场处罚率（本省）'] = 汇总报表['非现场处罚率（本省）'].apply(lambda x: format(x, '.2%'))
        汇总报表['非现场处罚率（外省）'] = 汇总报表['非现场处罚率（外省）'].apply(lambda x: format(x, '.2%'))
        汇总报表['处罚率(含抄告）'] = 汇总报表['处罚率(含抄告）'].apply(lambda x: format(x, '.2%'))
        汇总报表['合规率'] = 汇总报表['合规率'].apply(lambda x: format(x, '.2%'))

        汇总报表 = 汇总报表.fillna(0, inplace=False)
        汇总报表 = pd.DataFrame(汇总报表,
                            columns=['统计月份', '地市编码', '地市', '区县编码', '区县', '货车数', '超限数', '超限率', '超限100', '超限100%数货车数/万辆',
                                     '本月超限80吨以上', '本月超限90吨以上', '百吨王数', '超限100%遮挡车牌数量（辆）', '百吨王遮挡车牌数量（辆）', '站点数',
                                     '报修数', '设备完好率', '异常数', '数据完好率', '入库数(路政)', '本省入库数', '外省入库数', '现场处罚(路政)',
                                     '非现场处罚(路政)',
                                     '非现场处罚本省(路政)', '非现场处罚(路政)当年',
                                     '非现场处罚本省(路政)当年', '非现场处罚外省(路政)', '交警现场查处数', '交警非现场处罚数', '交警非现查处数本省',
                                     '需处罚数/非现入库数（总计）',
                                     '需处罚数/非现入库数（本省）',
                                     '需处罚数/非现入库数（外省）', '非现处罚数（总计）', '非现处罚数（本省）', '非现处罚数（外省）', '年处罚数', '外省抄告',
                                     '非现场处罚率（本省）', '非现场处罚率（外省）',
                                     '处罚率(含抄告）', '80吨以上总数', '80吨以上且满足', '80吨以上且审核通过', '合规率', '后面是源头数据', '数据站点总数',
                                     '源头货车数',
                                     '源头单位平均过车数（辆次）',
                                     '在线站点数', '设备上线率（%）', '20-50%数', '50-100%数', '100%以上数', '20-50%占比', '50-100%占比',
                                     '100%以上占比', '非现场处罚路政案发', '非现场处罚本省路政案发', '非现场处罚路政案发当年', '非现场处罚本省路政案发当年'])
        # df_站点在线明细 = self.data_station_source()
        df_站点在线明细 = pd.read_excel(r'{}\data_station_source.xlsx'.format(self.out_file_月报), sheet_name='df_站点在线明细')
        with pd.ExcelWriter(self.out_file_name) as writer1:
            汇总报表.to_excel(writer1, sheet_name='汇总', index=False)
            附件7.to_excel(writer1, sheet_name='区县汇总', index=False)
            附件7地市.to_excel(writer1, sheet_name='地市汇总', index=False)
            附件7地市省.to_excel(writer1, sheet_name='省汇总', index=False)
            货运源头监控数据省市.to_excel(writer1, sheet_name='货运源头监控数据省市', index=False)
            df_数据汇总.to_excel(writer1, sheet_name='站点数据明细', index=False)
            df_站点在线明细.to_excel(writer1, sheet_name='站点在线明细', index=False)
            df_报修点位统计.to_excel(writer1, sheet_name='报修', index=False)
            df_sheet1.to_excel(writer1, sheet_name='sheet1', index=False)
            站点设备完好率.to_excel(writer1, sheet_name='站点设备完好率', index=False)
            区县超限率排序.to_excel(writer1, sheet_name='区县超限率排序', index=False)
            T_20天与500筛选1.to_excel(writer1, sheet_name='在线天数大于等于20天货车数大于500的站点数据', index=False)
            df_源头站点明细.to_excel(writer1, sheet_name='源头站点数据明细', index=False)
            总重80吨以上明细.to_excel(writer1, sheet_name='总重80_90以上明细', index=False)
            超限100汇总.to_excel(writer1, sheet_name='超限100汇总', index=False)
            超限100明细.to_excel(writer1, sheet_name='超限100明细', index=False)
            百吨王明细.to_excel(writer1, sheet_name='百吨王明细', index=False)

    def creat_charts(self,df,city):

        df['检测时间'] = pd.to_datetime(df['检测时间'])
        df['检测时间日'] = df['检测时间'].map(lambda x: x.strftime('%d'))
        df['检测时间时'] = df['检测时间'].map(lambda x: x.strftime('%H'))
        df['案件状态'] = df['案件状态'].fillna('99')
        df['案件状态'] = df['案件状态'].astype('int')
        df.loc[df['案件状态'] == 99, '案件状态'] = '未采集'
        df.loc[df['案件状态'] == 0, '案件状态'] = '审核不通过'
        df.loc[df['案件状态'] == 1, '案件状态'] = '待初审'
        df.loc[df['案件状态'] == 2, '案件状态'] = '待审核'
        df.loc[df['案件状态'] == 3, '案件状态'] = '待判定'
        df.loc[df['案件状态'] == 4, '案件状态'] = '已告知'
        df.loc[df['案件状态'] == 5, '案件状态'] = '免处理'
        df.loc[df['案件状态'] == 6, '案件状态'] = '已立案'
        df.loc[df['案件状态'] == 12, '案件状态'] = '待告知'
        df.loc[df['案件状态'] == 13, '案件状态'] = '已结案'
        df.loc[df['案件状态'] == 9, '案件状态'] = '判定不处理'
        df.loc[df['案件状态'] == 15, '案件状态'] = '初审不通过'
        df['判定需处罚'] = df['判定需处罚'].fillna('2')
        df['判定需处罚'] = df['判定需处罚'].astype('int')
        df.loc[df['判定需处罚'] == 2, '判定需处罚'] = '未采集'
        df.loc[df['判定需处罚'] == 99, '判定需处罚'] = '初审'
        df.loc[df['判定需处罚'] == 1, '判定需处罚'] = '复审'
        cor = df['案件状态'].value_counts()
        # cor = cor[0:5]
        x = range(len(cor.index))
        name = cor.index
        y = cor.values
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        plt.figure(figsize=(12, 7))
        # 计算各店铺的商品数量
        cor.plot.bar(width=0.8, alpha=0.6, color='r')
        plt.title('80吨以上(案件状态)', fontsize=18)
        plt.ylabel('案件数', fontsize=14)
        for x, y in zip(x, y):
            plt.text(x, y + 0.1, '%d' % y, ha='center', va='bottom')
        plt.savefig(r'{}\{}80吨以上(案件状态).png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

        city_month = df['判定需处罚'].value_counts()
        plt.figure(figsize=(21, 12), dpi=80)
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        plt.subplot(1, 1, 1)
        plt.pie(city_month, labels=city_month.index, autopct='%.2f%%')
        # plt.legend()
        plt.title('审核状态', fontsize=18)
        plt.savefig(r'{}\{}审核状态.png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

        city_month = df[df['判定需处罚'] == '初审']['案件状态'].value_counts()
        plt.figure(figsize=(21, 12), dpi=80)
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        plt.subplot(1, 1, 1)
        plt.pie(city_month, labels=city_month.index, autopct='%.2f%%')
        # plt.legend()
        plt.title('初审占比', fontsize=18)
        plt.savefig(r'{}\{}初审占比.png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

        city_month = df[df['判定需处罚'] == '复审']['案件状态'].value_counts()
        plt.figure(figsize=(21, 12), dpi=80)
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        plt.subplot(1, 1, 1)
        plt.pie(city_month, labels=city_month.index, autopct='%.2f%%')
        # plt.legend()
        plt.title('复审占比', fontsize=18)
        plt.savefig(r'{}\{}复审占比.png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

        city_month = df[df['案件状态'] == '初审不通过']['审核意见'].value_counts().sort_values(ascending=False)
        plt.figure(figsize=(21, 12), dpi=80)
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        plt.subplot(1, 1, 1)
        plt.pie(city_month, labels=city_month.index, autopct='%.2f%%')
        # plt.legend()
        plt.title('审核意见(初始不通过)', fontsize=18)
        plt.savefig(r'{}\{}审核意见(初始不通过).png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

        # city_month=df['检测时间日'].value_counts().sort_index()
        city_month = df['检测时间日'].value_counts().sort_index()
        name = city_month.index
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        x = list(range(len(city_month.index)))
        y = city_month.values
        width = 0.3
        plt.figure(figsize=(15, 8), dpi=80)
        plt.plot(x, y, color='lightsalmon')
        plt.xticks(x, name)
        # plt.legend()
        plt.xlabel('天', fontsize=14)
        plt.ylabel('辆次', fontsize=14)
        plt.title('80吨以上(天）', fontsize=18)
        for x, y in zip(x, y):
            plt.text(x, y + 0.1, '%d' % y, ha='center', va='bottom')
        plt.savefig(r'{}\{}80吨以上(天）.png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

        # plt.xlabel(u'日', fontproperties=font) # 添加中文字符串
        city_month = df['检测时间时'].value_counts()
        city_month = city_month[0:12]
        name = city_month.index
        x = list(range(len(city_month.index)))
        y = city_month.values
        width = 0.3
        plt.figure(figsize=(15, 8), dpi=80)
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        plt.plot(x, y, color='lightsalmon')
        plt.xticks(x, name)
        # plt.legend()
        plt.xlabel('小时', fontsize=14)
        plt.ylabel('辆次', fontsize=14)
        plt.title('80吨以上(小时)', fontsize=18)
        for x, y in zip(x, y):
            plt.text(x, y + 0.1, '%d' % y, ha='center', va='bottom')
        plt.savefig(r'{}\{}80吨以上(小时).png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

        cor = df['站点名称'].value_counts().sort_values(ascending=False)
        cor = cor[0:10]
        cor = cor.sort_values(ascending=True)
        plt.figure(figsize=(8, 15))
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        x = range(len(cor.index))
        name = cor.index
        y = cor.values
        plt.barh(x, y, color='salmon')
        plt.yticks(x, name, fontsize=14)
        for x, y in zip(x, y):
            plt.text(y, x - 0.1, '%.2f' % y, ha='center', va='bottom')
        # plt.ylabel('站点名称', fontproperties=font)
        plt.title('80吨以上(前十站点)', fontsize=18)
        plt.savefig(r'{}\{}80吨以上(前十站点).png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

        cor = df['区县'].value_counts().sort_values(ascending=False)
        # cor = cor[0:5]
        x = range(len(cor.index))
        name = cor.index
        y = cor.values
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        plt.figure(figsize=(12, 7))
        # 计算各店铺的商品数量
        cor.plot.bar(width=0.8, alpha=0.6, color='r')
        plt.title('80吨以上(区县排名)', fontsize=18)
        plt.ylabel('辆次', fontsize=14)
        for x, y in zip(x, y):
            plt.text(x, y + 0.1, '%d' % y, ha='center', va='bottom')
        plt.savefig(r'{}\{}80吨以上(区县排名).png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

        cor = df[(~((df['车牌号码'].str.contains("牌") == True))) & (~((df['车牌号码'].str.contains("-") == True))) & (
            ~((df['车牌号码'].str.contains("未") == True)))]['车牌号码'].value_counts().sort_values(ascending=False)
        cor = cor[0:10]
        cor = cor.sort_values(ascending=True)
        plt.figure(figsize=(8, 15))
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        x = range(len(cor.index))
        name = cor.index
        y = cor.values
        plt.barh(x, y, color='salmon')
        plt.yticks(x, name, fontsize=14)
        for x, y in zip(x, y):
            plt.text(y, x - 0.1, '%.2f' % y, ha='center', va='bottom')
        # plt.ylabel('站点名称', fontproperties=font)
        plt.title('80吨以上(前十车牌)', fontsize=18)
        plt.savefig(r'{}\{}80吨以上(前十车牌).png'.format(self.charts_path, city), bbox_inches='tight')
        plt.close()

    def add_charts(self,city,city_path):
        workbook = load_workbook(filename=city_path)
        sheet = workbook.create_sheet('80吨以上数据分析')
        # sheet = workbook.active
        image_indices = list(range(len(sheet._images)))
        # 倒序遍历索引列表并删除图片对象
        for index in reversed(image_indices):
            del sheet._images[index]
        logo = Image(r'{}\{}80吨以上(天）.png'.format(self.charts_path, city))
        logo.height = 600
        logo.width = 1050
        sheet.add_image(logo, "A1")
        logo = Image(r'{}\{}80吨以上(小时).png'.format(self.charts_path, city))
        logo.height = 600
        logo.width = 1050
        sheet.add_image(logo, "Q1")

        logo = Image(r'{}\{}80吨以上(区县排名).png'.format(self.charts_path, city))
        logo.height = 600
        logo.width = 1050
        sheet.add_image(logo, "A33")
        logo = Image(r'{}\{}80吨以上(案件状态).png'.format(self.charts_path, city))
        logo.height = 600
        logo.width = 1050
        sheet.add_image(logo, "Q33")
        logo = Image(r'{}\{}审核状态.png'.format(self.charts_path, city))
        logo.height = 900
        logo.width = 900
        sheet.add_image(logo, "B66")
        logo = Image(r'{}\{}初审占比.png'.format(self.charts_path, city))
        logo.height = 900
        logo.width = 900
        sheet.add_image(logo, "Q66")
        logo = Image(r'{}\{}复审占比.png'.format(self.charts_path, city))
        logo.height = 900
        logo.width = 900
        sheet.add_image(logo, "B115")
        logo = Image(r'{}\{}审核意见(初始不通过).png'.format(self.charts_path, city))
        logo.height = 1000
        logo.width = 1100
        sheet.add_image(logo, "Q110")
        logo = Image(r'{}\{}80吨以上(前十站点).png'.format(self.charts_path, city))
        logo.height = 1000
        logo.width = 900
        sheet.add_image(logo, "A163")
        logo = Image(r'{}\{}80吨以上(前十车牌).png'.format(self.charts_path, city))
        logo.height = 1000
        logo.width = 700
        sheet.add_image(logo, "Q163")
        workbook.save(filename=city_path)

    ##地市分开
    def city_excel(self):
        excel_file = pd.read_excel(self.out_file_name, sheet_name=None)
        # 打印所有 sheet 的名称
        # print(excel_file.keys())
        # 读取指定 sheet 的数据
        汇总 = excel_file['汇总']
        汇总 = pd.DataFrame(汇总,
                          columns=['地市', '区县', '货车数', '超限数', '超限率', '本月超限80吨以上', '本月超限90吨以上', '超限100', '超限100%数货车数/万辆',
                                   '站点数',
                                   '报修数', '设备完好率', '异常数', '数据完好率', '百吨王数', '超限100%遮挡车牌数量（辆）', '百吨王遮挡车牌数量（辆）', '入库数(路政)',
                                   '本省入库数', '外省入库数', '现场处罚(路政)', '非现场处罚(路政)', '非现场处罚本省(路政)', '非现场处罚(路政)当年',
                                   '非现场处罚本省(路政)当年', '非现场处罚外省(路政)', '交警现场查处数', '交警非现场处罚数', '交警非现查处数本省',
                                   '需处罚数/非现入库数（总计）', '需处罚数/非现入库数（本省）', '需处罚数/非现入库数（外省）', '非现处罚数（总计）',
                                   '非现处罚数（本省）', '非现处罚数（外省）', '年处罚数', '外省抄告', '非现场处罚率（本省）', '非现场处罚率（外省）',
                                   '处罚率(含抄告）', '80吨以上总数', '80吨以上且满足', '80吨以上且审核通过', '合规率', '后面是源头数据', '数据站点总数',
                                   '源头货车数', '源头单位平均过车数（辆次）', '在线站点数', '设备上线率（%）', '20-50%数', '50-100%数', '100%以上数',
                                   '20-50%占比', '50-100%占比',
                                   '100%以上占比', '非现场处罚路政案发', '非现场处罚本省路政案发', '非现场处罚路政案发当年', '非现场处罚本省路政案发当年'])
        汇总 = 汇总[~((汇总.区县 == '上城'))]
        汇总 = 汇总[~((汇总.区县 == '下城'))]
        汇总 = 汇总[~((汇总.区县 == '江干'))]
        汇总 = 汇总[~((汇总.区县 == '拱墅'))]
        汇总 = 汇总[~((汇总.区县 == '西湖'))]
        汇总 = 汇总[~((汇总.区县 == '下沙'))]
        汇总 = 汇总[~((汇总.区县 == '滨江'))]
        站点数据明细 = excel_file['站点数据明细']
        站点在线明细 = excel_file['站点在线明细']
        报修 = excel_file['报修']
        总重80_90以上明细 = excel_file['总重80_90以上明细']
        超限100明细 = excel_file['超限100明细']
        百吨王明细 = excel_file['百吨王明细']
        源头站点数据明细 = excel_file['源头站点数据明细']
        df = self.总重80吨以上数据分析()
        city_list = [ '杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴', '金华', '衢州','舟山', '台州', '丽水']
        for city in city_list:
            汇总1 = 汇总[汇总['地市'] == city]
            站点在线明细1 = 站点在线明细[站点在线明细['地市'] == city]
            站点数据明细1 = 站点数据明细[站点数据明细['地市'] == city]
            报修1 = 报修[报修['地市'] == city]
            总重80_90以上明细1 = 总重80_90以上明细[总重80_90以上明细['地市'] == city]
            超限100明细1 = 超限100明细[超限100明细['地市'] == city]
            百吨王明细1 = 百吨王明细[百吨王明细['地市'] == city]
            源头站点数据明细1 = 源头站点数据明细[源头站点数据明细['city'] == city]
            with pd.ExcelWriter(r'{}\{}{}月通报数据{}v1.0.xlsx'.format(self.out_file_月报,city, self.this_month, self.today2)) as writer1:
                汇总1.to_excel(writer1, sheet_name='汇总', index=False)
                站点数据明细1.to_excel(writer1, sheet_name='站点数据明细', index=False)
                站点在线明细1.to_excel(writer1, sheet_name='站点在线明细', index=False)
                报修1.to_excel(writer1, sheet_name='报修', index=False)
                总重80_90以上明细1.to_excel(writer1, sheet_name='总重80_90以上明细', index=False)
                超限100明细1.to_excel(writer1, sheet_name='超限100明细', index=False)
                百吨王明细1.to_excel(writer1, sheet_name='百吨王明细', index=False)
                源头站点数据明细1.to_excel(writer1, sheet_name='源头站点数据明细', index=False)
            self.write_excel_file2(r'{}\{}{}月通报数据{}v1.0.xlsx'.format(self.out_file_月报,city, self.this_month, self.today2))
            city_path = r'{}\{}{}月通报数据{}v1.0.xlsx'.format(self.out_file_月报,city, self.this_month, self.today2)
            df1 = df[df['地市'] == '{}'.format(city)]
            if df1['地市'].empty == 'True':
                self.To_emil(city_path, city, self.this_month, self.today2)
                continue
            # self.creat_charts(df1,city)
            # self.add_charts(city, city_path)
            self.To_emil(city_path, city, self.this_month, self.today2)

    # import pysnooper
    # @pysnooper.snoop()
    def ShouYe(self):
        from datetime import datetime
        ks = datetime.now()
        print('首页开始时间', ks)
        self.__init__()
        # self.data_station1()
        # self.总重80吨以上明细()
        # self.超限100明细()
        # self.百吨王明细()
        # self.data_station_source()
        # self.case_statistic()
        # self.Key_freight_sources()
        self.data_station()

        self.t_bas_basic_data_report()
        # self.city_excel()
        js = datetime.now()
        sjc = js - ks
        print('首页运行耗时', sjc)
        from threading import Timer
        import datetime
        """定时1天"""
        now_time = datetime.datetime.now()
        next_time = now_time + datetime.timedelta(days=+1)
        next_year = next_time.date().year
        next_month = next_time.date().month
        next_day = next_time.date().day
        next_time2 = datetime.datetime.strptime(
            str(next_year) + "-" + str(next_month) + "-" + str(next_day) + " " + "7:00:00", "%Y-%m-%d %H:%M:%S")
        timer_start_time2 = (next_time2 - now_time).total_seconds()
        print('下次首页运行耗时', next_time2)
        t2 = Timer(timer_start_time2, self.ShouYe)  # 此处使用递归调用实现
        t2.start()

if __name__ == "__main__":
    R=Reports()
    R.ShouYe()
