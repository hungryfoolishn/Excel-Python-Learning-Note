from tkinter import *
from tkinter import filedialog
import pymysql
import pandas as pd
from pandas.tseries.offsets import Day, MonthEnd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

'''本程序不可以在远程pycharm上进行调试，会报错'''

def excel_setting(ws1, data_list, title_list, hz, l_list, r_list, c_w_list, c_w_z_list):
    # 起始行号设置
    r1 = 2
    c1 = 1
    # 是否有汇总决定是否启用hz
    # hz = 1
    # 格式标题头设置
    title_font_style_1 = Font(name='宋体', size=15, bold=True, color='FF000000')
    title_font_style_2 = Font(name='宋体', size=12, bold=True, color='FF000000')
    # 普通内容字体设置
    plain_font_style = Font(name='宋体', size=12)
    cnter_alignment_style = Alignment(horizontal='center')
    border_style = Border(top=Side(border_style='thin', color='FF000000'),
                          left=Side(border_style='thin', color='FF000000'),
                          right=Side(border_style='thin', color='FF000000'),
                          bottom=Side(border_style='thin', color='FF000000'),
                          )
    left_alignment_style = Alignment(horizontal='left')
    right_alignment_style = Alignment(horizontal='right')

    # 单个数据块循环插入
    for yf1, bt in zip(data_list, title_list):
        # 单块的数据块插入

        lt = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L',
              13: 'M',
              14: 'N', 15: 'O', 16: 'P', 16: 'Q'}
        for i in range(yf1.shape[1]):
            ws1.cell(row=r1, column=c1 + i).value = yf1.columns[i]
        for i in range(yf1.shape[1]):
            for j in range(yf1.shape[0]):
                ws1.cell(row=r1 + 1 + j, column=c1 + i).value = yf1.iloc[j, i]
        ws1['A' + str(r1 - 1)] = bt
        ws1.merge_cells('A' + str(r1 - 1) + ':' + lt[yf1.shape[1]] + str(r1 - 1))
        if hz == 1:
            ws1.merge_cells('A' + str(r1 + len(yf1)) + ':' + 'B' + str(r1 + len(yf1)))
        # 整体格式设置
        for row in ws1['A' + str(r1 - 1):lt[yf1.shape[1]] + str(r1 + yf1.shape[0])]:
            for c in row:
                c.font = plain_font_style
                c.border = border_style
                c.alignment = cnter_alignment_style
        # 特殊块格式设置-靠左-特殊的长文本
        for l in l_list:
            for row in ws1[l + str(r1 + 1):l + str(r1 + yf1.shape[0] - hz)]:
                for c in row:
                    c.alignment = left_alignment_style
        # 特殊块格式设置-靠右-数值靠右
        for r in r_list:
            for row in ws1[r + str(r1 + 1):r + str(r1 + yf1.shape[0])]:
                for c in row:
                    c.alignment = right_alignment_style
        # 特殊块格式设置-表头
        for row in ws1[r1 - 1]:
            row.font = title_font_style_1
        for row in ws1[r1]:
            row.font = title_font_style_2
        r1 = r1 + len(yf1) + 6

        # print(r1)
    # 列名宽度，具体情况具体调整
    for i, j in zip(c_w_list, c_w_z_list):
        ws1.column_dimensions[i].width = j
    pass

def set_columns_date_frame(df,select_rows,tail_rows):
    #得到一个包含列名的大的数据块
    ff=df.values
    #数据块分为两块：不规则列名块和规则数据块
    #第一步处理列名
    if tail_rows==0:
        f1 = pd.DataFrame(ff[0:select_rows, :]).fillna(method='ffill', axis=0)
        #第二步处理后列名和数据块打包
        f2 = pd.DataFrame(ff[select_rows:, :], columns=f1.values[-1])
    else:
        f1 = pd.DataFrame(ff[0:select_rows, :]).fillna(method='ffill', axis=0)
        # 第二步处理后列名和数据块打包
        f2 = pd.DataFrame(ff[select_rows:-tail_rows, :], columns=f1.values[-1])
    return f2

def get_data_from_sql(sql, cs_list):
    global db
    # db = pymysql.connect(host='192.168.2.119',user='zjzhzcuser',passwd='F4dus0ee',
    #         port=3306,charset='utf8',database='db_manage_overruns')
    db = pymysql.connect(host="192.168.1.229", port=3306, user='root', password='zcits123456',
                         database='db_manage_overruns')

    cursor = db.cursor()
    #print('1')
    #print(cs_list)
    cursor.execute(sql, cs_list)
    #print('2')
    data = cursor.fetchall()
    #print('3')
    data = pd.DataFrame(list(data))
    if data.empty == False:
        data.columns = [i[0] for i in cursor.description]
        return data
    else:
        return data


def set_columns_date_frame(df, select_rows, tail_rows):
    # 得到一个包含列名的大的数据块
    ff = df.values
    # 数据块分为两块：不规则列名块和规则数据块
    # 第一步处理列名
    if tail_rows == 0:
        f1 = pd.DataFrame(ff[0:select_rows, :]).fillna(method='ffill', axis=0)
        # 第二步处理后列名和数据块打包
        f2 = pd.DataFrame(ff[select_rows:, :], columns=f1.values[-1])
    else:
        f1 = pd.DataFrame(ff[0:select_rows, :]).fillna(method='ffill', axis=0)
        # 第二步处理后列名和数据块打包
        f2 = pd.DataFrame(ff[select_rows:-tail_rows, :], columns=f1.values[-1])
    return f2


def qq():
    pass


def func1():
    # a=filedialog.asksaveasfilename()#返回文件名

    # a =filedialog.asksaveasfile()#会创建文件

    # a =filedialog.askopenfilename()#返回文件名

    # a =filedialog.askopenfile()#返回文件流对象

    # a =filedialog.askdirectory()#返回目录名

    a = filedialog.askopenfilenames()  # 可以返回多个文件名

    # a=filedialog.askopenfiles()#多个文件流对象


def sr_1():
    global a1
    a1 = filedialog.askopenfilenames()  # 可以返回多个文件名的地址的元组
    sr_e_1.set(a1)


def sr_2():
    c_e_1.set("")
    global a2
    a2 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
    sr_e_2.set(a2)


def sr_3():
    # 市级归0
    # sr_e_5.set("")
    # sr_e_6.set("")
    # c_e_1.set("")
    global a3
    a3 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
    sr_e_3.set(a3)


# def sr_4():
#     c_e_1.set("")
#     global a4
#     a4 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
#     sr_e_4.set(a4)
def sr_5():
    # 省级归0
    sr_d_1.set("")
    sr_d_2.set("")
    sr_e_3.set("")
    # 输出归0
    c_e_1.set("")
    global a5
    a5 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
    sr_e_5.set(a5)


def sr_6():
    # 省级归0
    sr_d_1.set("")
    sr_d_2.set("")
    sr_e_3.set("")
    # 输出归0
    c_e_1.set("")
    global a6
    a6 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
    sr_e_6.set(a6)


def tjfx():
    E1.config(state='readonly')
    E4.config(state='readonly')
    E3.config(state='readonly')

    cs = sr_e_1.get()

    #qishi_date = sr_d_1.get()
    #qishi_date = sr_d_1.get()
    jiezi_date = sr_d_2.get()

    yd=sr_e_3.get()

    dizhi = sc_e_1.get()
    # 总表
    # sql_808 = "SELECT b.city as '地市',a.area_city as '地市编码',b.county as '区县',a.area_county as '区县编码',a.site_name as '站点名称',a.out_station_time as '检测时间',a.car_no as '车牌号码',a.total_weight as '总重',a.limit_weight as '限重',a.overrun as '超重',a.axis as '轴数',a.overrun_rate as '超限率' \
    #             FROM t_bas_over_data_31 a \
    #             LEFT JOIN t_code_area b ON b.county_code = a.area_county \
    #             WHERE a.total_weight >=80 AND a.out_station_time >= %s AND a.out_station_time <%s and a.site_name in %s \
    #             ORDER BY a.area_city,a.area_county,a.site_name"

    wc_list = []
    if cs != '' and yd!='' and len(jiezi_date) == 7 and dizhi != '':
        # 时间转换
        try:

            city_list = cs.split('-')
            # #print('1')
            # qishi_time = datetime.strptime(jiezi_date, '%Y-%m')
            # #print('2')
            # jiezi_time = datetime.strptime(jiezi_date, '%Y-%m')
            # #print('3')
            # jiezi_time = jiezi_time + MonthEnd() + Day()
            # #print('4')
            # city_tuple=tuple(set_columns_date_frame(pd.read_excel(yd),2,0)['站点名称'])

            try:
                # 试验区
                df_big = yd
                # 大区块开始分

                b1_8 = df_big[df_big['总重'] >= 80].groupby(['地市', '地市编码'])['总重'].count().rename('80吨以上')
                b1_9 = df_big[df_big['总重'] >= 90].groupby(['地市', '地市编码'])['总重'].count().rename('90吨以上')
                b1 = pd.concat([b1_8, b1_9], axis=1).reset_index().sort_values(by=['地市编码'], ascending=True)

                # 一级省市汇总表
                b1 = b1[['地市', '80吨以上', '90吨以上']]
                b1.insert(0, '序号', [i + 1 for i in range(len(b1))])

                b11 = pd.DataFrame({'序号': ['汇总'], '80吨以上': [b1['80吨以上'].sum()], '90吨以上': [b1['90吨以上'].sum()]})
                b12 = pd.concat([b1, b11], axis=0)

                wb1 = Workbook()
                excel_setting(wb1.active, [b12], ['至' + jiezi_date + '浙江省80吨、90吨汇总表'], 1, [], [], [], [])
                wb1.save(dizhi + '\\' + "汇总表.xlsx")
                wc_list.append('月度汇总表正常生成')
            except:
                wc_list.append('月度汇总表未生成')
            for j in city_list:
                try:
                    # 二级区县汇总表
                    df = df_big[df_big['地市'] == j]
                    if df.empty == False:

                        b2_8 = df[df['总重'] >= 80].groupby(['地市', '区县', '区县编码'])['总重'].count().rename('80吨以上')
                        b2_9 = df[df['总重'] >= 90].groupby(['地市', '区县', '区县编码'])['总重'].count().rename('90吨以上')
                        b2 = pd.concat([b2_8, b2_9], axis=1).reset_index().sort_values(by=['区县编码'], ascending=True)
                        b2 = b2[['地市', '区县', '80吨以上', '90吨以上']]
                        b2.insert(0, '序号', [i + 1 for i in range(len(b2))])
                        b22 = pd.DataFrame(
                            {'序号': ['汇总'], '80吨以上': [b2['80吨以上'].sum()], '90吨以上': [b2['90吨以上'].sum()]})
                        b222 = pd.concat([b2, b22], axis=0)


                        b3_8 = df[df['总重'] >= 80].groupby(['地市', '区县', '区县编码', '站点名称'])['总重'].count().rename('80吨以上')
                        b3_9 = df[df['总重'] >= 90].groupby(['地市', '区县', '区县编码', '站点名称'])['总重'].count().rename('90吨以上')
                        b3 = pd.concat([b3_8, b3_9], axis=1).reset_index().sort_values(by=['区县编码', '站点名称'], ascending=True)
                        b3 = b3[['地市', '区县', '站点名称', '80吨以上', '90吨以上']]
                        b3.insert(0, '序号', [i + 1 for i in range(len(b3))])
                        b33 = pd.DataFrame(
                            {'序号': ['汇总'], '80吨以上': [b3['80吨以上'].sum()], '90吨以上': [b3['90吨以上'].sum()]})
                        b333 = pd.concat([b3, b33], axis=0)


                        b80 = df[df['总重'] >= 80].sort_values(by=['区县编码', '站点名称'], ascending=True)
                        b80 = b80[['地市', '区县', '站点名称', '检测时间', '车牌号码', '总重', '限重', '超重', '轴数', '超限率']]
                        b80.insert(0, '序号', [i + 1 for i in range(len(b80))])

                        b90 = df[df['总重'] >= 90].sort_values(by=['区县编码', '站点名称'], ascending=True)
                        b90 = b90[['地市', '区县', '站点名称', '检测时间', '车牌号码', '总重', '限重', '超重', '轴数', '超限率']]
                        b90.insert(0, '序号', [i + 1 for i in range(len(b90))])
                        # print(b90.head(6))
                        wb = Workbook()
                        ws1 = wb.create_sheet(j + '市80-90吨汇总表')
                        ws2 = wb.create_sheet(j + '市80-90吨站点汇总表')
                        ws3 = wb.create_sheet(j + '市80吨明细表')
                        ws4 = wb.create_sheet(j + '市90吨明细表')
                        # print('5')

                        excel_setting(ws1, [b222], [jiezi_date + '单月汇总表'], 1, [], [], [], [])

                        excel_setting(ws2, [b333], [jiezi_date + '单月站点汇总表'], 1, ['D'], [], ['D'],
                                      [40])

                        excel_setting(ws3, [b80], [jiezi_date + '单月80吨明细表'], 0, ['D'],
                                      ['G', 'H', 'I', 'J', 'K'], ['D', 'E'], [40, 24])

                        excel_setting(ws4, [b90], [jiezi_date + '单月90吨明细表'], 0, ['D'],
                                      ['G', 'H', 'I', 'J', 'K'], ['D', 'E'], [40, 24])
                        wb.remove(wb.active)
                        wb.save(dizhi + '\\' + j + "市80-90吨汇总及明细表" + ".xlsx")
                        wc_list.append(j + '相关表已生成')
                    else:
                        wc_list.append(j + '无相关表')

                except:
                    wc_list.append(j + '出现问题')
            pd.DataFrame({'完成情况': wc_list}).to_csv(
                dizhi + '\\'+
                jiezi_date.split('-')[-2] + jiezi_date.split('-')[-1] + '日志' + '.txt', sep='\t', index=False)
            c_e_1.set("报表已经处理完成，请查看！！！")
            db.close()
        except:
            c_e_1.set("参数格式有误！")

    else:
        c_e_1.set("参数设置遗漏！")


def sc_10():
    a = filedialog.asksaveasfile()  # 会创建文件 适用于单个文件的输出


def sc_1():
    c_e_1.set("")
    global dz
    dz = filedialog.askdirectory()  # 返回目录名 适用于多个文件的输出
    sc_e_1.set(dz)


def czhs():
    # 文本状态释放
    E1.config(state='normal')
    #E4.config(state='normal')
    E3.config(state='normal')
    # 条件归0
    sr_e_1.set("杭州-宁波-温州-嘉兴-湖州-绍兴-金华-衢州-舟山-台州-丽水")
    # 省级归0
    # sr_d_1.set("")
    # sr_d_2.set("")
    # sr_e_3.set("")
    # 市级归0
    # sr_e_5.set("")
    # sr_e_6.set("")
    # 输出结果归0
    c_e_1.set("")


root = Tk()
root.resizable(width=False, height=False)
s1 = StringVar()
s1.set("")
s2 = StringVar()
s2.set("")
root.title('80-90吨单月汇总专用')

# 以下为省级报表输出框架


k3 = Frame(root, width=500, height=200)
k3.grid(row=10, column=0, rowspan=2)

# qzcs=LabelFrame(k3,text='条件^-^',width=500,height=200)
# qzcs.grid(row=0,column=0,rowspan=2)
#
# sr_e_1=StringVar()
# sr_e_1.set("")
# Label(qzcs,text='请输入省会或地市',width=15).grid(row=0,column=0,rowspan=1,pady=2,padx=4)
# E1=Entry(qzcs,textvariable=sr_e_1,width=40)
# E1.grid(row=0,column=1,rowspan=1,pady=2,padx=4)
# #Button(k3,text="选择文件",width=15,command=sr_1).grid(row=0,column=2,rowspan=1,pady=2,padx=4)
# Button(qzcs,text="重-置",width=15,command=czhs).grid(row=0,column=2,rowspan=1,pady=2,padx=4)
#
# sr_e_2=StringVar()
# sr_e_2.set("")
# Label(qzcs,text='选择月度汇总表',width=15).grid(row=1,column=0,rowspan=1,pady=2,padx=4)
# Entry(qzcs,textvariable=sr_e_2,width=40,state='readonly').grid(row=1,column=1,rowspan=1,pady=2,padx=4)
# Button(qzcs,text="选择文件",width=15,command=sr_2).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

Frame(k3, width=550, height=5, bg='red').grid(row=3, column=0, rowspan=1, pady=2, padx=4)

sjbb = LabelFrame(k3, text='省级报表', width=500, height=200)
sjbb.grid(row=4, column=0, rowspan=2)

sr_e_1 = StringVar()
sr_e_1.set("杭州-宁波-温州-嘉兴-湖州-绍兴-金华-衢州-舟山-台州-丽水")
Label(sjbb, text='地市列表', width=15).grid(row=0, column=0, rowspan=1, pady=2, padx=4)
E1 = Entry(sjbb, textvariable=sr_e_1, width=40)
E1.grid(row=0, column=1, rowspan=1, pady=2, padx=4)
Button(sjbb, text="重-置", width=15, command=czhs).grid(row=0, column=2, rowspan=1, pady=2, padx=4)

# sr_d_1 = StringVar()
# sr_d_1.set("")
# Label(sjbb, text='输入起始时间', width=15).grid(row=2, column=0, rowspan=1, pady=2, padx=4)
# E2 = Entry(sjbb, textvariable=sr_d_1, width=40)
# E2.grid(row=2, column=1, rowspan=1, pady=2, padx=4)
# Label(sjbb, text='例：2022-04', width=15).grid(row=2, column=2, rowspan=1, pady=2, padx=4)

sr_d_2 = StringVar()
sr_d_2.set("")
Label(sjbb, text='输入当月时间', width=15).grid(row=4, column=0, rowspan=1, pady=2, padx=4)
E3 = Entry(sjbb, textvariable=sr_d_2, width=40)
E3.grid(row=4, column=1, rowspan=1, pady=2, padx=4)
Label(sjbb, text='例：2022-05', width=15).grid(row=4, column=2, rowspan=1, pady=2, padx=4)

sr_e_3=StringVar()
sr_e_3.set("")
Label(sjbb,text='选择当月汇总表',width=15).grid(row=8,column=0,rowspan=1,pady=2,padx=4)
E4=Entry(sjbb,textvariable=sr_e_3,width=40,state='readonly')
E4.grid(row=8,column=1,rowspan=1,pady=2,padx=4)
Button(sjbb,text="选择文件",width=15,command=sr_3).grid(row=8,column=2,rowspan=1,pady=2,padx=4)

# sr_e_4=StringVar()
# sr_e_4.set("")
# Label(sjbb,text='选择报修数文件',width=15).grid(row=9,column=0,rowspan=1,pady=2,padx=4)
# Entry(sjbb,textvariable=sr_e_4,width=40,state='readonly').grid(row=9,column=1,rowspan=1,pady=2,padx=4)
# Button(sjbb,text="选择文件",width=15,command=sr_4).grid(row=9,column=2,rowspan=1,pady=2,padx=4)
# Frame(k3,width=300,height=5,bg='red').grid(row=6,column=0,rowspan=1,pady=2,padx=4)
#
#
# dsbb=LabelFrame(k3,text='市级报表',width=500,height=200)
# dsbb.grid(row=8,column=0,rowspan=2)
#
# sr_e_6=StringVar()
# sr_e_6.set("")
# Label(dsbb,text='地市超限100%数据',width=15).grid(row=10,column=0,rowspan=1,pady=2,padx=4)
# Entry(dsbb,textvariable=sr_e_6,width=40,state='readonly').grid(row=10,column=1,rowspan=1,pady=2,padx=4)
# Button(dsbb,text="选择文件",width=15,command=sr_6).grid(row=10,column=2,rowspan=1,pady=2,padx=4)
#
# sr_e_5=StringVar()
# sr_e_5.set("")
# Label(dsbb,text='选择区县模版',width=15).grid(row=11,column=0,rowspan=1,pady=2,padx=4)
# Entry(dsbb,textvariable=sr_e_5,width=40,state='readonly').grid(row=11,column=1,rowspan=1,pady=2,padx=4)
# Button(dsbb,text="选择文件",width=15,command=sr_5).grid(row=11,column=2,rowspan=1,pady=2,padx=4)

Frame(root, width=550, height=5, bg='red').grid(row=20, column=0, rowspan=1, pady=2, padx=4)

k6 = LabelFrame(root, text="为输出找个目录吧！", width=500, height=200)
k6.grid(row=25, column=0, rowspan=2)

sc_e_1 = StringVar()
sc_e_1.set("")
Label(k6, text='输出到', width=15).grid(row=0, column=0, rowspan=1, pady=2, padx=4)
Entry(k6, textvariable=sc_e_1, width=40, state='readonly').grid(row=0, column=1, rowspan=1, pady=2, padx=4)
Button(k6, text="选择目录", width=15, command=sc_1).grid(row=0, column=2, rowspan=1, pady=2, padx=4)
# Entry(k6,textvariable=sc_dwj_e_1,width=40).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
# Button(k6,text="处理情况",width=15,command=sc_dwj).grid(row=1,column=0,rowspan=1,pady=2,padx=4)
# Button(k6,text="保存结果",width=15,command=sc_dwj).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

Frame(root, width=550, height=5, bg='red').grid(row=27, column=0, rowspan=1, pady=2, padx=4)

k8 = LabelFrame(root, text="进行分析并获取结果", width=500, height=200)
k8.grid(row=30, column=0, rowspan=2)

c_e_1 = StringVar()
c_e_1.set("")
Label(k8, text='处理结果', width=15).grid(row=0, column=0, rowspan=1, pady=2, padx=4)
Entry(k8, textvariable=c_e_1, width=40, state='readonly').grid(row=0, column=1, rowspan=1, pady=2, padx=4)
Button(k8, text="提交分析", foreground="green", width=15, command=tjfx).grid(row=0, column=2, rowspan=1, pady=2, padx=4)

# 以下为区级报表生成框架
# dsbb=LabelFrame(root,text='地市报表输出',width=500,height=200)
# dsbb.grid(row=2,column=0,rowspan=2)
#
# l3=LabelFrame(dsbb,text="输入需要分析的参数 ^-^",width=500,height=200,bg='Red')
# l3.grid(row=200,column=0,rowspan=2)
#
#
# sr_e_1=StringVar()
# sr_e_1.set("")
# Label(l3,text='请输入省会或城市',width=15).grid(row=0,column=0,rowspan=1,pady=2,padx=4)
# Entry(l3,textvariable=sr_e_1,width=40).grid(row=0,column=1,rowspan=1,pady=2,padx=4)
#
# sr_e_6=StringVar()
# sr_e_6.set("")
# Label(l3,text='地市超限100%数据',width=15).grid(row=10,column=0,rowspan=1,pady=2,padx=4)
# Entry(l3,textvariable=sr_e_6,width=40,state='readonly').grid(row=10,column=1,rowspan=1,pady=2,padx=4)
# Button(l3,text="选择文件",width=15,command=sr_6).grid(row=10,column=2,rowspan=1,pady=2,padx=4)
#
# sr_e_5=StringVar()
# sr_e_5.set("")
# Label(l3,text='选择市区相应模版',width=15).grid(row=11,column=0,rowspan=1,pady=2,padx=4)
# Entry(l3,textvariable=sr_e_5,width=40,state='readonly').grid(row=11,column=1,rowspan=1,pady=2,padx=4)
# Button(l3,text="选择文件",width=15,command=sr_5).grid(row=11,column=2,rowspan=1,pady=2,padx=4)

# Text(k8,width=40,height=10).grid(row=1,column=0,rowspan=1,columnspan=1,pady=2,padx=4)
# Entry(k8,textvariable=sc_dwj_e_1,width=40).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
# Button(k8,text="结果预览",width=15).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
# Button(k8,text="保存结果",width=15,command=sc_dwj).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

root.mainloop()