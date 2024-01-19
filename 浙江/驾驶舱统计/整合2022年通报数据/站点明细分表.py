from tkinter import *
from tkinter import filedialog
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment
#表头处理函数
def set_columns_date_frame(df,select_rows,tail_rows):
    #得到一个包含列名的大的数据块
    ff=df.values
    #数据块分为两块：不规则列名块和规则数据块
    #第一步处理列名
    if tail_rows==0:
        f1 = pd.DataFrame(ff[0:select_rows, :]).fillna(method='ffill', axis=0)
        #第二步处理后列名和数据块打包
        f2 = pd.DataFrame(ff[select_rows:, :], columns=f1.values[-1])
    else:
        f1 = pd.DataFrame(ff[0:select_rows, :]).fillna(method='ffill', axis=0)
        # 第二步处理后列名和数据块打包
        f2 = pd.DataFrame(ff[select_rows:-tail_rows, :], columns=f1.values[-1])
    return f2

def excel_setting(ws1, data_list, title_list, hz, l_list, r_list, c_w_list, c_w_z_list):
    # 起始行号设置
    r1 = 3
    c1 = 1
    # 是否有汇总决定是否启用hz
    # hz = 1
    # 格式标题头设置
    title_font_style_1 = Font(name='宋体', size=15, bold=True, color='FF000000')
    title_font_style_2 = Font(name='宋体', size=12, bold=True, color='FF000000')
    # 普通内容字体设置
    plain_font_style = Font(name='宋体', size=12)
    cnter_alignment_style = Alignment(horizontal='center',vertical='center',wrap_text=True)
    border_style = Border(top=Side(border_style='thin', color='FF000000'),
                          left=Side(border_style='thin', color='FF000000'),
                          right=Side(border_style='thin', color='FF000000'),
                          bottom=Side(border_style='thin', color='FF000000'),
                          )
    left_alignment_style = Alignment(horizontal='left',vertical='center',wrap_text=False)
    right_alignment_style = Alignment(horizontal='right',vertical='center',wrap_text=False)
    center_alignment_style=Alignment(horizontal='center',vertical='center',wrap_text=False)
    # 单个数据块循环插入
    for yf1, bt in zip(data_list, title_list):
        # 单块的数据块插入

        lt = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J',
              11: 'K', 12: 'L',13: 'M',14: 'N', 15: 'O', 16: 'P', 17: 'Q',18:'R',19:'S',20:'T',
              21:'U',22:'V',23:'W',24:'X',25:'Y',26:'Z'}
        for i in range(yf1.shape[1]):
            ws1.cell(row=r1, column=c1 + i).value = yf1.columns[i]
        for i in range(yf1.shape[1]):
            for j in range(yf1.shape[0]):
                ws1.cell(row=r1 + 1 + j, column=c1 + i).value = yf1.iloc[j, i]
        ws1['A' + str(r1 - 2)] = bt
        ws1['A' + str(r1 - 1)] = '序号'
        ws1['B' + str(r1 - 1)] = '站点名称'
        ws1['C' + str(r1 - 1)] = '地市'
        ws1['D' + str(r1 - 1)] = '区县'
        ws1['E' + str(r1 - 1)] = '在线情况'
        ws1['H' + str(r1 - 1)] = '超限情况'
        ws1['Q' + str(r1 - 1)] = '满足处罚'
        ws1['R' + str(r1 - 1)] = '是否直传'
        ws1['S' + str(r1 - 1)] = '状态'



        ws1.merge_cells('A' + str(r1 - 2) + ':' + lt[yf1.shape[1]] + str(r1 - 2))
        ws1.merge_cells('A' + str(r1 - 1) + ':' + 'A' + str(r1))
        ws1.merge_cells('B' + str(r1 - 1) + ':' + 'B' + str(r1))
        ws1.merge_cells('C' + str(r1 - 1) + ':' + 'C' + str(r1))
        ws1.merge_cells('D' + str(r1 - 1) + ':' + 'D' + str(r1))
        ws1.merge_cells('Q' + str(r1 - 1) + ':' + 'Q' + str(r1))
        ws1.merge_cells('R' + str(r1 - 1) + ':' + 'R' + str(r1))
        ws1.merge_cells('S' + str(r1 - 1) + ':' + 'S' + str(r1))
        ws1.merge_cells('E' + str(r1 - 1) + ':' + 'G' + str(r1 - 1))
        ws1.merge_cells('H' + str(r1 - 1) + ':' + 'P' + str(r1 - 1))



        if hz == 1:
            ws1.merge_cells('A' + str(r1 + len(yf1)) + ':' + 'B' + str(r1 + len(yf1)))
        # 整体格式设置
        for row in ws1['A' + str(r1 - 2):lt[yf1.shape[1]] + str(r1 + yf1.shape[0])]:
            for c in row:
                c.font = plain_font_style
                c.border = border_style
                c.alignment = cnter_alignment_style
        # 特殊块格式设置-靠左-特殊的长文本
        for l in l_list:
            for row in ws1[l + str(r1 + 1):l + str(r1 + yf1.shape[0] - hz)]:
                for c in row:
                    c.alignment = left_alignment_style
        # 特殊块格式设置-靠右-数值靠右
        for r in r_list:
            for row in ws1[r + str(r1 + 1):r + str(r1 + yf1.shape[0])]:
                for c in row:
                    c.alignment = right_alignment_style
        # 特殊块格式设置-表头
        for row in ws1[r1 - 2]:
            row.font = title_font_style_1
        for row in ws1[r1-1]:
            row.font = title_font_style_2
        for row in ws1[r1]:
            row.font = title_font_style_2
        #表尾格式设置
        if len(yf1['理应在线天数'].dropna())<len(yf1['理应在线天数']):
            ws1['E' + str(r1 + len(yf1['理应在线天数'].dropna()) + 1)] = '站点维修'
            ws1.merge_cells('E' + str(r1 + len(yf1['理应在线天数'].dropna()) + 1) + ':' + 'P' + str(r1 + len(yf1['理应在线天数'])))

            ws1['E' + str(r1 + len(yf1['理应在线天数'].dropna())+1)].alignment=center_alignment_style
        # str(r1 + len(yf1['理应在线天数'].dropna()) + 1)
        hong=yf1[yf1['实际在线天数']<=10]['序号'].tolist()
        huang=yf1[(yf1['实际在线天数']>10)&(yf1['货车数']<500)]['序号'].tolist()
        for i in hong:
            ws1['F'+str(r1+i)].fill=PatternFill(fill_type='solid',fgColor='FFFF2100')

        for i in huang:
            ws1['H'+str(r1+i)].fill=PatternFill(fill_type='solid',fgColor='FFFFFF00')
        #rule1=CellIsRule(operator='between',formula=[0,10],fill=PatternFill(end_color='FFFF2100'))
        #ws1.conditional_formatting.add('F'+str(r1+1)+':F'+str(r1+len(yf1['理应在线天数'].dropna())),rule1)

        #rule2 = CellIsRule(operator='lessThanOrEqual', formula=[500], fill=PatternFill(end_color='FFFFFF00'))
        #ws1.conditional_formatting.add('H' + str(r1 + 1) + ':H' + str(r1+len(yf1['理应在线天数'].dropna())), rule2)


        r1 = r1 + len(yf1) + 6

        # print(r1)
    # 列名宽度，具体情况具体调整
    for i, j in zip(c_w_list, c_w_z_list):
        ws1.column_dimensions[i].width = j

def qq():
    pass

def func1():
    #a=filedialog.asksaveasfilename()#返回文件名

    #a =filedialog.asksaveasfile()#会创建文件

    #a =filedialog.askopenfilename()#返回文件名

    #a =filedialog.askopenfile()#返回文件流对象

    #a =filedialog.askdirectory()#返回目录名

    a =filedialog.askopenfilenames()#可以返回多个文件名

    #a=filedialog.askopenfiles()#多个文件流对象
def sr_1():
    global a1
    a1 = filedialog.askopenfilenames()  # 可以返回多个文件名的地址的元组
    sr_e_1.set(a1)
def sr_2():
    global a2
    a2 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
    sr_e_2.set(a2)
def sr_3():
    sr_e_5.set("")
    c_e_1.set("")
    global a3
    a3 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
    sr_e_3.set(a3)
def sr_4():
    global a4
    a4 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
    sr_e_4.set(a4)
def sr_6():
    global a6
    a6 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
    sr_e_6.set(a6)
def sr_5():
    sr_e_3.set("")
    c_e_1.set("")
    global a5
    a5 = filedialog.askopenfilename()  # 可以返回多个文件名的地址的元组
    sr_e_5.set(a5)
def tjfx():
    EE.config(state='readonly')
    yf = sr_e_1.get()
    zc = sr_e_2.get()
    yc = sr_e_4.get()
    zj = sr_e_6.get()
    cs = sc_e_1.get()
    if 0<len(yf)<=2 and zc!='' and yc!='' and zj!='' and cs!='':
        '''定义月份'''
        #yf = 6
        '''读取抓拍的站点文件'''
        ss = pd.read_excel(zj)
        cf_list = ss['station_name'].tolist()

        '''读取正常站点的数据'''
        df1 = pd.read_excel(zc)
        df1 = set_columns_date_frame(df1, 2, 0)
        df1['状态'] = '启用'

        '''读取报修点位的数据'''
        df2 = pd.read_excel(yc)
        # df2=pd.read_excel(r"C:\Users\Admin\Desktop\临时数据处理表\湖州7月报修点位统计表 (1).xls")
        df2 = set_columns_date_frame(df2, 1, 0)
        df2.columns = ['地市', '区县', '站点名称', '状态']
        df2['状态'] = '维修'
        # print(df2.head(70))
        df3 = pd.concat([df1, df2], axis=0)

        '''义乌单选备用'''
        # df3.loc[df3['区县']=='义乌','地市']='义乌'
        df3.loc[df3['站点名称'].isin(cf_list), '最后接受时间'] = '是'
        df3.loc[~(df3['站点名称'].isin(cf_list)), '最后接受时间'] = '否'

        df3.insert(len(df3.columns) - 1, '是否直传', '')
        #df3.to_excel(cs+'\\'+yf+'月份所有站点超限情况明细.xlsx')
        city_list = df3['地市'].dropna().drop_duplicates().tolist()

        wb = Workbook()
        for i in city_list:
            tt = wb.create_sheet(i)
            dd = df3[df3['地市'] == i]
            dd.insert(0, '序号', [i + 1 for i in range(len(dd))])
            # dd.insert(16, '满足处罚', [i + 1 for i in range(len(dd))])
            # print(dd.columns)
            dd = dd[['序号', '站点名称', '地市', '区县', '理应在线天数', '实际在线天数',
                     '在线率', '货车数', '超限数', '超限10%除外数', '超限10%除外超限率(%)',
                     '超限20%除外数', '超限20%除外超限率(%)', '百吨王数', '超限100%数', '超限率(%)', '最后接受时间', '是否直传', '状态']]
            dd.columns = ['序号', '站点名称', '地市', '区县', '理应在线天数', '实际在线天数',
                          '在线率', '货车数', '超限数', '剔除10%超限数（不包含临界点）', '剔除10%后的超限率',
                          '剔除20%超限数（不包含临界点）', '剔除20%后的超限率', '百吨王数', '超限100%数', '超限率(%)', '满足处罚', '是否直传', '状态']

            excel_setting(tt, [dd], [i + yf + '月份站点超限明细'], 0, ['B'],
                          ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P'], ['B', 'J', 'L', 'C', 'D'],
                          [43, 9.5, 9.5, 11, 11])

        wb.remove(wb.active)
        wb.save(cs+'\\'+yf+'月份站点超限明细.xlsx')




        c_e_1.set("处理已完成，请查看！！！")
    else:
        c_e_1.set("参数设置出错！")

def sc_10():
    a = filedialog.asksaveasfile()  # 会创建文件 适用于单个文件的输出
def sc_1():
    c_e_1.set("")
    global b1
    b1 = filedialog.askdirectory()  # 返回目录名 适用于多个文件的输出
    sc_e_1.set(b1)
def czhs():
    EE.config(state='normal')
    sr_e_1.set("")
    #sr_e_5.set("")
    #sr_e_3.set("")
    c_e_1.set("")



root=Tk()
root.resizable(width=False, height=False)
s1=StringVar()
s1.set("")
s2=StringVar()
s2.set("")
root.title('超限站点明细分表')





k3=LabelFrame(root,text="输入需要分析的参数 ^-^",width=500,height=200)
k3.grid(row=1,column=0,rowspan=2)


sr_e_1=StringVar()
sr_e_1.set("6")
Label(k3,text='请输入月份',width=15).grid(row=0,column=0,rowspan=1,pady=2,padx=4)
EE=Entry(k3,textvariable=sr_e_1,width=40)
EE.grid(row=0,column=1,rowspan=1,pady=2,padx=4)
#Button(k3,text="选择文件",width=15,command=sr_1).grid(row=0,column=2,rowspan=1,pady=2,padx=4)
Button(k3,text="重-置",width=15,command=czhs).grid(row=0,column=2,rowspan=1,pady=2,padx=4)

sr_e_2=StringVar()
sr_e_2.set("")
Label(k3,text='选择月度汇总表',width=15).grid(row=1,column=0,rowspan=1,pady=2,padx=4)
Entry(k3,textvariable=sr_e_2,width=40,state='readonly').grid(row=1,column=1,rowspan=1,pady=2,padx=4)
Button(k3,text="选择文件",width=15,command=sr_2).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

sr_e_4=StringVar()
sr_e_4.set("")
Label(k3,text='选择报修数文件',width=15).grid(row=2,column=0,rowspan=1,pady=2,padx=4)
Entry(k3,textvariable=sr_e_4,width=40,state='readonly').grid(row=2,column=1,rowspan=1,pady=2,padx=4)
Button(k3,text="选择文件",width=15,command=sr_4).grid(row=2,column=2,rowspan=1,pady=2,padx=4)

sr_e_6=StringVar()
sr_e_6.set("")
Label(k3,text='选择证据站点',width=15).grid(row=3,column=0,rowspan=1,pady=2,padx=4)
Entry(k3,textvariable=sr_e_6,width=40,state='readonly').grid(row=3,column=1,rowspan=1,pady=2,padx=4)
Button(k3,text="选择文件",width=15,command=sr_6).grid(row=3,column=2,rowspan=1,pady=2,padx=4)


Frame(root,width=550,height=5,bg='red').grid(row=13,column=0,rowspan=1,pady=2,padx=4)

k6=LabelFrame(root,text="为输出找个目录吧！",width=500,height=200)
k6.grid(row=25,column=0,rowspan=2)

sc_e_1=StringVar()
sc_e_1.set("")
Label(k6,text='输出到',width=15).grid(row=0,column=0,rowspan=1,pady=2,padx=4)
Entry(k6,textvariable=sc_e_1,width=40,state='readonly').grid(row=0,column=1,rowspan=1,pady=2,padx=4)
Button(k6,text="选择目录",width=15,command=sc_1).grid(row=0,column=2,rowspan=1,pady=2,padx=4)
# Entry(k6,textvariable=sc_dwj_e_1,width=40).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
# Button(k6,text="处理情况",width=15,command=sc_dwj).grid(row=1,column=0,rowspan=1,pady=2,padx=4)
# Button(k6,text="保存结果",width=15,command=sc_dwj).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

Frame(root,width=550,height=5,bg='red').grid(row=28,column=0,rowspan=1,pady=2,padx=4)


k8=LabelFrame(root,text="进行分析并获取结果",width=500,height=200)
k8.grid(row=30,column=0,rowspan=2)

c_e_1=StringVar()
c_e_1.set("")
Label(k8,text='处理结果',width=15).grid(row=0,column=0,rowspan=1,pady=2,padx=4)
Entry(k8,textvariable=c_e_1,width=40,state='readonly').grid(row=0,column=1,rowspan=1,pady=2,padx=4)
Button(k8,text="提交分析",width=15,command=tjfx).grid(row=0,column=2,rowspan=1,pady=2,padx=4)

#Text(k8,width=40,height=10).grid(row=1,column=0,rowspan=1,columnspan=1,pady=2,padx=4)
#Entry(k8,textvariable=sc_dwj_e_1,width=40).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
#Button(k8,text="结果预览",width=15).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
#Button(k8,text="保存结果",width=15,command=sc_dwj).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

root.mainloop()