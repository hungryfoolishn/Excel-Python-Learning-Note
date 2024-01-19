from tkinter import *
from tkinter import filedialog
import pymysql
import pandas as pd
from pandas.tseries.offsets import Day, MonthEnd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment


def excel_setting(ws1, data_list, title_list, hz, l_list, r_list, c_w_list, c_w_z_list):
    # ��ʼ�к�����
    r1 = 2
    c1 = 1
    # �Ƿ��л��ܾ����Ƿ�����hz
    # hz = 1
    # ��ʽ����ͷ����
    title_font_style_1 = Font(name='����', size=15, bold=True, color='FF000000')
    title_font_style_2 = Font(name='����', size=12, bold=True, color='FF000000')
    # ��ͨ������������
    plain_font_style = Font(name='����', size=12)
    cnter_alignment_style = Alignment(horizontal='center')
    border_style = Border(top=Side(border_style='thin', color='FF000000'),
                          left=Side(border_style='thin', color='FF000000'),
                          right=Side(border_style='thin', color='FF000000'),
                          bottom=Side(border_style='thin', color='FF000000'),
                          )
    left_alignment_style = Alignment(horizontal='left')
    right_alignment_style = Alignment(horizontal='right')

    # �������ݿ�ѭ������
    for yf1, bt in zip(data_list, title_list):
        # ��������ݿ����

        lt = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L',
              13: 'M',
              14: 'N', 15: 'O', 16: 'P', 16: 'Q'}
        for i in range(yf1.shape[1]):
            ws1.cell(row=r1, column=c1 + i).value = yf1.columns[i]
        for i in range(yf1.shape[1]):
            for j in range(yf1.shape[0]):
                ws1.cell(row=r1 + 1 + j, column=c1 + i).value = yf1.iloc[j, i]
        ws1['A' + str(r1 - 1)] = bt
        ws1.merge_cells('A' + str(r1 - 1) + ':' + lt[yf1.shape[1]] + str(r1 - 1))
        if hz == 1:
            ws1.merge_cells('A' + str(r1 + len(yf1)) + ':' + 'B' + str(r1 + len(yf1)))
        # �����ʽ����
        for row in ws1['A' + str(r1 - 1):lt[yf1.shape[1]] + str(r1 + yf1.shape[0])]:
            for c in row:
                c.font = plain_font_style
                c.border = border_style
                c.alignment = cnter_alignment_style
        # ������ʽ����-����-����ĳ��ı�
        for l in l_list:
            for row in ws1[l + str(r1 + 1):l + str(r1 + yf1.shape[0] - hz)]:
                for c in row:
                    c.alignment = left_alignment_style
        # ������ʽ����-����-��ֵ����
        for r in r_list:
            for row in ws1[r + str(r1 + 1):r + str(r1 + yf1.shape[0])]:
                for c in row:
                    c.alignment = right_alignment_style
        # ������ʽ����-��ͷ
        for row in ws1[r1 - 1]:
            row.font = title_font_style_1
        for row in ws1[r1]:
            row.font = title_font_style_2
        r1 = r1 + len(yf1) + 6

        # print(r1)
    # ������ȣ���������������
    for i, j in zip(c_w_list, c_w_z_list):
        ws1.column_dimensions[i].width = j
    pass


def get_data_from_sql(sql, cs_list):
    cursor = db.cursor()
    print('1')
    print(cs_list)
    cursor.execute(sql, cs_list)
    print('2')
    data = cursor.fetchall()
    print('3')
    data = pd.DataFrame(list(data))
    if data.empty == False:
        data.columns = [i[0] for i in cursor.description]
        return data
    else:
        return data


def set_columns_date_frame(df, select_rows, tail_rows):
    # �õ�һ�����������Ĵ�����ݿ�
    ff = df.values
    # ���ݿ��Ϊ���飺������������͹������ݿ�
    # ��һ����������
    if tail_rows == 0:
        f1 = pd.DataFrame(ff[0:select_rows, :]).fillna(method='ffill', axis=0)
        # �ڶ�����������������ݿ���
        f2 = pd.DataFrame(ff[select_rows:, :], columns=f1.values[-1])
    else:
        f1 = pd.DataFrame(ff[0:select_rows, :]).fillna(method='ffill', axis=0)
        # �ڶ�����������������ݿ���
        f2 = pd.DataFrame(ff[select_rows:-tail_rows, :], columns=f1.values[-1])
    return f2


db = pymysql.connect(
    host='172.19.116.150',
    user='zjzhzcuser',
    passwd='F4dus0ee',
    port=11806,
    charset='utf8',
    database='db_manage_overruns'
)


# db = pymysql.connect(host="192.168.1.229", port=3306, user='root', password='zcits123456',
#                          database='db_manage_overruns')


def qq():
    pass


def func1():
    # a=filedialog.asksaveasfilename()#�����ļ���

    # a =filedialog.asksaveasfile()#�ᴴ���ļ�

    # a =filedialog.askopenfilename()#�����ļ���

    # a =filedialog.askopenfile()#�����ļ�������

    # a =filedialog.askdirectory()#����Ŀ¼��

    a = filedialog.askopenfilenames()  # ���Է��ض���ļ���

    # a=filedialog.askopenfiles()#����ļ�������


def sr_1():
    global a1
    a1 = filedialog.askopenfilenames()  # ���Է��ض���ļ����ĵ�ַ��Ԫ��
    sr_e_1.set(a1)


def sr_2():
    c_e_1.set("")
    global a2
    a2 = filedialog.askopenfilename()  # ���Է��ض���ļ����ĵ�ַ��Ԫ��
    sr_e_2.set(a2)


def sr_3():
    # �м���0
    sr_e_5.set("")
    sr_e_6.set("")
    c_e_1.set("")
    global a3
    a3 = filedialog.askopenfilename()  # ���Է��ض���ļ����ĵ�ַ��Ԫ��
    sr_e_3.set(a3)


# def sr_4():
#     c_e_1.set("")
#     global a4
#     a4 = filedialog.askopenfilename()  # ���Է��ض���ļ����ĵ�ַ��Ԫ��
#     sr_e_4.set(a4)
def sr_5():
    # ʡ����0
    sr_d_1.set("")
    sr_d_2.set("")
    sr_e_3.set("")
    # �����0
    c_e_1.set("")
    global a5
    a5 = filedialog.askopenfilename()  # ���Է��ض���ļ����ĵ�ַ��Ԫ��
    sr_e_5.set(a5)


def sr_6():
    # ʡ����0
    sr_d_1.set("")
    sr_d_2.set("")
    sr_e_3.set("")
    # �����0
    c_e_1.set("")
    global a6
    a6 = filedialog.askopenfilename()  # ���Է��ض���ļ����ĵ�ַ��Ԫ��
    sr_e_6.set(a6)


def tjfx():
    E1.config(state='readonly')
    E2.config(state='readonly')
    E3.config(state='readonly')

    cs = sr_e_1.get()

    qishi_date = sr_d_1.get()
    jiezi_date = sr_d_2.get()

    dizhi = sc_e_1.get()
    print(cs, qishi_date, jiezi_date, dizhi)
    # # �㽭ʡ�л��ܱ�
    # sql_1 = "select b.city as '����',sum(case when a.total_weight>80 then 1 else 0 end) as '80������',\
    #         sum(case when a.total_weight>90 then 1 else 0 end) as '90������' \
    #         from t_bas_over_data_31 a \
    #         left join t_code_area b on b.county_code = a.area_county \
    #         left join t_sys_station c on c.station_code = a.out_station \
    #         where a.out_station_time >= %s and a.out_station_time <%s   and a.is_unusual= 0  and c.is_deleted = 0 and c.station_status = 0 and c.station_code IS NOT null group by b.city order by a.area_city"
    # # �㽭���ػ��ܱ�
    # sql_2 = "select b.city as '����',b.county as '����',sum(case when a.total_weight>80 then 1 else 0 end) as '80������',\
    #         sum(case when a.total_weight>90 then 1 else 0 end) as '90������' \
    #         from t_bas_over_data_31 a \
    #         left join t_code_area b on b.county_code = a.area_county \
    #         left join t_sys_station c on c.station_code = a.out_station \
    #         where b.city=%s and a.out_station_time >= %s and a.out_station_time <%s  and a.is_unusual= 0 and c.is_deleted = 0 and c.station_status = 0 and c.station_code IS NOT null \
    #         group by b.county order by a.area_county"
    # # �㽭����վ����ܱ�
    # sql_3 = "select b.city as '����',b.county as '����',a.site_name as 'վ������',sum(case when a.total_weight>80 then 1 else 0 end) as '80������',\
    #         sum(case when a.total_weight>90 then 1 else 0 end) as '90������' \
    #         from t_bas_over_data_31 a \
    #         left join t_code_area b on b.county_code = a.area_county \
    #         left join t_sys_station c on c.station_code = a.out_station \
    #         where b.city=%s and a.out_station_time >= %s and a.out_station_time <%s   and a.is_unusual= 0  and c.is_deleted = 0 and c.station_status = 0 and c.station_code IS NOT null \
    #         group by b.county,a.site_name order by a.area_county"
    # # �㽭����80����ϸ��
    # sql_80 = "SELECT b.city as '����',b.county as '����',a.site_name as 'վ������',a.out_station_time as '���ʱ��',a.car_no as '���ƺ���',a.total_weight as '����',a.limit_weight as '����',a.overrun as '����',a.axis as '����',a.overrun_rate as '������' ,a.record_code\
    #         FROM t_bas_over_data_31 a \
    #         LEFT JOIN t_code_area b ON b.county_code = a.area_county \
    #         left join t_sys_station c on c.station_code = a.out_station \
    #         WHERE a.total_weight >80 AND b.city = %s AND a.out_station_time >= %s AND a.out_station_time <%s  and a.is_unusual= 0 and c.is_deleted = 0 and c.station_status = 0 and c.station_code IS NOT null  ORDER BY a.area_county,a.site_name"
    # # �㽭����90����ϸ��
    # sql_90 = "SELECT record_code,b.city as '����',b.county as '����',a.site_name as 'վ������',a.out_station_time as '���ʱ��',a.car_no as '���ƺ���',a.total_weight as '����',a.limit_weight as '����',a.overrun as '����',a.axis as '����',a.overrun_rate as '������',a.record_code \
    #         FROM t_bas_over_data_31 a \
    #         LEFT JOIN t_code_area b ON b.county_code = a.area_county \
    #         left join t_sys_station c on c.station_code = a.out_station \
    #         WHERE a.total_weight >90 AND b.city = %s AND a.out_station_time >= %s AND a.out_station_time <%s  and a.is_unusual= 0 and c.is_deleted = 0 and c.station_status = 0 and c.station_code IS NOT null ORDER BY a.area_county,a.site_name"
    # # �ܱ�
    sql_808 = "SELECT record_code,b.city as '����',a.area_city as '���б���',b.county as '����',a.area_county as '���ر���',a.site_name as 'վ������',a.out_station_time as '���ʱ��',a.car_no as '���ƺ���',a.total_weight as '����',a.limit_weight as '����',a.overrun as '����',a.axis as '����',a.overrun_rate as '������' \
                FROM t_bas_over_data_31 a \
                LEFT JOIN t_code_area b ON b.county_code = a.area_county \
                left join t_sys_station c on c.station_code = a.out_station \
                WHERE a.total_weight >80 AND a.out_station_time >= %s AND a.out_station_time <%s   and a.is_unusual= 0 and c.is_deleted = 0 and c.station_status = 0 and c.station_code IS NOT null ORDER BY a.area_city,a.area_county,a.site_name"


    wc_list = []
    if cs != '' and dizhi != '':
        # ʱ��ת��
        try:
            # db = pymysql.connect(host='192.168.2.119',user='zjzhzcuser',passwd='F4dus0ee',
            #         port=3306,charset='utf8',database='db_manage_overruns')
            # # db = pymysql.connect(host="192.168.1.229", port=3306, user='root', password='zcits123456',
            # #                      database='db_manage_overruns')
            city_list = cs.split('-')
            # print('1')
            qishi_time = datetime.strptime(qishi_date, '%Y-%m-%d')
            # print('2')
            jiezi_time = datetime.strptime(jiezi_date, '%Y-%m-%d')
            # print('3')
            # jiezi_time = jiezzi_time + MonthEnd() + Day()
            # print('4')

            try:
                # ������
                df_big = get_data_from_sql(sql_808, [qishi_time, jiezi_time])
                # �����鿪ʼ��
                df_���ݻ��� = pd.read_excel(r"C:\Users\liu.wenjie\Desktop\�±�\202304\4���ݻ��ܱ�.xlsx")
                station_name = df_���ݻ���['վ������']
                df_big = df_big[df_big.loc[:, 'վ������'].isin(station_name)]

                b1_8 = df_big[df_big['����'] > 80].groupby(['����', '���б���', '����', '���ر���'])['����'].count().rename('80������')
                b1_9 = df_big[df_big['����'] > 90].groupby(['����', '���б���', '����', '���ر���'])['����'].count().rename('90������')
                b1 = pd.concat([b1_8, b1_9], axis=1).reset_index().sort_values(by=['���ر���'], ascending=True)

                # һ��ʡ�л��ܱ�
                b11 = b1[['����', '����', '���ر���', '80������', '90������']]
                b11.insert(0, '���', [i + 1 for i in range(len(b1))])

                b111 = pd.DataFrame({'���': ['����'], '80������': [b11['80������'].sum()], '90������': [b11['90������'].sum()]})
                b13 = pd.concat([b11, b111], axis=0)


                # �����鿪ʼ��


                b1_8 = df_big[df_big['����'] > 80].groupby(['����', '���б���'])['����'].count().rename('80������')
                b1_9 = df_big[df_big['����'] > 90].groupby(['����', '���б���'])['����'].count().rename('90������')
                b1 = pd.concat([b1_8, b1_9], axis=1).reset_index().sort_values(by=['���б���'], ascending=True)

                # һ��ʡ�л��ܱ�
                b1 = b1[['����', '80������', '90������']]
                b1.insert(0, '���', [i + 1 for i in range(len(b1))])

                b11 = pd.DataFrame({'���': ['����'], '80������': [b1['80������'].sum()], '90������': [b1['90������'].sum()]})
                b12 = pd.concat([b1, b11], axis=0)

                b80 = df_big[df_big['����'] > 80].sort_values(by=['���ر���', 'վ������'], ascending=True)
                b80 = b80[['����', '����', 'վ������', '���ʱ��', '���ƺ���', '����', '����', '����', '����', '������', 'record_code']]
                b80.insert(0, '���', [i + 1 for i in range(len(b80))])
                print(b80)
                with pd.ExcelWriter(r'C:\Users\liu.wenjie\Desktop\�±�\202304\80-90\80������ϸ.xlsx') as writer1:
                     b12.to_excel(writer1, sheet_name='���л���', index=False)
                     b13.to_excel(writer1, sheet_name='���ػ���', index=False)
                     b80.to_excel(writer1, sheet_name='������ϸ', index=False)



            except:
                wc_list.append('�¶Ȼ��ܱ�δ����')
            for j in city_list:
                try:
                    # �������ػ��ܱ�
                    df = df_big[df_big['����'] == j]
                    if df.empty == False:

                        b2_8 = df[df['����'] > 80].groupby(['����', '����', '���ر���'])['����'].count().rename('80������')
                        b2_9 = df[df['����'] > 90].groupby(['����', '����', '���ر���'])['����'].count().rename('90������')
                        b2 = pd.concat([b2_8, b2_9], axis=1).reset_index().sort_values(by=['���ر���'], ascending=True)
                        b2 = b2[['����', '����', '80������', '90������']]
                        b2.insert(0, '���', [i + 1 for i in range(len(b2))])
                        b22 = pd.DataFrame(
                            {'���': ['����'], '80������': [b2['80������'].sum()], '90������': [b2['90������'].sum()]})
                        b222 = pd.concat([b2, b22], axis=0)

                        b3_8 = df[df['����'] > 80].groupby(['����', '����', '���ر���', 'վ������'])['����'].count().rename('80������')
                        b3_9 = df[df['����'] > 90].groupby(['����', '����', '���ر���', 'վ������'])['����'].count().rename('90������')
                        b3 = pd.concat([b3_8, b3_9], axis=1).reset_index().sort_values(by=['���ر���', 'վ������'],
                                                                                       ascending=True)
                        b3 = b3[['����', '����', 'վ������', '80������', '90������']]
                        b3.insert(0, '���', [i + 1 for i in range(len(b3))])
                        b33 = pd.DataFrame(
                            {'���': ['����'], '80������': [b3['80������'].sum()], '90������': [b3['90������'].sum()]})
                        b333 = pd.concat([b3, b33], axis=0)

                        b80 = df[df['����'] >= 80].sort_values(by=['���ر���', 'վ������'], ascending=True)
                        b80 = b80[['����', '����', 'վ������', '���ʱ��', '���ƺ���', '����', '����', '����', '����', '������', 'record_code']]
                        b80.insert(0, '���', [i + 1 for i in range(len(b80))])

                        b90 = df[df['����'] > 90].sort_values(by=['���ر���', 'վ������'], ascending=True)
                        b90 = b90[['����', '����', 'վ������', '���ʱ��', '���ƺ���', '����', '����', '����', '����', '������', 'record_code']]
                        b90.insert(0, '���', [i + 1 for i in range(len(b90))])
                        # print(b90.head(6))
                        wb = Workbook()
                        ws1 = wb.create_sheet(j + '��80-90�ֻ��ܱ�')
                        ws2 = wb.create_sheet(j + '��80-90��վ����ܱ�')
                        ws3 = wb.create_sheet(j + '��80����ϸ��')
                        ws4 = wb.create_sheet(j + '��90����ϸ��')
                        # print('5')

                        excel_setting(ws1, [b222], [qishi_date + '��' + jiezi_date + '���ܱ�'], 1, [], [], [], [])

                        excel_setting(ws2, [b333], [qishi_date + '��' + jiezi_date + 'վ����ܱ�'], 1, ['D'], [], ['D'],
                                      [40])

                        excel_setting(ws3, [b80], [qishi_date + '��' + jiezi_date + '�ڼ�80����ϸ��'], 0, ['D'],
                                      ['G', 'H', 'I', 'J', 'K'], ['D', 'E'], [40, 24])

                        excel_setting(ws4, [b90], [qishi_date + '��' + jiezi_date + '�ڼ�90����ϸ��'], 0, ['D'],
                                      ['G', 'H', 'I', 'J', 'K'], ['D', 'E'], [40, 24])
                        wb.remove(wb.active)
                        wb.save(dizhi + '\\' + j + "��80-90�ֻ��ܼ���ϸ��" + ".xlsx")
                        wc_list.append(j + '��ر�������')
                    else:
                        wc_list.append(j + '����ر�')

                except:
                    wc_list.append(j + '��������')
            pd.DataFrame({'������': wc_list}).to_csv(
                dizhi + '\\' + qishi_date.split('-')[-2] + qishi_date.split('-')[-1] + '_' +
                jiezi_date.split('-')[-2] + jiezi_date.split('-')[-1] + '��־' + '.txt', sep='\t', index=False)
            c_e_1.set("�����Ѿ�������ɣ���鿴������")
            db.close()
        except:
            c_e_1.set("������ʽ����")

    else:
        c_e_1.set("����������©��")


def sc_10():
    a = filedialog.asksaveasfile()  # �ᴴ���ļ� �����ڵ����ļ������


def sc_1():
    c_e_1.set("")
    global dz
    dz = filedialog.askdirectory()  # ����Ŀ¼�� �����ڶ���ļ������
    sc_e_1.set(dz)


def czhs():
    # �ı�״̬�ͷ�
    E1.config(state='normal')
    E2.config(state='normal')
    E3.config(state='normal')
    # ������0
    sr_e_1.set("����-����-����-����-����-����-��-����-��ɽ-̨��-��ˮ")
    # ʡ����0
    # sr_d_1.set("")
    # sr_d_2.set("")
    # sr_e_3.set("")
    # �м���0
    # sr_e_5.set("")
    # sr_e_6.set("")
    # ��������0
    c_e_1.set("")


root = Tk()
root.resizable(width=False, height=False)
s1 = StringVar()
s1.set("")
s2 = StringVar()
s2.set("")
root.title('80-90���¶Ȼ���ר��')

# ����Ϊʡ������������


k3 = Frame(root, width=500, height=200)
k3.grid(row=10, column=0, rowspan=2)

# qzcs=LabelFrame(k3,text='����^-^',width=500,height=200)
# qzcs.grid(row=0,column=0,rowspan=2)
#
# sr_e_1=StringVar()
# sr_e_1.set("")
# Label(qzcs,text='������ʡ������',width=15).grid(row=0,column=0,rowspan=1,pady=2,padx=4)
# E1=Entry(qzcs,textvariable=sr_e_1,width=40)
# E1.grid(row=0,column=1,rowspan=1,pady=2,padx=4)
# #Button(k3,text="ѡ���ļ�",width=15,command=sr_1).grid(row=0,column=2,rowspan=1,pady=2,padx=4)
# Button(qzcs,text="��-��",width=15,command=czhs).grid(row=0,column=2,rowspan=1,pady=2,padx=4)
#
# sr_e_2=StringVar()
# sr_e_2.set("")
# Label(qzcs,text='ѡ���¶Ȼ��ܱ�',width=15).grid(row=1,column=0,rowspan=1,pady=2,padx=4)
# Entry(qzcs,textvariable=sr_e_2,width=40,state='readonly').grid(row=1,column=1,rowspan=1,pady=2,padx=4)
# Button(qzcs,text="ѡ���ļ�",width=15,command=sr_2).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

Frame(k3, width=550, height=5, bg='red').grid(row=3, column=0, rowspan=1, pady=2, padx=4)

sjbb = LabelFrame(k3, text='ʡ������', width=500, height=200)
sjbb.grid(row=4, column=0, rowspan=2)

sr_e_1 = StringVar()
sr_e_1.set("����-����-����-����-����-����-��-����-��ɽ-̨��-��ˮ")
Label(sjbb, text='�����б�', width=15).grid(row=0, column=0, rowspan=1, pady=2, padx=4)
E1 = Entry(sjbb, textvariable=sr_e_1, width=40)
E1.grid(row=0, column=1, rowspan=1, pady=2, padx=4)
# Button(k3,text="ѡ���ļ�",width=15,command=sr_1).grid(row=0,column=2,rowspan=1,pady=2,padx=4)
Button(sjbb, text="��-��", width=15, command=czhs).grid(row=0, column=2, rowspan=1, pady=2, padx=4)

sr_d_1 = StringVar()
sr_d_1.set("")
Label(sjbb, text='������ʼʱ��', width=15).grid(row=2, column=0, rowspan=1, pady=2, padx=4)
E2 = Entry(sjbb, textvariable=sr_d_1, width=40)
E2.grid(row=2, column=1, rowspan=1, pady=2, padx=4)
Label(sjbb, text='����2022-04', width=15).grid(row=2, column=2, rowspan=1, pady=2, padx=4)

sr_d_2 = StringVar()
sr_d_2.set("")
Label(sjbb, text='�����ֹʱ��', width=15).grid(row=4, column=0, rowspan=1, pady=2, padx=4)
E3 = Entry(sjbb, textvariable=sr_d_2, width=40)
E3.grid(row=4, column=1, rowspan=1, pady=2, padx=4)
Label(sjbb, text='����2022-05', width=15).grid(row=4, column=2, rowspan=1, pady=2, padx=4)

# sr_e_3=StringVar()
# sr_e_3.set("")
# Label(sjbb,text='ѡ�����ģ��',width=15).grid(row=8,column=0,rowspan=1,pady=2,padx=4)
# Entry(sjbb,textvariable=sr_e_3,width=40,state='readonly').grid(row=8,column=1,rowspan=1,pady=2,padx=4)
# Button(sjbb,text="ѡ���ļ�",width=15,command=sr_3).grid(row=8,column=2,rowspan=1,pady=2,padx=4)

# sr_e_4=StringVar()
# sr_e_4.set("")
# Label(sjbb,text='ѡ�������ļ�',width=15).grid(row=9,column=0,rowspan=1,pady=2,padx=4)
# Entry(sjbb,textvariable=sr_e_4,width=40,state='readonly').grid(row=9,column=1,rowspan=1,pady=2,padx=4)
# Button(sjbb,text="ѡ���ļ�",width=15,command=sr_4).grid(row=9,column=2,rowspan=1,pady=2,padx=4)
# Frame(k3,width=300,height=5,bg='red').grid(row=6,column=0,rowspan=1,pady=2,padx=4)
#
#
# dsbb=LabelFrame(k3,text='�м�����',width=500,height=200)
# dsbb.grid(row=8,column=0,rowspan=2)
#
# sr_e_6=StringVar()
# sr_e_6.set("")
# Label(dsbb,text='���г���100%����',width=15).grid(row=10,column=0,rowspan=1,pady=2,padx=4)
# Entry(dsbb,textvariable=sr_e_6,width=40,state='readonly').grid(row=10,column=1,rowspan=1,pady=2,padx=4)
# Button(dsbb,text="ѡ���ļ�",width=15,command=sr_6).grid(row=10,column=2,rowspan=1,pady=2,padx=4)
#
# sr_e_5=StringVar()
# sr_e_5.set("")
# Label(dsbb,text='ѡ������ģ��',width=15).grid(row=11,column=0,rowspan=1,pady=2,padx=4)
# Entry(dsbb,textvariable=sr_e_5,width=40,state='readonly').grid(row=11,column=1,rowspan=1,pady=2,padx=4)
# Button(dsbb,text="ѡ���ļ�",width=15,command=sr_5).grid(row=11,column=2,rowspan=1,pady=2,padx=4)

Frame(root, width=550, height=5, bg='red').grid(row=20, column=0, rowspan=1, pady=2, padx=4)

k6 = LabelFrame(root, text="Ϊ����Ҹ�Ŀ¼�ɣ�", width=500, height=200)
k6.grid(row=25, column=0, rowspan=2)

sc_e_1 = StringVar()
sc_e_1.set("")
Label(k6, text='�����', width=15).grid(row=0, column=0, rowspan=1, pady=2, padx=4)
Entry(k6, textvariable=sc_e_1, width=40, state='readonly').grid(row=0, column=1, rowspan=1, pady=2, padx=4)
Button(k6, text="ѡ��Ŀ¼", width=15, command=sc_1).grid(row=0, column=2, rowspan=1, pady=2, padx=4)
# Entry(k6,textvariable=sc_dwj_e_1,width=40).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
# Button(k6,text="�������",width=15,command=sc_dwj).grid(row=1,column=0,rowspan=1,pady=2,padx=4)
# Button(k6,text="������",width=15,command=sc_dwj).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

Frame(root, width=550, height=5, bg='red').grid(row=27, column=0, rowspan=1, pady=2, padx=4)

k8 = LabelFrame(root, text="���з�������ȡ���", width=500, height=200)
k8.grid(row=30, column=0, rowspan=2)

c_e_1 = StringVar()
c_e_1.set("")
Label(k8, text='������', width=15).grid(row=0, column=0, rowspan=1, pady=2, padx=4)
Entry(k8, textvariable=c_e_1, width=40, state='readonly').grid(row=0, column=1, rowspan=1, pady=2, padx=4)
Button(k8, text="�ύ����", foreground="green", width=15, command=tjfx).grid(row=0, column=2, rowspan=1, pady=2, padx=4)

# ����Ϊ�����������ɿ��
# dsbb=LabelFrame(root,text='���б������',width=500,height=200)
# dsbb.grid(row=2,column=0,rowspan=2)
#
# l3=LabelFrame(dsbb,text="������Ҫ�����Ĳ��� ^-^",width=500,height=200,bg='Red')
# l3.grid(row=200,column=0,rowspan=2)
#
#
# sr_e_1=StringVar()
# sr_e_1.set("")
# Label(l3,text='������ʡ������',width=15).grid(row=0,column=0,rowspan=1,pady=2,padx=4)
# Entry(l3,textvariable=sr_e_1,width=40).grid(row=0,column=1,rowspan=1,pady=2,padx=4)
#
# sr_e_6=StringVar()
# sr_e_6.set("")
# Label(l3,text='���г���100%����',width=15).grid(row=10,column=0,rowspan=1,pady=2,padx=4)
# Entry(l3,textvariable=sr_e_6,width=40,state='readonly').grid(row=10,column=1,rowspan=1,pady=2,padx=4)
# Button(l3,text="ѡ���ļ�",width=15,command=sr_6).grid(row=10,column=2,rowspan=1,pady=2,padx=4)
#
# sr_e_5=StringVar()
# sr_e_5.set("")
# Label(l3,text='ѡ��������Ӧģ��',width=15).grid(row=11,column=0,rowspan=1,pady=2,padx=4)
# Entry(l3,textvariable=sr_e_5,width=40,state='readonly').grid(row=11,column=1,rowspan=1,pady=2,padx=4)
# Button(l3,text="ѡ���ļ�",width=15,command=sr_5).grid(row=11,column=2,rowspan=1,pady=2,padx=4)

# Text(k8,width=40,height=10).grid(row=1,column=0,rowspan=1,columnspan=1,pady=2,padx=4)
# Entry(k8,textvariable=sc_dwj_e_1,width=40).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
# Button(k8,text="���Ԥ��",width=15).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
# Button(k8,text="������",width=15,command=sc_dwj).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

root.mainloop()