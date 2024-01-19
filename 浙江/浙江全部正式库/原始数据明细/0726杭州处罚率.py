import pymysql
import pandas as pd
import  time
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment

def excel_setting(ws1, data_list, title_list, hz, l_list, r_list, c_w_list, c_w_z_list):
    # 起始行号设置
    r1 = 2
    c1 = 1
    # 是否有汇总决定是否启用hz
    # hz = 1
    # 格式标题头设置
    title_font_style_1 = Font(name='宋体', size=15, bold=True, color='FF000000')
    title_font_style_2 = Font(name='宋体', size=12, bold=True, color='FF000000')
    # 普通内容字体设置
    plain_font_style = Font(name='宋体', size=12)
    cnter_alignment_style = Alignment(horizontal='center',vertical='center',wrap_text=False)
    border_style = Border(top=Side(border_style='thin', color='FF000000'),
                          left=Side(border_style='thin', color='FF000000'),
                          right=Side(border_style='thin', color='FF000000'),
                          bottom=Side(border_style='thin', color='FF000000'),
                          )
    left_alignment_style = Alignment(horizontal='left',vertical='center', wrap_text=False)
    right_alignment_style = Alignment(horizontal='right',vertical='center', wrap_text=False)
    center_alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # 单个数据块循环插入
    for yf1, bt in zip(data_list, title_list):
        # 单块的数据块插入

        lt = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J',
              11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T',
              21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z'}
        for i in range(yf1.shape[1]):
            ws1.cell(row=r1, column=c1 + i).value = yf1.columns[i]
        for i in range(yf1.shape[1]):
            for j in range(yf1.shape[0]):
                ws1.cell(row=r1 + 1 + j, column=c1 + i).value = yf1.iloc[j, i]
        ws1['A' + str(r1 - 1)] = bt
        #ws1.merge_cells('A' + str(r1 - 1) + ':' + lt[yf1.shape[1]] + str(r1 - 1))

        #待启用合并模版
        # ws1['E' + str(r1 + len(yf1['理应在线天数'].dropna() + 1))] = '站点维修'
        # ws1.merge_cells('E' + str(r1 + len(yf1['理应在线天数'].dropna() + 1)) + ':' + 'P' + str(r1 + len(yf1['理应在线天数'])))

        if hz == 1:
            ws1.merge_cells('A' + str(r1 + len(yf1)) + ':' + 'B' + str(r1 + len(yf1)))
        # 整体格式设置
        for row in ws1['A' + str(r1 - 1):lt[yf1.shape[1]] + str(r1 + yf1.shape[0])]:
            for c in row:
                c.font = plain_font_style
                c.border = border_style
                c.alignment = cnter_alignment_style
        # 特殊块格式设置-靠左-特殊的长文本
        for l in l_list:
            for row in ws1[l + str(r1 + 1):l + str(r1 + yf1.shape[0] - hz)]:
                for c in row:
                    c.alignment = left_alignment_style
        # 特殊块格式设置-靠右-数值靠右
        for r in r_list:
            for row in ws1[r + str(r1 + 1):r + str(r1 + yf1.shape[0])]:
                for c in row:
                    c.alignment = right_alignment_style
        # 特殊块格式设置-表头
        for row in ws1[r1 - 1]:
            row.font = title_font_style_1
        for row in ws1[r1]:
            row.font = title_font_style_2
        #待启用的着色模块
        # hong = yf1[yf1['实际在线天数'] <= 10]['序号'].tolist()
        # huang = yf1[(yf1['实际在线天数'] > 10) & (yf1['货车数'] < 500)]['序号'].tolist()
        # for i in hong:
        #     ws1['F' + str(r1 + i)].fill = PatternFill(fill_type='solid', fgColor='FFFF2100')
        #     print('F' + str(r1 + i))
        # for i in huang:
        #     ws1['H' + str(r1 + i)].fill = PatternFill(fill_type='solid', fgColor='FFFFFF00')


        r1 = r1 + len(yf1) + 6

        # print(r1)
    # 列名宽度，具体情况具体调整
    for i, j in zip(c_w_list, c_w_z_list):
        ws1.column_dimensions[i].width = j
    pass



db = pymysql.connect(
    host='192.168.2.119',
    user='zjzhzcuser',
    passwd='F4dus0ee',
    port=3306,
    charset='utf8',
    database='db_manage_overruns'
    )


def get_df_from_db(sql):
    cursor = db.cursor()  # 使用cursor()方法获取用于执行SQL语句的游标
    cursor.execute(sql)  # 执行SQL语句
    """
    使用fetchall函数以元组形式返回所有查询结果并打印出来
    fetchone()返回第一行，fetchmany(n)返回前n行
    游标执行一次后则定位在当前操作行，下一次操作从当前操作行开始
    """
    data = cursor.fetchall()
    # 下面为将获取的数据转化为dataframe格式
    columnDes = cursor.description  # 获取连接对象的描述信息
    columnNames = [columnDes[i][0] for i in range(len(columnDes))]  # 获取列名
    df = pd.DataFrame([list(i) for i in data], columns=columnNames)  # 得到的data为二维元组，逐行取出，转化为列表，再转化为df
    """
    使用完成之后需关闭游标和数据库连接，减少资源占用,cursor.close(),db.close()
    db.commit()若对数据库进行了修改，需进行提交之后再关闭
    """
    return df


'''获取满足惩罚条件的非现场点位'''
sql1 = "select b.city,b.county,a.site_name,a.status,a.car_no,a.car_no_color,\
        a.total_weight,a.overrun,a.overrun_rate,a.axis,a.out_station_time,\
        a.vehicle_city,a.vehicle_county,a.car_holder_addr,t.punish_money \
        from t_bas_over_data_collection_31 as a \
        left join t_code_area as b \
        on a.area_county=b.county_code \
        left join t_case_sign_result as t \
        on a.record_code=t.record_code \
        where a.out_station_time>='2021-07-15 00:00:00' and a.out_station_time<'2022-04-01 00:00:00' \
        and a.status in(4,6,12,13) \
        and b.city='杭州'"
sql2="select * from t_bas_over_data_collection_31 where car_no='浙G87068' "
#df= get_df_from_db(sql1)
df= get_df_from_db(sql2)

df.to_excel(r"C:\Users\Administrator\Desktop\lll\杭州车牌.xlsx",index=False)
print(df)