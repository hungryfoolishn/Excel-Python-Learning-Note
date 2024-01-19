import pymysql
import pandas as pd
import  time
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment

def excel_setting(ws1, data_list, title_list, hz, l_list, r_list, c_w_list, c_w_z_list):
    # 起始行号设置
    r1 = 2
    c1 = 1
    # 是否有汇总决定是否启用hz
    # hz = 1
    # 格式标题头设置
    title_font_style_1 = Font(name='宋体', size=15, bold=True, color='FF000000')
    title_font_style_2 = Font(name='宋体', size=12, bold=True, color='FF000000')
    # 普通内容字体设置
    plain_font_style = Font(name='宋体', size=12)
    cnter_alignment_style = Alignment(horizontal='center',vertical='center',wrap_text=False)
    border_style = Border(top=Side(border_style='thin', color='FF000000'),
                          left=Side(border_style='thin', color='FF000000'),
                          right=Side(border_style='thin', color='FF000000'),
                          bottom=Side(border_style='thin', color='FF000000'),
                          )
    left_alignment_style = Alignment(horizontal='left',vertical='center', wrap_text=False)
    right_alignment_style = Alignment(horizontal='right',vertical='center', wrap_text=False)
    center_alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # 单个数据块循环插入
    for yf1, bt in zip(data_list, title_list):
        # 单块的数据块插入

        lt = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J',
              11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T',
              21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z'}
        for i in range(yf1.shape[1]):
            ws1.cell(row=r1, column=c1 + i).value = yf1.columns[i]
        for i in range(yf1.shape[1]):
            for j in range(yf1.shape[0]):
                ws1.cell(row=r1 + 1 + j, column=c1 + i).value = yf1.iloc[j, i]
        ws1['A' + str(r1 - 1)] = bt
        #ws1.merge_cells('A' + str(r1 - 1) + ':' + lt[yf1.shape[1]] + str(r1 - 1))

        #待启用合并模版
        # ws1['E' + str(r1 + len(yf1['理应在线天数'].dropna() + 1))] = '站点维修'
        # ws1.merge_cells('E' + str(r1 + len(yf1['理应在线天数'].dropna() + 1)) + ':' + 'P' + str(r1 + len(yf1['理应在线天数'])))

        if hz == 1:
            ws1.merge_cells('A' + str(r1 + len(yf1)) + ':' + 'B' + str(r1 + len(yf1)))
        # 整体格式设置
        for row in ws1['A' + str(r1 - 1):lt[yf1.shape[1]] + str(r1 + yf1.shape[0])]:
            for c in row:
                c.font = plain_font_style
                c.border = border_style
                c.alignment = cnter_alignment_style
        # 特殊块格式设置-靠左-特殊的长文本
        for l in l_list:
            for row in ws1[l + str(r1 + 1):l + str(r1 + yf1.shape[0] - hz)]:
                for c in row:
                    c.alignment = left_alignment_style
        # 特殊块格式设置-靠右-数值靠右
        for r in r_list:
            for row in ws1[r + str(r1 + 1):r + str(r1 + yf1.shape[0])]:
                for c in row:
                    c.alignment = right_alignment_style
        # 特殊块格式设置-表头
        for row in ws1[r1 - 1]:
            row.font = title_font_style_1
        for row in ws1[r1]:
            row.font = title_font_style_2
        #待启用的着色模块
        # hong = yf1[yf1['实际在线天数'] <= 10]['序号'].tolist()
        # huang = yf1[(yf1['实际在线天数'] > 10) & (yf1['货车数'] < 500)]['序号'].tolist()
        # for i in hong:
        #     ws1['F' + str(r1 + i)].fill = PatternFill(fill_type='solid', fgColor='FFFF2100')
        #     print('F' + str(r1 + i))
        # for i in huang:
        #     ws1['H' + str(r1 + i)].fill = PatternFill(fill_type='solid', fgColor='FFFFFF00')


        r1 = r1 + len(yf1) + 6

        # print(r1)
    # 列名宽度，具体情况具体调整
    for i, j in zip(c_w_list, c_w_z_list):
        ws1.column_dimensions[i].width = j
    pass



db = pymysql.connect(
    host='192.168.2.119',
    user='zjzhzcuser',
    passwd='F4dus0ee',
    port=3306,
    charset='utf8',
    database='db_manage_overruns'
    )


def get_df_from_db(sql):
    cursor = db.cursor()  # 使用cursor()方法获取用于执行SQL语句的游标
    cursor.execute(sql)  # 执行SQL语句
    """
    使用fetchall函数以元组形式返回所有查询结果并打印出来
    fetchone()返回第一行，fetchmany(n)返回前n行
    游标执行一次后则定位在当前操作行，下一次操作从当前操作行开始
    """
    data = cursor.fetchall()
    # 下面为将获取的数据转化为dataframe格式
    columnDes = cursor.description  # 获取连接对象的描述信息
    columnNames = [columnDes[i][0] for i in range(len(columnDes))]  # 获取列名
    df = pd.DataFrame([list(i) for i in data], columns=columnNames)  # 得到的data为二维元组，逐行取出，转化为列表，再转化为df
    """
    使用完成之后需关闭游标和数据库连接，减少资源占用,cursor.close(),db.close()
    db.commit()若对数据库进行了修改，需进行提交之后再关闭
    """
    return df


'''获取满足惩罚条件的非现场点位'''
sql1 = "SELECT station_name FROM t_sys_station where is_match_station =1 and station_type = 31"
t_station_name= get_df_from_db(sql1)
station_name=t_station_name[['station_name']]
station_name.to_excel(r"C:\Users\Administrator\Desktop\tttt\满足证据站点明细表.xlsx",index=False)
print(t_station_name)


'''获取超限100%明细数据'''
cs ='浙江'
h = input("请输入拉出表路径(C:/Users/Administrator/Desktop/系统拉出表/省级月报)：")

df_数据汇总 = pd.read_excel("{}/月数据汇总表.xls".format(h))
df_数据汇总.columns = df_数据汇总.iloc[2]
df_数据汇总 = df_数据汇总.iloc[3:].reset_index(drop=False)

q = input("请输入起始年月日(比如2022-04-01)：")
s = input("请输入截止年月日(比如2022-05-01)：")
""" 引入原始表 """
start = time.time()
print('请等待2分钟左右......')
sql = "SELECT * FROM t_sys_station"
t_sys_station = get_df_from_db(sql)
U_汇总_站点表 = pd.merge(df_数据汇总, t_sys_station, left_on='站点名称', right_on='station_name', how='left')
station_code = U_汇总_站点表['station_code']
sql = "SELECT * FROM t_code_area  "
t_code_area = get_df_from_db(sql)
sql = "SELECT * FROM t_bas_over_data_31 where is_unusual = 0 and out_station_time >= '{} 00:00:00' and  out_station_time <'{} 00:00:00' and allow is null  and is_unusual=0 and overrun_rate>=100 and  total_weight<=100 ".format(
    q, s)
t_bas_over_data_31 = get_df_from_db(sql)
end = time.time()
time = end - start
print(time)

"""超限1000全省明细"""

U_过车_区域表 = pd.merge(t_bas_over_data_31, t_code_area, left_on='area_county', right_on='county_code', how='left')
U_过车_站点表 = pd.merge(U_过车_区域表, t_sys_station, left_on='out_station', right_on='station_code', how='left')
U_过车_站点表 = U_过车_站点表[U_过车_站点表.loc[:, 'out_station'].isin(station_code)]

i = {
    "330100": 1.1,
    "330200": 1,
    "330300": 1,
    "330400": 1.1,
    "330500": 1,
    "330600": 1,
    "330700": 1,
    "330800": 1.1,
    "330900": 1.1,
    "331000": 1.1,
    "331100": 1.1,
    "330122": 1.2,
    "330183": 1.21,
    "330523": 1.03,
    "330603": 1.1,
    "330604": 1.2,
    "330624": 1.2,
    "330681": 1.1,
    "330703": 1.2,
    "330782": 1.1
}
for item in i.items():
    key = item[0]
    value = item[1]
    U_过车_站点表.loc[((U_过车_站点表['area_county_x'] == key) | (U_过车_站点表['area_city_x'] == key)) & (
                U_过车_站点表['total_weight'] < 100), 'vehicle_brand'] = U_过车_站点表['limit_weight'].map(
        lambda x: float(x) * value).round(4)
U_过车_站点表['limit_weight'] = U_过车_站点表['limit_weight'].astype('float')
U_过车_站点表['total_weight'] = U_过车_站点表['total_weight'].astype('float')
U_过车_站点表['vehicle_brand'] = U_过车_站点表['vehicle_brand'].astype('float')
U_过车_站点表['超限率100%'] = U_过车_站点表.apply(lambda x: (x['total_weight'] - x['vehicle_brand']) * 100 / x['vehicle_brand'],
                                     axis=1).round(2)
U_过车_站点表 = U_过车_站点表[(U_过车_站点表['超限率100%'] >= 100)
]
U_过车_站点表 = pd.DataFrame(U_过车_站点表,
                        columns=["city", "county", "car_no", "out_station_time", "axis", "total_weight", "overrun",
                                 "overrun_rate", "site_name", "status_x", "is_collect", "is_unusual", "record_code",
                                 "photo1", "photo2", "photo3", "vedio"])

U_过车_站点表 = U_过车_站点表.sort_values(by=['city', 'county', 'out_station_time'], ascending=False).reset_index(drop=True)
U_过车_站点表 = U_过车_站点表.drop_duplicates(subset=['record_code'])
U_过车_站点表.loc[U_过车_站点表['status_x'] == 1, 'status_x'] = '待初审'
U_过车_站点表.loc[U_过车_站点表['status_x'] == 2, 'status_x'] = '待审核'
U_过车_站点表.loc[U_过车_站点表['status_x'] == 9, 'status_x'] = '判定不处理'
U_过车_站点表.loc[U_过车_站点表['status_x'] == 15, 'status_x'] = '初审不通过'
U_过车_站点表.loc[U_过车_站点表['is_collect'] == 1, 'is_collect'] = '满足'
U_过车_站点表.loc[U_过车_站点表['is_collect'] == 0, 'is_collect'] = '不满足'

"""车牌处理"""
U_过车_站点表['car_no'].fillna('无牌', inplace=True)
U_过车_站点表['字节数'] = U_过车_站点表['car_no'].str.len()
U_过车_站点表.loc[U_过车_站点表['字节数'] <= 5, 'car_no'] = '无牌'
U_过车_站点表.loc[~((U_过车_站点表['字节数'] <= 5) & (U_过车_站点表['is_collect'] == '满足')), 'photo1'] = ''
U_过车_站点表.loc[~((U_过车_站点表['字节数'] <= 5) & (U_过车_站点表['is_collect'] == '满足')), 'photo2'] = ''
U_过车_站点表.loc[~((U_过车_站点表['字节数'] <= 5) & (U_过车_站点表['is_collect'] == '满足')), 'photo3'] = ''
U_过车_站点表.loc[~((U_过车_站点表['字节数'] <= 5) & (U_过车_站点表['is_collect'] == '满足')), 'vedio'] = ''

U_过车_站点表.insert(0,'序号',[i+1 for i in range(len(U_过车_站点表))])
U_过车_站点表.columns=['序号', '地市', '区县', '车牌号', '检测时间', '车轴', '总重', '超限', '超限率', '站点名称', '状态',
       '证据满足情况', '是否故意遮挡车牌', '流水号','photo1', 'photo2', 'photo3', 'vedio', '字节数']
U_过车_站点表['是否故意遮挡车牌']=''
'''明细表放入工作表中'''
wb1=Workbook()
ws1=wb1.active
excel_setting(ws1,[U_过车_站点表],[123], 0,['O','P','Q','M','N'], [], [], [])
ws1.delete_rows(1)
wb1.save(r"C:\Users\Administrator\Desktop\tttt\超限100%明细表.xlsx")

#U_过车_站点表.to_excel(r"C:\Users\Administrator\Desktop\lll\超限100%明细表.xlsx",index=False)
print(U_过车_站点表.head())

'''百吨王明细表'''

""" 引入原始表 """

sql = "SELECT * FROM t_sys_station "
t_sys_station = get_df_from_db(sql)
sql = "SELECT * FROM t_code_area  "
t_code_area = get_df_from_db(sql)

sql = "SELECT * FROM t_bas_over_data_31 where is_unusual = 0 and out_station_time >= '{} 00:00:00' and  out_station_time <'{} 00:00:00' and allow is null  and is_unusual=0 and area_province=330000 and total_weight >= 100 ".format(q,s)
t_bas_over_data_31 = get_df_from_db(sql)


U_过车_区域表 = pd.merge(t_bas_over_data_31, t_code_area, left_on='area_county', right_on='county_code', how='left')
U_过车_站点表 = pd.merge(U_过车_区域表, t_sys_station, left_on='out_station', right_on='station_code', how='left')
U_过车_站点表 = U_过车_站点表[(U_过车_站点表['station_status'] == 0)]
U_过车_站点表 = U_过车_站点表[~U_过车_站点表.loc[:, 'site_name'].isin(i)]

U_过车_站点表 = U_过车_站点表[(U_过车_站点表['province'] == '浙江')
]
U_过车_站点表 = pd.DataFrame(U_过车_站点表,
                        columns=["city", "county", "car_no", "out_station_time", "axis", "total_weight", "overrun",
                                 "overrun_rate", "site_name", "status_x", "is_collect", "is_unusual", "record_code",
                                 "photo1", "photo2", "photo3", "vedio"])
U_过车_站点表 = U_过车_站点表.sort_values(by=['city', 'county', 'out_station_time']).reset_index(drop=True)
U_过车_站点表 = U_过车_站点表.drop_duplicates(subset=['record_code'])
#百吨王车辆统计 = U_过车_站点表.groupby(['city','county'])['record_code'].count().reset_index(name='百吨王（辆次）')
'''状态替换'''
U_过车_站点表.loc[U_过车_站点表['status_x'] == 1, 'status_x'] = '待初审'
U_过车_站点表.loc[U_过车_站点表['status_x'] == 2, 'status_x'] = '待审核'
U_过车_站点表.loc[U_过车_站点表['status_x'] == 9, 'status_x'] = '判定不处理'
U_过车_站点表.loc[U_过车_站点表['status_x'] == 15, 'status_x'] = '初审不通过'
U_过车_站点表.loc[U_过车_站点表['is_collect'] == 1, 'is_collect'] = '满足'
U_过车_站点表.loc[U_过车_站点表['is_collect'] == 0, 'is_collect'] = '不满足'

U_过车_站点表.insert(0,'序号',[i+1 for i in range(len(U_过车_站点表))])
U_过车_站点表.columns=['序号', '地市', '区县', '车牌号', '检测时间', '车轴', '总重', '超限', '超限率', '站点名称', '状态',
       '证据满足情况', '超限记录核实', '流水号', 'photo1', 'photo2', 'photo3', 'vedio']
U_过车_站点表['超限记录核实']=''
U_过车_站点表.insert(len(U_过车_站点表.columns)-5,'是否遮挡车牌','')

print(U_过车_站点表.columns)
wb2=Workbook()
ws2=wb2.active
excel_setting(ws2,[U_过车_站点表],[123], 0,['O','P','Q','R','S'], [], [], [])
ws2.delete_rows(1)
wb2.save(r"C:\Users\Administrator\Desktop\tttt\百吨王明细表.xlsx")
#U_过车_站点表.to_excel(r"C:\Users\Administrator\Desktop\lll\百吨王明细表.xlsx",index=False)
print(U_过车_站点表)

'''案件处理累计表'''
""" 引入原始表 """
q='2022-01-01'
#q = input("请输入起始年月日(比如2022-01-01)：")
#s = input("请输入截止年月日(比如2022-05-01)：")
#cs = input("请输入省市名：")

# start=time.time()
sql = "SELECT * FROM t_code_area"
t_code_area=get_df_from_db(sql)

sql = "SELECT * FROM t_bas_over_data_collection_31 where valid_time >= '{} 00:00:00' and valid_time < '{} 00:00:00'".format(q,s)
t_bas_over_data_collection_31=get_df_from_db(sql)
sql = "SELECT * FROM t_case_sign_result where close_case_time >= '{} 00:00:00' and close_case_time < '{} 00:00:00'".format(q,s)
t_case_sign_result=get_df_from_db(sql)

q=q.replace('-','')
s=s.replace('-','')
sql = "SELECT * FROM t_bas_over_data_collection_makecopy where insert_time >={} and insert_time <{}".format(q,s)
t_bas_over_data_collection_makecopy=get_df_from_db(sql)

sql = "SELECT * FROM t_sys_station "
t_sys_station=get_df_from_db(sql)

# end =time.time()
# print(end-start)

U_检测点_区域表 = pd.merge(t_bas_over_data_collection_31,t_code_area,left_on='area_county',right_on='county_code',how='left')

U_案件处罚_区域表= pd.merge(t_case_sign_result,t_code_area,left_on='area_county',right_on='county_code',how='left')

U_外省抄告_区域表 = pd.merge(t_bas_over_data_collection_makecopy,t_code_area,left_on='area_county',right_on='county_code',how='left')

T_入库查询=U_检测点_区域表.loc[(U_检测点_区域表['law_judgment']=="1")]


'''滴滴滴'''
T_现场处罚 = U_案件处罚_区域表[(U_案件处罚_区域表.record_type == 99)
                        & (U_案件处罚_区域表.insert_type == 5)]

T_非现场处罚 = U_案件处罚_区域表[(U_案件处罚_区域表.record_type == 31)
                     & (U_案件处罚_区域表.insert_type == 1)
                     & (U_案件处罚_区域表.data_source == 1)
                     & (U_案件处罚_区域表.case_type == 1)]

T_外省抄告 = U_外省抄告_区域表


T_入库查询 = T_入库查询.sort_values(['area_city','area_county'], ascending=False)
print('---1---')
print(T_入库查询.columns)
T_入库数 = T_入库查询.groupby(['area_city','city','area_county','county']).count()

T_入库数 = T_入库数.loc[:, ['id_x']]
T_入库数.rename(columns={'id_x': '入库数(系统)'}, inplace=True)

T_现场处罚 = T_现场处罚.drop_duplicates(['CASE_NUM'])
T_现场处罚计数 = T_现场处罚.groupby(['city','county'])['CASE_NUM'].count()
T_现场处罚计数 = T_现场处罚计数.to_frame()
T_现场处罚计数.rename(columns={'CASE_NUM': '现场处罚(系统)'}, inplace=True)

T_非现场处罚 = T_非现场处罚.drop_duplicates(['CASE_NUM'])
T_非现场处罚计数 = T_非现场处罚.groupby(['city','county'])['CASE_NUM'].count()
T_非现场处罚计数 = T_非现场处罚计数.to_frame()
T_非现场处罚计数.rename(columns={'CASE_NUM': '非现场处罚(系统)'}, inplace=True)

T_外省抄告 = T_外省抄告.groupby(['city','county']).count()
T_外省抄告计数 = T_外省抄告.loc[:, ['id_x']]
T_外省抄告计数.rename(columns={'id_x': '外省抄告(系统)'}, inplace=True)
W_案件统计 = pd.merge(T_入库数, T_现场处罚计数, left_on=['city','county'], right_on=['city','county'], how='outer')
W_案件统计 = pd.merge(W_案件统计, T_非现场处罚计数, left_on=['city','county'], right_on=['city','county'], how='outer')
W_案件统计 = pd.merge(W_案件统计, T_外省抄告计数, left_on=['city','county'], right_on=['city','county'], how='outer')
W_案件统计[['外省抄告(系统)']] = W_案件统计[['外省抄告(系统)']].apply(np.int64)
W_案件统计.loc[W_案件统计['外省抄告(系统)'] < 0, '外省抄告(系统)'] = 0

'''其他市案件统计情况'''
W_案件统计 =W_案件统计.reset_index()
#W_案件统计= W_案件统计.drop(index=(W_案件统计.loc[(W_案件统计['county'] == '市辖区')].index))
W_案件统计.to_excel(r"C:\Users\Administrator\Desktop\tttt\案件wei明细表.xlsx",index=False)
print(W_案件统计.head(30))





