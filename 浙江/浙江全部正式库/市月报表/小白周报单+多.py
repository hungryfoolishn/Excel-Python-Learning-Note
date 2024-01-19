from tkinter import *
from tkinter import filedialog
from tkinter import ttk
import pandas as pd
import pymysql
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment
from pandas.tseries.offsets import Day

cc=pd.DataFrame()
cc['city']=['杭州', '杭州', '杭州', '杭州', '杭州', '杭州', '杭州', '杭州', '杭州', '杭州', '杭州', '杭州', '杭州', '湖州', '湖州', '湖州', '湖州', '湖州', '嘉兴', '嘉兴', '嘉兴', '嘉兴', '嘉兴', '嘉兴', '嘉兴', '金华', '金华', '金华', '金华', '金华', '金华', '金华', '金华', '金华', '丽水', '丽水', '丽水', '丽水', '丽水', '丽水', '丽水', '丽水', '丽水', '宁波', '宁波', '宁波', '宁波', '宁波', '宁波', '宁波', '宁波', '宁波', '宁波', '宁波', '衢州', '衢州', '衢州', '衢州', '衢州', '衢州', '绍兴', '绍兴', '绍兴', '绍兴', '绍兴', '绍兴', '台州', '台州', '台州', '台州', '台州', '台州', '台州', '台州', '台州', '温州', '温州', '温州', '温州', '温州', '温州', '温州', '温州', '温州', '温州', '温州', '舟山', '舟山', '舟山', '舟山', '杭州', '湖州', '嘉兴', '金华', '丽水', '宁波', '衢州', '绍兴', '台州', '温州', '舟山', '杭州', '杭州', '宁波', '杭州', '绍兴', '湖州', '杭州']
cc['county']=['滨江', '淳安', '富阳', '拱墅', '建德', '江干', '临安', '上城', '桐庐', '西湖', '下城', '萧山', '余杭', '安吉', '德清', '南浔', '吴兴', '长兴', '海宁', '海盐', '嘉善', '南湖', '平湖', '桐乡', '秀洲', '东阳', '金东', '兰溪', '磐安', '浦江', '武义', '婺城', '义乌', '永康', '缙云', '景宁', '莲都', '龙泉', '青田', '庆元', '松阳', '遂昌', '云和', '北仑', '慈溪', '奉化', '海曙', '江北', '江东', '宁海', '象山', '鄞州', '余姚', '镇海', '常山', '江山', '开化', '柯城', '龙游', '衢江', '柯桥', '上虞', '嵊州', '新昌', '越城', '诸暨', '黄岩', '椒江', '临海', '路桥', '三门', '天台', '温岭', '仙居', '玉环', '苍南', '洞头', '乐清', '龙湾', '鹿城', '瓯海', '平阳', '瑞安', '泰顺', '文成', '永嘉', '岱山', '定海', '普陀', '嵊泗', '市辖区', '市辖区', '市辖区', '市辖区', '市辖区', '市辖区', '市辖区', '市辖区', '市辖区', '市辖区', '市辖区', '下沙', '钱塘新区', '东钱湖', '临平', '绍兴县', '南太湖新区', '建德']
dd=pd.DataFrame()
dd['年月']=['2022-01', '2022-01', '2022-01', '2022-01', '2022-01', '2022-02', '2022-02', '2022-02', '2022-02', '2022-03', '2022-03', '2022-03', '2022-03', '2022-04', '2022-04', '2022-04', '2022-04', '2022-05', '2022-05', '2022-05', '2022-05', '2022-05', '2022-06', '2022-06', '2022-06', '2022-06', '2022-07', '2022-07', '2022-07', '2022-07', '2022-08', '2022-08', '2022-08', '2022-08', '2022-08', '2022-09', '2022-09', '2022-09', '2022-09', '2022-10', '2022-10', '2022-10', '2022-10', '2022-10', '2022-11', '2022-11', '2022-11', '2022-11', '2022-12', '2022-12', '2022-12', '2022-12', '2023-01', '2023-01', '2023-01', '2023-01', '2023-01', '2023-02', '2023-02', '2023-02', '2023-02', '2023-03', '2023-03', '2023-03', '2023-03', '2023-04', '2023-04', '2023-04', '2023-04']
dd['日期']=['2022-01-03', '2022-01-10', '2022-01-17', '2022-01-24', '2022-01-31', '2022-02-07', '2022-02-14', '2022-02-21', '2022-02-28', '2022-03-07', '2022-03-14', '2022-03-21', '2022-03-28', '2022-04-04', '2022-04-11', '2022-04-18', '2022-04-25', '2022-05-02', '2022-05-09', '2022-05-16', '2022-05-23', '2022-05-30', '2022-06-06', '2022-06-13', '2022-06-20', '2022-06-27', '2022-07-04', '2022-07-11', '2022-07-18', '2022-07-25', '2022-08-01', '2022-08-08', '2022-08-15', '2022-08-22', '2022-08-29', '2022-09-05', '2022-09-12', '2022-09-19', '2022-09-26', '2022-10-03', '2022-10-10', '2022-10-17', '2022-10-24', '2022-10-31', '2022-11-07', '2022-11-14', '2022-11-21', '2022-11-28', '2022-12-05', '2022-12-12', '2022-12-19', '2022-12-26', '2023-01-02', '2023-01-09', '2023-01-16', '2023-01-23', '2023-01-30', '2023-02-06', '2023-02-13', '2023-02-20', '2023-02-27', '2023-03-06', '2023-03-13', '2023-03-20', '2023-03-27', '2023-04-03', '2023-04-10', '2023-04-17', '2023-04-24']


def zhsj(dq):
    city_name = dq[5].drop_duplicates().values.tolist()
    county_name = dq[4].drop_duplicates().values.tolist()
    dq.columns = ['out_station_time', 'site_name', 'overrun_rate', 'is_truck', 'county', 'city']
    bins = [-1, 0, 20, 50, 100, 1000]

    dq[['overrun_rate']] = dq[['overrun_rate']].astype(float)
    nn = dq.groupby(['site_name']).apply(
        lambda x: pd.cut(x['overrun_rate'], bins).value_counts(sort=False)).stack().unstack()

    nn.columns = ['非超限数', "超限<=20%车辆数", "20%<超限<=50%车辆数", "50%超限<=100%车辆数", "超限>100%车辆数"]

    nn["货车流量"] = nn.sum(axis=1)
    nn["超限货车"] = nn[["超限<=20%车辆数", "20%<超限<=50%车辆数", "50%超限<=100%车辆数", "超限>100%车辆数"]].sum(axis=1)
    nn['平均超限率(%)'] = nn["超限货车"] / nn["货车流量"] * 100
    nn['超限<=20%占比(%)'] = nn["超限<=20%车辆数"] / nn["超限货车"] * 100
    nn['20%<超限<=50%占比(%)'] = nn["20%<超限<=50%车辆数"] / nn["超限货车"] * 100
    nn['50%超限<=100%占比(%)'] = nn["50%超限<=100%车辆数"] / nn["超限货车"] * 100
    nn['超限>100%占比(%)'] = nn["超限>100%车辆数"] / nn["超限货车"] * 100

    # 数据格式处理区
    nn['平均超限率(%)'] = nn['平均超限率(%)'].round(2)
    nn['超限<=20%占比(%)'] = nn['超限<=20%占比(%)'].round(2)
    nn['20%<超限<=50%占比(%)'] = nn['20%<超限<=50%占比(%)'].round(2)
    nn['50%超限<=100%占比(%)'] = nn['50%超限<=100%占比(%)'].round(2)
    nn['超限>100%占比(%)'] = nn['超限>100%占比(%)'].round(2)
    nn.index.name = '站点'
    nn = nn.reset_index()

    mm = nn[['站点', '货车流量', '超限货车', '平均超限率(%)', '超限<=20%车辆数', '超限<=20%占比(%)', '20%<超限<=50%车辆数', '20%<超限<=50%占比(%)',
             '50%超限<=100%车辆数', '50%超限<=100%占比(%)', '超限>100%车辆数', '超限>100%占比(%)']]
    mm=mm.sort_values('货车流量',ascending=False)
    mm_hz = pd.DataFrame(columns=mm.columns)
    mm_hz['站点'] = ['汇总']
    mm_hz['货车流量'] = mm['货车流量'].sum()
    mm_hz['超限货车'] = mm['超限货车'].sum()
    mm_hz['平均超限率(%)'] = mm['超限货车'].sum() / mm['货车流量'].sum() * 100
    mm_hz['超限<=20%车辆数'] = mm["超限<=20%车辆数"].sum()
    mm_hz['超限<=20%占比(%)'] = mm["超限<=20%车辆数"].sum() / mm["超限货车"].sum() * 100
    mm_hz['20%<超限<=50%车辆数'] = mm["20%<超限<=50%车辆数"].sum()
    mm_hz['20%<超限<=50%占比(%)'] = mm["20%<超限<=50%车辆数"].sum() / mm["超限货车"].sum() * 100
    mm_hz['50%超限<=100%车辆数'] = mm["50%超限<=100%车辆数"].sum()
    mm_hz['50%超限<=100%占比(%)'] = mm["50%超限<=100%车辆数"].sum() / mm["超限货车"].sum() * 100
    mm_hz['超限>100%车辆数'] = mm["超限>100%车辆数"].sum()
    mm_hz['超限>100%占比(%)'] = mm["超限>100%车辆数"].sum() / mm["超限货车"].sum() * 100
    # 数据格式处理区
    mm_hz['平均超限率(%)'] = mm_hz['平均超限率(%)'].round(2)
    mm_hz['超限<=20%占比(%)'] = mm_hz['超限<=20%占比(%)'].round(2)
    mm_hz['20%<超限<=50%占比(%)'] = mm_hz['20%<超限<=50%占比(%)'].round(2)
    mm_hz['50%超限<=100%占比(%)'] = mm_hz['50%超限<=100%占比(%)'].round(2)
    mm_hz['超限>100%占比(%)'] = mm_hz['超限>100%占比(%)'].round(2)

    boss = pd.concat([mm, mm_hz])
    # ll = i + "成功读取数据"
    # log.append(ll)
    # print('==========')
    # 全表
    boss = boss.fillna(value=0)
    boss=boss.reset_index()
    # print(boss.head(5))
    # boss.to_excel(sc_dz+'\\'+i+'重点'+qishi_date.split('-')[-2]+qishi_date.split('-')[-1]+'_'+jiezi_date.split('-')[-2]+jiezi_date.split('-')[-1]+'.xlsx',index=True)
    boss_1 = pd.DataFrame()
    boss_1['序号'] = [i + 1 for i in range(len(boss))]
    boss_1['站点名称'] = boss['站点']
    boss_1['城市'] = city_name * len(boss)
    boss_1['区县'] = county_name * len(boss)
    boss_1['货车流量'] = boss['货车流量']
    boss_1['超限车次'] = boss['超限货车']
    boss_1['超限率(%)'] = boss['平均超限率(%)']
    boss_1['20%-50%'] = boss['20%<超限<=50%车辆数']
    boss_1['50%-100%'] = boss['50%超限<=100%车辆数']
    boss_1['100%以上'] = boss['超限>100%车辆数']

    return boss_1





def show_data_2(*args):
    sc_dwj_e_1.set("")
    county_list_1 = cc.loc[cc['city'] == cbx_1.get()]['county'].drop_duplicates().values.tolist()
    cbx_2["values"] =county_list_1
    #cbx_2.current(0)
def show_data_time(*args):
    sc_dwj_e_1.set("")
    cbx_4["values"] = dd.loc[dd['年月'] == cbx_3.get()]['日期'].drop_duplicates().values.tolist()

def show_data_1(*args):
    sc_dwj_e_1.set("")
    sc_dwj_e_0.set("等待点击进行分析")
def tjfx():
    log = []
    dg = cbx_2.get()
    # aa=['柯桥']
    sc_dz = b1
    qishi_date = cbx_4.get()
    qishi_time = datetime.strptime(qishi_date, '%Y-%m-%d')
    jiezi_time = qishi_time + Day(7)
    jiezi_d = qishi_time + Day(6)
    jiezi_date = jiezi_d.strftime('%Y-%m-%d')
    cj = sc_dwj_e_4.get().split('-')
    try:


        if len(cj)>1:
            bx_list=sc_dwj_e_4.get().split('-')
        else:
            bx_list=[dg]

        for i in bx_list:

            try:

                db = pymysql.connect(host='192.168.2.119',user='zjzhzcuser',passwd='F4dus0ee',port=3306,charset='utf8',database='db_manage_overruns')

                sql = "select out_station_time,site_name,overrun_rate,is_truck,county,city\
                                                       from t_bas_pass_data_31_12 as a\
                                                       left join t_code_area as b\
                                                       on a.area_county=b.county_code where county= %s and is_truck='1' and out_station_time >= %s and out_station_time <%s"
                sql1 = "select out_station_time,site_name,overrun_rate,is_truck,county,city\
                                                           from t_bas_pass_data_71 as a\
                                                           left join t_code_area as b\
                                                           on a.area_county=b.county_code where county= %s and is_truck='1' and out_station_time >= %s and out_station_time <%s"
                # print('aaaaaaaaaaa')

                cursor = db.cursor()

                cursor.execute(sql, [i, qishi_time, jiezi_time])
                # 传递单个参数时 cursor.execute(sql,'B00140N5CS')
                # print('bbbbbbbbbbbbbbbbb')
                data = cursor.fetchall()

                cursor = db.cursor()
                cursor.execute(sql1, [i, qishi_time, jiezi_time])
                data_data = cursor.fetchall()

                dq1 = pd.DataFrame(list(data))
                dq2 = pd.DataFrame(list(data_data))
                # print(dq1)
                # print(dq2)
                # print(len(dq1))
                if dq1.empty and dq2.empty:
                    ll = i + " 无非现 无重点"
                    log.append(ll)

                elif dq1.empty and dq2.empty == False:
                    boss_1 = zhsj(dq2)

                    # 格式处理区
                    wb = Workbook()
                    ws = wb.active
                    for r in dataframe_to_rows(boss_1, index=False, header=True):
                        ws.append(r)
                    # print('gggggg')
                    ws.insert_rows(1)
                    ws['A1'] = '重点货运源' + ' ' + qishi_date + ' 周报'
                    ws.merge_cells('A1:J1')
                    # 几列单独调整列宽
                    ws.column_dimensions['A'].width = 10
                    ws.column_dimensions['B'].width = 35
                    ws.column_dimensions['C'].width = 10
                    ws.column_dimensions['D'].width = 10
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 15
                    ws.column_dimensions['H'].width = 15
                    ws.column_dimensions['I'].width = 15
                    ws.column_dimensions['j'].width = 15
                    # 设置所有的居中
                    for row in ws[1:boss_1.shape[0] + 2]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="center")
                            # 设置第一行单独加粗''''""""
                    # 设置局部靠左
                    for row in ws['B3':'B' + str(boss_1.shape[0] + 1)]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="left")
                            # 设置第一行单独加粗''''""""
                    # 设置局部靠右
                    for row in ws['E3':'J' + str(boss_1.shape[0] + 2)]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="right")
                            # 设置第一行单独加粗''''""""
                    for c in ws[1]:
                        c.font = Font(name="宋体", size=15, bold=True)
                    for c in ws[2]:
                        c.font = Font(name="宋体", size=12, bold=True)
                    wb.save(sc_dz + '\\' + i + qishi_date.split('-')[-2] + qishi_date.split('-')[-1] + '_' +
                            jiezi_date.split('-')[-2] + jiezi_date.split('-')[-1] + '.xlsx')

                    ll = i + " 无非现 有重点"
                    log.append(ll)

                elif dq1.empty == False and dq2.empty == True:
                    boss_1 = zhsj(dq1)
                    # 格式处理区
                    wb = Workbook()
                    ws = wb.active
                    for r in dataframe_to_rows(boss_1, index=False, header=True):
                        ws.append(r)
                    # print('gggggg')
                    ws.insert_rows(1)
                    ws['A1'] = '非现场站点' + ' ' + qishi_date + ' 周报'
                    ws.merge_cells('A1:J1')
                    # 几列单独调整列宽
                    ws.column_dimensions['A'].width = 10
                    ws.column_dimensions['B'].width = 35
                    ws.column_dimensions['C'].width = 10
                    ws.column_dimensions['D'].width = 10
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 15
                    ws.column_dimensions['H'].width = 15
                    ws.column_dimensions['I'].width = 15
                    ws.column_dimensions['j'].width = 15
                    # 设置所有的居中
                    for row in ws[1:boss_1.shape[0] + 2]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="center")
                            # 设置第一行单独加粗''''""""
                    # 设置局部靠左
                    for row in ws['B3':'B' + str(boss_1.shape[0] + 1)]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="left")
                            # 设置第一行单独加粗''''""""
                    # 设置局部靠右
                    for row in ws['E3':'J' + str(boss_1.shape[0] + 2)]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="right")
                            # 设置第一行单独加粗''''""""
                    for c in ws[1]:
                        c.font = Font(name="宋体", size=15, bold=True)
                    for c in ws[2]:
                        c.font = Font(name="宋体", size=12, bold=True)
                    wb.save(sc_dz + '\\' + i + qishi_date.split('-')[-2] + qishi_date.split('-')[-1] + '_' +
                            jiezi_date.split('-')[-2] + jiezi_date.split('-')[-1] + '.xlsx')

                    ll = i + " 有非现 无重点"
                    log.append(ll)

                # and out_station_time >= '2021-01-01 00:00:00' and out_station_time <='2022-05-06 00:00:00'
                # 城市备用
                # dq=dq[['out_station_time','site_name','overrun_rate','is_truck','city','county']]
                # 区县备用
                else:
                    boss_1 = zhsj(dq1)
                    data_2 = zhsj(dq2)

                    # 格式处理区
                    wb = Workbook()
                    ws = wb.active
                    for r in dataframe_to_rows(boss_1, index=False, header=True):
                        ws.append(r)
                    # print('gggggg')

                    # 表格标记
                    biaoji = boss_1.shape[0] + 5

                    # 第二张表的数据
                    for j in range(data_2.shape[1]):
                        ws.cell(row=boss_1.shape[0] + 5, column=1 + j).value = data_2.columns[j]
                    for m in range(data_2.shape[0]):
                        for n in range(data_2.shape[1]):
                            ws.cell(row=boss_1.shape[0] + 6 + m, column=1 + n).value = data_2.iloc[m, n]

                    ws.insert_rows(1)
                    ws['A1'] = '非现场站点' + ' ' + qishi_date + ' 周报'
                    ws.merge_cells('A1:j1')
                    ws['A' + str(biaoji)] = '重点货运源' + ' ' + qishi_date + ' 周报'
                    ws.merge_cells('A' + str(biaoji) + ':' + 'j' + str(biaoji))
                    # 几列单独调整列宽
                    ws.column_dimensions['A'].width = 10
                    ws.column_dimensions['B'].width = 35
                    ws.column_dimensions['C'].width = 10
                    ws.column_dimensions['D'].width = 10
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 15
                    ws.column_dimensions['H'].width = 15
                    ws.column_dimensions['I'].width = 15
                    ws.column_dimensions['j'].width = 15
                    # 设置所有的居中
                    for row in ws[1:biaoji - 3]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="center")
                            # 设置第一行单独加粗''''""""
                    for row in ws[biaoji:ws.max_row]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="center")
                            # 设置第一行单独加粗
                    # 设置局部靠左
                    for row in ws['B3':'B' + str(biaoji - 4)]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="left")
                            # 设置第一行单独加粗''''""""
                    for row in ws['B' + str(biaoji + 2):'B' + str(ws.max_row - 1)]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="left")
                            # 设置第一行单独加粗
                    # 设置局部靠右
                    for row in ws['E3':'J' + str(biaoji - 3)]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="right")
                            # 设置第一行单独加粗''''""""
                    for row in ws['E' + str(biaoji + 2):'J' + str(ws.max_row)]:
                        for c in row:
                            # 字体设置
                            c.font = Font(name='宋体', size=12)
                            # 边框线设置
                            c.border = Border(left=Side(border_style='thin', color='FF000000'),
                                              right=Side(border_style='thin', color='FF000000'),
                                              top=Side(border_style='thin', color='FF000000'),
                                              bottom=Side(border_style='thin', color='FF000000'), )
                            # 设置对齐方式
                            c.alignment = Alignment(horizontal="right")
                            # 设置第一行单独加粗''''""""
                    for c in ws[1]:
                        c.font = Font(name="宋体", size=15, bold=True)
                    for c in ws[2]:
                        c.font = Font(name="宋体", size=12, bold=True)
                    for c in ws[biaoji]:
                        c.font = Font(name="宋体", size=15, bold=True)
                    for c in ws[biaoji + 1]:
                        c.font = Font(name="宋体", size=12, bold=True)
                    # print("是这里有错误吗？")
                    wb.save(sc_dz + '\\' + i + qishi_date.split('-')[-2] + qishi_date.split('-')[-1] + '_' +
                            jiezi_date.split('-')[-2] + jiezi_date.split('-')[-1] + '.xlsx')
                    ll = i + " 有非现 有重点"
                    log.append(ll)


                #print(log)
                db.close()
                #sc_dwj_e_1.set("本次周报全部生成！！！")

            except:
                ll = i + " 发生了某些未知错误"
                log.append(ll)

        llll = pd.DataFrame()
        llll['文件结果'] = log
        llll.to_csv(b1 + '\\' + qishi_date.split('-')[-2] + qishi_date.split('-')[-1] + '_' +
                    jiezi_date.split('-')[-2] + jiezi_date.split('-')[-1] + '日志' + '.txt', sep='\t', index=False)
        sc_dwj_e_1.set("本次周报全部生成！！！")
    except:
        sc_dwj_e_1.set("参数设置遗漏，请检查重新尝试！")
    #sc_dwj_e_1.set("发生了某些错误，请重新尝试！")


#sc_dwj_e_1.set("本次周报全部生成！！！")
#sc_dwj_e_1.set("发生了某些错误，请重新尝试！")
def dbhs():
    pass

def sc_1():
    sc_dwj_e_1.set("")
    global b1
    b1 = filedialog.askdirectory()  # 返回目录名 适用于多个文件的输出
    sc_e_1.set(b1)
root=Tk()
s1=StringVar()
s1.set("")
s2=StringVar()
s2.set("")
root.title('区县-周报生成系统')
# kk=Frame(root,width=500,height=200)
# kk.grid(row=0,column=0,rowspan=1)


k5=LabelFrame(root,text="输入确认",width=500,height=200,bg='Red')
k5.grid(row=20,column=0,rowspan=2)

Label(k5,text='城市',width=15).grid(row=10,column=0,rowspan=1,pady=2,padx=4)
cbx_1 =ttk.Combobox(k5, width=30, height=8)
cbx_1.grid(column=1, row=10)
cbx_1.configure(state="readonly")
#第一个下拉框值设置
cbx_1["values"] = cc['city'].drop_duplicates().values.tolist()
#cbx_1.current(1)
cbx_1.bind("<<ComboboxSelected>>",show_data_2)
Label(k5,text='区县',width=15).grid(row=10,column=2,rowspan=1,pady=2,padx=4)
cbx_2 =ttk.Combobox(k5, width=30, height=8)
cbx_2.grid(column=3, row=10)
cbx_2.configure(state="readonly")
cbx_2.bind("<<ComboboxSelected>>",show_data_1)

Label(k5,text='多表模式',width=15).grid(row=20,column=0,rowspan=1,pady=2,padx=4)
sc_dwj_e_4=StringVar()
sc_dwj_e_4.set("")
Entry(k5,textvariable=sc_dwj_e_4,width=85).grid(row=20,column=1,columnspan=3,pady=2,padx=4)
#Label(k5,text='截止日期',width=15).grid(row=20,column=3,rowspan=1,pady=2,padx=4)
# sc_dwj_e_5=StringVar()
# sc_dwj_e_5.set("")
# Entry(k5,textvariable=sc_dwj_e_5,width=33).grid(row=20,column=3,rowspan=1,pady=2,padx=4)

Label(k5,text='年月选择',width=15).grid(row=30,column=0,rowspan=1,pady=2,padx=4)
cbx_3 =ttk.Combobox(k5, width=30, height=8)
cbx_3.grid(column=1, row=30)
cbx_3.configure(state="readonly")
cbx_3["values"]=dd['年月'].drop_duplicates().values.tolist()
cbx_3.bind("<<ComboboxSelected>>",show_data_time)

Label(k5,text='日期选择',width=15).grid(row=30,column=2,rowspan=1,pady=2,padx=4)
cbx_4 =ttk.Combobox(k5, width=30, height=8)
cbx_4.grid(column=3, row=30)
cbx_4.configure(state="readonly")
cbx_4.bind("<<ComboboxSelected>>",show_data_1)


k7=LabelFrame(root,text="输出保存",width=500,height=200,bg='Red')
k7.grid(row=30,column=0,rowspan=2)

sc_e_1=StringVar()
sc_e_1.set("")
Label(k7,text='文件地址',width=15).grid(row=0,column=0,rowspan=1,pady=2,padx=4)
Entry(k7,textvariable=sc_e_1,state='readonly',width=40).grid(row=0,column=1,rowspan=1,pady=2,padx=4)
Button(k7,text="结果保存",width=15,command=sc_1).grid(row=0,column=2,rowspan=1,pady=2,padx=4)
#Entry(k7,textvariable=sc_dml_e_1,width=40).grid(row=0,column=3,rowspan=1,pady=2,padx=4)

k8=LabelFrame(root,text="分析模块",width=500,height=200,bg='Red')
k8.grid(row=35,column=0,rowspan=2)

sc_dwj_e_0=StringVar()
sc_dwj_e_0.set("")
sc_dwj_e_1=StringVar()
sc_dwj_e_1.set("")
Label(k8,text='程序状态',width=15).grid(row=0,column=0,rowspan=1,pady=2,padx=4)
Entry(k8,textvariable=sc_dwj_e_0,width=40,state='readonly').grid(row=0,column=1,rowspan=1,pady=2,padx=4)
Label(k8,text='分析结果',width=15).grid(row=2,column=0,rowspan=1,pady=2,padx=4)
Entry(k8,textvariable=sc_dwj_e_1,width=40,state='readonly').grid(row=2,column=1,rowspan=1,pady=2,padx=4)
Button(k8,text="提交分析",width=15,command=tjfx).grid(row=2,column=2,rowspan=1,pady=2,padx=4)

#Text(k8,width=40,height=10).grid(row=1,column=0,rowspan=1,columnspan=1,pady=2,padx=4)
#Entry(k8,textvariable=sc_dwj_e_1,width=40).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
#Button(k8,text="结果预览",width=15).grid(row=1,column=1,rowspan=1,pady=2,padx=4)
#Button(k8,text="保存结果",width=15,command=sc_dwj).grid(row=1,column=2,rowspan=1,pady=2,padx=4)

root.mainloop()

