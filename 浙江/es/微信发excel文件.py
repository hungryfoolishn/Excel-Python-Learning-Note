
import openpyxl,  PyOfficeRobot  # 导入两个库，第一个处理excel，第二个用到微信发消息功能，第二个库还有很多强大便捷的功能。

wb = openpyxl.load_workbook(r'C:\Users\stayhungary\Desktop\t_bas_over_data_31_81.xlsx')  # 括号中写汇总文件地址
ws = wb['Sheet1']  # 获取excel表单
Names = ws['B']  # 获取表单中第二列，我的表单第二列是微信昵称，可以根据实际进行调整
max_row = ws.max_row  # 获取excel的最大行数

for Name in Names:  # 第5行代码获得的昵称需要遍历
    Name = Name.value  # 遍历出来的是元祖，需要用value进行取值
    if Name == '程序员晚枫':  # 我的excel B1单元格写的是程序员晚枫，大家可以根据实际调整
        continue
    else:
        wb = openpyxl.load_workbook(r'C:\Users\stayhungary\Desktop\t_bas_over_data_31_81.xlsx')
        ws = wb['Sheet1']
        for j in range(max_row, 1, -1):  # 倒着遍历，方便删除时不错序
            if ws[f'B{j}'].value != Name:
                ws.delete_rows(j)  # 删除行
        file_path = r'C:\Users\stayhungary\Desktop\t_bas_over_data_31_82.xlsx'  # 重命名
        wb.save(file_path)  # 保存excel
        PyOfficeRobot.file.send_file(who=Name, file=file_path)  # 通过微信发送文件，分别在括号里写上微信昵称和文件地址。


