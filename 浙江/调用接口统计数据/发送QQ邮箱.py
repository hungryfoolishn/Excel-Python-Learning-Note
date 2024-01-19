# coding: utf-8


import time
import pandas as pd
import requests
import base64
import datetime
import pymysql
import json
import csv
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import Font,Border,Side,Alignment
import pandas as pd
from openpyxl import load_workbook

def retry(max_attempts, delay=1):
    def decorator(func):
        def wrapper(*args, **kwargs):
            attempts = 0
            while attempts < max_attempts:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    attempts += 1
                    print(f"Attempt {attempts} failed: {e}")
                    time.sleep(delay)
            print(f"Function failed after {max_attempts} attempts")
        return wrapper
    return decorator


# 获取数据
def get_data(sql):
    # if index == 3:
    #     print (sql['where'])
    #     return
    data = {
        "type": "query",
        "tableName": sql['tableName'],
        "where": (base64.b64encode(sql['where'].encode())).decode(),
        "columns": sql['columns'],
        "isEncry": "1"
    }
    url = 'https://yhxc.jtyst.zj.gov.cn:7443/zc-interface/db/excuteSql'
    headers = {'content-type': 'application/json'}
    res = requests.post(url, json=data, headers=headers)
    res = json.loads(res.text)
    data = res['data']
    data=pd.DataFrame(data)
    return  data

def write_excel_file(file_path):

    wb = load_workbook(file_path)


    # 初始化worksheet对象
    ws = wb.active

    # 定义样式
    fontstyle = Font(name="Microsoft YaHei UI", size=12,  bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),bottom=Side(style="thin"))
    align = Alignment(horizontal="center", vertical="center")
    fontstyle2 = Font(name="Microsoft YaHei UI", size=12)
    side = Side(
        style="medium",
        # 边框样式，可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
        color="ff66dd",  # 边框颜色，16进制rgb表示
    )
    for i in ('B', 'C', 'D', 'E', 'F'):
        ws.column_dimensions[i].width = 20
    ws.column_dimensions['D'].width = 30
    row_nu=range(ws.max_row)
    # 设置单元格样式
    for i in range(ws.max_column):
        ws.cell(row=1, column=i + 1).font = fontstyle
        ws.cell(row=1, column=i + 1).border = border
        ws.cell(row=1, column=i + 1).alignment = align
        for j in row_nu:
            ws.row_dimensions[j+1].height = 20  # 设置第2行高度为30
            ws.cell(row=j + 2, column=i+1).font = fontstyle2
            ws.cell(row=j+2, column=2).alignment = align
            ws.cell(row=j+1, column=i+1).border = border
    ws.delete_cols(1)


    #保存到Excel文件
    wb.save(file_path)

@retry(max_attempts=3, delay=2)
def 每日全省案件处罚数据():
    import datetime
    now = datetime.datetime.now()
    date = datetime.datetime.strftime(now - datetime.timedelta(days=1), '%Y-%m-%d')  # 获取日期
    # 当前时间  必填
    start_time = '{}'.format(date)
    end_time = '{}'.format(date)

    ##--交警现场
    sql_t_code_area = {
        "tableName": "   t_code_area    ",
        "where": "  is_deleted =0 and province_code= '330000' ",
        "columns": "county_code as area_county,city_code,city as 地市,county as 区县"
    }

    ##--交通现场查处数


    sql_交通现场查处数 = {
        "tableName": " t_bas_over_data_collection_sign c",
        "where": " c.record_type = 99 AND c.insert_type = 5  AND c.close_case_time between '{} 00:00:00'  and '{} 23:59:59'  GROUP BY c.area_county  ".format(start_time,end_time),
        "columns": "c.area_county as area_county ,count( DISTINCT ( CASE_NUM ) ) AS 交通现场查处数 "
    }

    ##--交通现场查处数明细

    sql_交通非现查处数 = {
        "tableName": "t_case_sign_result a  ",
        "where": " record_type = 31 AND insert_type = 1 AND data_source = 1 AND case_type = 1 AND close_case_time between '{} 00:00:00' AND  '{} 23:59:59' AND area_province = '330000' GROUP BY  dept_county".format(
            start_time, end_time),
        "columns": "dept_county as area_county,count(DISTINCT ( CASE_NUM )) AS 交通非现查处数,  count(DISTINCT IF( car_no LIKE '%浙%', CASE_NUM, NULL )) AS 交通非现查处数本省 "
    }

    ##--交警现场
    sql_交警现场 = {
        "tableName": "   t_bas_police_road_site a    ",
        "where": "   a.punish_time between '{} 00:00:00' AND  '{} 23:59:59'  and case_status=2  GROUP BY area_county  ".format(
            start_time, end_time),
        "columns": "area_county,count(DISTINCT case_number) as 交警现场查处数"
    }
    t_code_area=get_data(sql_t_code_area)

    交通现场查处数 = get_data(sql_交通现场查处数)
    交通现场查处数 = pd.DataFrame(交通现场查处数,
                           columns=['area_county', '交通现场查处数'])

    交通非现查处数 = get_data(sql_交通非现查处数)
    交通非现查处数 = pd.DataFrame(交通非现查处数,
                           columns=['area_county', '交通非现查处数','交通非现查处数本省'])
    交警现场 = get_data(sql_交警现场)
    交警现场 = pd.DataFrame(交警现场,
                           columns=['area_county', '交警现场查处数'])
    U_all = pd.merge(t_code_area, 交通非现查处数, on=['area_county'], how='left')
    U_all = pd.merge(U_all, 交通现场查处数, on=['area_county'], how='left')
    U_all = pd.merge(U_all, 交警现场, on=['area_county'], how='left')
    U_all = pd.DataFrame(U_all,
                           columns=['地市', '区县', '交通非现查处数', '交通非现查处数本省', '交通现场查处数', '交警现场查处数', 'city_code', 'area_county'])
    U_all=U_all.fillna(0)
    U_all = U_all.sort_values(by=['area_county'],
                                    ascending=True).reset_index(drop=True)
    U_all地市 = U_all.groupby(['city_code', '地市']).sum().reset_index()
    U_all地市 = pd.DataFrame(U_all地市,
                           columns=['地市', '交通非现查处数', '交通非现查处数本省', '交通现场查处数', '交警现场查处数'])
    U_all省=U_all地市.copy()
    U_all省['省']='浙江'
    U_all省 = U_all省.groupby(['省']).sum().reset_index()
    U_all省 = pd.DataFrame(U_all省,
                           columns=['省', '交通非现查处数', '交通非现查处数本省', '交通现场查处数', '交警现场查处数'])

    U_all省.rename(
        columns={'省': '地市'}, inplace=True)
    U_all省=pd.concat([U_all地市,U_all省])
    print(U_all省)
    with pd.ExcelWriter(r'C:\Users\stayhungary\Desktop\每日全省案件处罚数据.xlsx') as writer1:
        U_all省.to_excel(writer1, sheet_name='地市汇总', index=True)
        U_all地市.to_excel(writer1, sheet_name='地市汇总', index=True)
        U_all.to_excel(writer1, sheet_name='每日全省案件处罚数据', index=True)
        交通现场查处数.to_excel(writer1, sheet_name='交通现场查处数', index=True)
        交警现场.to_excel(writer1, sheet_name='交警现场', index=True)
    file_path=r'C:\Users\stayhungary\Desktop\每日全省案件处罚数据.xlsx'
    write_excel_file(file_path)
    from threading import Timer
    import datetime
    """定时1天"""
    now_time = datetime.datetime.now()
    next_time = now_time + datetime.timedelta(days=+1)
    next_year = next_time.date().year
    next_month = next_time.date().month
    next_day = next_time.date().day
    next_time2 = datetime.datetime.strptime(
        str(next_year) + "-" + str(next_month) + "-" + str(next_day) + " " + "7:00:00", "%Y-%m-%d %H:%M:%S")
    timer_start_time2 = (next_time2 - now_time).total_seconds()
    print('下次复检开始时间', next_time2)
    t2 = Timer(timer_start_time2, 每日全省案件处罚数据)  # 此处使用递归调用实现
    t2.start()



@retry(max_attempts=3, delay=2)
def To_emil():
    import datetime
    now = datetime.datetime.now()
    date = datetime.datetime.strftime(now - datetime.timedelta(days=1), '%Y%m%d')  # 获取日期
    # 邮件配置信息
    smtp_server = 'smtp.qq.com'
    smtp_port = 465
    smtp_ssl = True
    smtp_user = '1399120443@qq.com'
    smtp_password = 'rycnskfchlpkijad'  # 邮箱授权码

    # 发送邮件信息
    sender = '1399120443@qq.com'
    receivers = ['liu.wenjie@zcits.com']

    # 邮件正文
    mail_content = '每日全省案件处罚数据({}).xlsx'.format(date)
    message = MIMEMultipart()

    # 邮件信息配置
    message['From'] = '1399120443@qq.com'
    message['To'] = 'liu.wenjie@zcits.com'
    message['Subject'] = Header("每日全省案件处罚数据({})".format(date), 'utf-8')

    csv_file = open(r'C:\Users\stayhungary\Desktop\每日全省案件处罚数据.xlsx', 'rb').read()
    csv = MIMEApplication(csv_file)
    csv.add_header('Content-Disposition', 'attachment', filename='每日全省案件处罚数据({}).xlsx'.format(date))
    message.attach(csv)

    # 发送邮件
    try:
        smtpObj = smtplib.SMTP_SSL(smtp_server, smtp_port)
        smtpObj.login(smtp_user, smtp_password)
        smtpObj.sendmail(sender, receivers, message.as_string())
        print("邮件发送成功")
    except smtplib.SMTPException as e:
        print("Error: 邮件发送失败: ", e)
    from threading import Timer
    import datetime
    """定时1天"""
    now_time = datetime.datetime.now()
    next_time = now_time + datetime.timedelta(days=+1)
    next_year = next_time.date().year
    next_month = next_time.date().month
    next_day = next_time.date().day
    next_time2 = datetime.datetime.strptime(
        str(next_year) + "-" + str(next_month) + "-" + str(next_day) + " " + "7:05:00", "%Y-%m-%d %H:%M:%S")
    timer_start_time2 = (next_time2 - now_time).total_seconds()
    print('下次复检开始时间', next_time2)
    t2 = Timer(timer_start_time2, To_emil)  # 此处使用递归调用实现
    t2.start()


if __name__ == "__main__":
    每日全省案件处罚数据()
    To_emil()
