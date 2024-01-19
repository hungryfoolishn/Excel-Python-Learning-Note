# coding: utf-8


import time
import pandas as pd
import requests
import base64
import datetime
import pymysql
import json
import csv
import time
import pandas as pd
import requests
import base64
import json
from openpyxl.styles import Font,Border,Side,Alignment
import pandas as pd
from openpyxl import load_workbook


start_time='2023-10-01'
end_time='2023-10-25'




# 获取数据
def get_data(sql):
    # if index == 3:
    #     print (sql['where'])
    #     return
    data = {
        "type": "query",
        "tableName": sql['tableName'],
        "where": (base64.b64encode(sql['where'].encode())).decode(),
        "columns": sql['columns'],
        "isEncry": "1"
    }
    url = 'https://yhxc.jtyst.zj.gov.cn:7443/zc-interface/db/excuteSql'
    headers = {'content-type': 'application/json'}
    res = requests.post(url, json=data, headers=headers)
    res = json.loads(res.text)
    data = res['data']
    data=pd.DataFrame(data)
    return  data


def 汇总数据():
    ##--交通现场查处数明细

    月报80吨以上明细 = {
        "type": "query",
        "tableName": "t_bas_over_data_31 a left join t_code_area b on a.area_county = b.county_code  ",
        "where": " out_station_time between '{} 00:00:00' and '{} 23:59:59' and allow is null and total_weight between 80 and 300 ".format(start_time,end_time),
        "columns": "*"
    }

    月报超限100以上明细 = {
        "type": "query",
        "tableName": "t_bas_over_data_31 a left join t_code_area b on a.area_county = b.county_code  ",
        "where": " out_station_time between '{} 00:00:00' and '{} 23:59:59' and allow is null and total_weight <100 and overrun_rate>=100 ".format(start_time,end_time),
        "columns": "*"
    }
    月报80吨以上明细=get_data(月报80吨以上明细)
    月报80吨以上明细 = pd.DataFrame(月报80吨以上明细,
                      columns=['record_code','city', 'county', 'out_station_time', 'car_no', 'total_weight', 'limit_weight', 'overrun',
                               'axis', 'site_name', 'direction', 'overrun_rate', 'photo1', 'photo2', 'photo3', 'vedio',
                               'insert_time',  'has_car_no', 'area_county', 'allow'])
    月报80吨以上明细['total_weight'] = 月报80吨以上明细['total_weight'].astype('float')
    月报80吨以上明细['limit_weight'] = 月报80吨以上明细['limit_weight'].astype('float')
    月报80吨以上明细['overrun'] = 月报80吨以上明细['overrun'].astype('float')
    月报80吨以上明细.rename(columns={'record_code': '记录号','out_station_time': '检测时间','car_no': '车牌'
                        ,'total_weight': '总重','limit_weight': '限重','overrun': '超限吨数','axis': '轴数','site_name': '站点名称'
        , 'direction': '行驶方向','overrun_rate': '超限率','insert_time': '插入时间','photo1': '车头照','photo2': '车身照','photo3': '车尾照'
                              ,'vedio': '视频','city': '地市'
        ,'county': '区县','has_car_no': '有无车牌'
                        }, inplace=True)
    月报80吨以上明细 = 月报80吨以上明细.sort_values(by=['area_county'],
                                    ascending=True).reset_index(drop=True)
    月报超限100以上明细=get_data(月报超限100以上明细)
    月报超限100以上明细 = pd.DataFrame(月报超限100以上明细,
                      columns=['record_code','city', 'county', 'out_station_time', 'car_no', 'total_weight', 'limit_weight', 'overrun',
                               'axis', 'site_name', 'direction', 'overrun_rate', 'photo1', 'photo2', 'photo3', 'vedio',
                               'insert_time',  'has_car_no', 'area_county', 'allow'])
    月报超限100以上明细['total_weight'] = 月报超限100以上明细['total_weight'].astype('float')
    月报超限100以上明细['limit_weight'] = 月报超限100以上明细['limit_weight'].astype('float')
    月报超限100以上明细['overrun'] = 月报超限100以上明细['overrun'].astype('float')
    月报超限100以上明细.rename(columns={'record_code': '记录号','out_station_time': '检测时间','car_no': '车牌'
                        ,'total_weight': '总重','limit_weight': '限重','overrun': '超限吨数','axis': '轴数','site_name': '站点名称'
        , 'direction': '行驶方向','overrun_rate': '超限率','insert_time': '插入时间','photo1': '车头照','photo2': '车身照','photo3': '车尾照'
                              ,'vedio': '视频','city': '地市'
        ,'county': '区县','has_car_no': '有无车牌'
                        }, inplace=True)
    月报超限100以上明细 = 月报超限100以上明细.sort_values(by=['area_county'],
                                    ascending=True).reset_index(drop=True)



    # data2=get_data(sql_80吨以上非现查处数)
    # data3=get_data(sql_交警现场)
    with pd.ExcelWriter(r'C:\Users\stayhungary\Desktop\1025月报明细1.0.xlsx') as writer1:
        月报80吨以上明细.to_excel(writer1, sheet_name='月报80吨以上明细', index=True)
        月报超限100以上明细.to_excel(writer1, sheet_name='月报超限100以上明细', index=True)
        # data2.to_excel(writer1, sheet_name='sql_80吨以上非现查处数', index=True)
        # data3.to_excel(writer1, sheet_name='sql_80吨以上交警现场', index=True)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def write_excel_file(file_path):

    wb = load_workbook(file_path)


    # 初始化worksheet对象
    ws = wb.active

    # 定义样式
    fontstyle = Font(name="Microsoft YaHei UI", size=12,  bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),bottom=Side(style="thin"))
    align = Alignment(horizontal="center", vertical="center")
    fontstyle2 = Font(name="Microsoft YaHei UI", size=12)
    side = Side(
        style="medium",
        # 边框样式，可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
        color="ff66dd",  # 边框颜色，16进制rgb表示
    )
    # for i in ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
    #     ws.column_dimensions[i].width = 30
    # ws.column_dimensions['D'].width = 30
    for c in range(1, 21):
        # 把列序数转变为字母
        w = get_column_letter(c)
        # 调整列宽
        ws.column_dimensions[w].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['H'].width = 40

    # df = pd.read_excel(file_path)
    # df.loc[len(df)] = list(df.columns)
    # for col in df.columns:
    #     index = list(df.columns).index(col)  # 列序号
    #     letter = get_column_letter(index + 1)  # 列字母
    #     collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
    #     ws.column_dimensions[letter].width = collen * 1.2

    # 设置单元格样式
    row_nu=range(ws.max_row)
    for i in range(ws.max_column):
        ws.cell(row=1, column=i + 1).font = fontstyle
        ws.cell(row=1, column=i + 1).border = border
        ws.cell(row=1, column=i + 1).alignment = align
        for j in row_nu:
            ws.row_dimensions[j+1].height = 20  # 设置第2行高度为30
            ws.cell(row=j + 2, column=i+1).font = fontstyle2
            ws.cell(row=j+2, column=2).alignment = align
            ws.cell(row=j+1, column=i+1).border = border
    ws.delete_cols(1)


    #保存到Excel文件
    wb.save(file_path)


def 分表():
    总重80吨以上明细 = pd.read_excel(r'C:\Users\stayhungary\Desktop\1025月报明细1.0.xlsx', sheet_name='月报80吨以上明细')
    超限100以上明细 = pd.read_excel(r'C:\Users\stayhungary\Desktop\1025月报明细1.0.xlsx', sheet_name='月报超限100以上明细')
    index80 = 总重80吨以上明细[['地市','区县']].drop_duplicates(['地市', '区县'])
    index100 = 超限100以上明细[['地市','区县']].drop_duplicates(['地市', '区县'])

    for 地市,区县 in index80.iterrows():
        print(区县[0],区县[1])
        总重80吨以上明细区县=总重80吨以上明细[(总重80吨以上明细.地市==区县[0]) &(总重80吨以上明细.区县==区县[1]) ]
        总重80吨以上明细区县 = pd.DataFrame(总重80吨以上明细区县,
                                   columns=['记录号','检测时间','车牌', '总重', '限重','超限吨数', '轴数',  '站点名称' ,
                                            '行驶方向', '超限率', '插入时间',  '车头照','车身照', '车尾照','视频',  '地市'
                                              ,'区县','有无车牌', 'area_county', 'allow'])


        with pd.ExcelWriter(r'G:\智诚\2023日常给出数据\月报原始数据明细\10月80吨以上及超限100明细数据\各区县80吨以上明细\10月{}-{}80吨以上明细.xlsx'.format(区县[0],区县[1])) as writer1:
            总重80吨以上明细区县.to_excel(writer1, sheet_name='月报80吨以上明细', index=True)
        file_path=r'G:\智诚\2023日常给出数据\月报原始数据明细\10月80吨以上及超限100明细数据\各区县80吨以上明细\10月{}-{}80吨以上明细.xlsx'.format(区县[0],区县[1])
        write_excel_file(file_path)
    for 地市,区县 in index100.iterrows():
        print(区县[0],区县[1])
        超限100以上明细区县=超限100以上明细[(超限100以上明细.地市==区县[0]) &(超限100以上明细.区县==区县[1]) ]
        超限100以上明细区县 = pd.DataFrame(超限100以上明细区县,
                                   columns=['记录号','检测时间','车牌', '总重', '限重','超限吨数', '轴数',  '站点名称' ,
                                            '行驶方向', '超限率', '插入时间',  '车头照','车身照', '车尾照','视频',  '地市'
                                              ,'区县','有无车牌', 'area_county', 'allow'])


        with pd.ExcelWriter(r'G:\智诚\2023日常给出数据\月报原始数据明细\10月80吨以上及超限100明细数据\各区县超限100以明细\10月{}-{}超限100以明细.xlsx'.format(区县[0],区县[1])) as writer1:
            超限100以上明细区县.to_excel(writer1, sheet_name='超限100以上明细', index=True)
        file_path=r'G:\智诚\2023日常给出数据\月报原始数据明细\10月80吨以上及超限100明细数据\各区县超限100以明细\10月{}-{}超限100以明细.xlsx'.format(区县[0],区县[1])
        write_excel_file(file_path)
if __name__ == "__main__":
    # 汇总数据()
    # file_path=r'G:\智诚\2023日常给出数据\月报原始数据明细\10月80吨以上及超限100明细数据\各区县80吨以上明细\10月杭州-萧山80吨以上明细.xlsx'
    # write_excel_file(file_path)

    分表()

