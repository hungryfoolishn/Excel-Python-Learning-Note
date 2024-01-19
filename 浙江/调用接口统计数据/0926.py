# coding: utf-8
from openpyxl.styles import Font,Border,Side,Alignment
import pandas as pd
from openpyxl import load_workbook


def write_excel_file(file_path):
    # 初始化worksheet对象
    wb = load_workbook(file_path)
    ws = wb.active

    # 定义样式
    fontstyle = Font(name="Microsoft YaHei UI", size=12,  bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),bottom=Side(style="thin"))
    align = Alignment(horizontal="center", vertical="center")
    fontstyle2 = Font(name="Microsoft YaHei UI", size=12)
    side = Side(
        style="medium",
        # 边框样式，可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
        color="ff66dd",  # 边框颜色，16进制rgb表示
    )
    for i in ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T'):
        ws.column_dimensions[i].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['Q'].width = 30
    row_nu=range(ws.max_row)
    # 设置单元格样式
    for i in range(ws.max_column):
        ws.cell(row=1, column=i + 1).font = fontstyle
        ws.cell(row=1, column=i + 1).border = border
        ws.cell(row=1, column=i + 1).alignment = align
        for j in row_nu:
            ws.row_dimensions[j+1].height = 20  # 设置第2行高度为30
            ws.cell(row=j + 2, column=i+1).font = fontstyle2
            ws.cell(row=j+2, column=2).alignment = align
            ws.cell(row=j+1, column=i+1).border = border
    ws.delete_cols(1)
    #保存到Excel文件
    wb.save(file_path)

def 格式处理(data1):
    data1['字节数'] = data1['身份证号'].str.len()
    data1['字节数2'] = data1['行政处罚决定书文号'].str.len()
    data1 = data1.drop_duplicates(subset=['record_code'])
    data1 = data1.drop_duplicates(subset=['行政处罚决定书文号'])
    data1 = data1[(data1['罚款金额'] >0)]
    data1 = data1[(data1['字节数'] ==18)]
    data1['处罚决定日期'] = pd.to_datetime(data1['处罚决定日期'],format='%Y-%m-%d')
    本季度前3次以上车牌= data1[(data1['处罚决定日期'] <= '2023-07-01 00:00:00')].groupby(["车牌号"])[
        'record_code'].count().reset_index(name='本季度前3次以上车牌')
    本季度前3次以上车牌=本季度前3次以上车牌[(本季度前3次以上车牌['本季度前3次以上车牌'] >=3)]
    本季度前3次以上车牌=本季度前3次以上车牌['车牌号']

    全部3次以上车牌= data1[(data1['处罚决定日期'] >= '2023-01-01 00:00:00')].groupby(["车牌号"])[
        'record_code'].count().reset_index(name='全部3次以上车牌')
    全部3次以上车牌=全部3次以上车牌[(全部3次以上车牌['全部3次以上车牌'] >=3)]
    全部3次以上车牌=全部3次以上车牌['车牌号']
    新增车牌 = data1[data1.loc[:, '车牌号'].isin(全部3次以上车牌)]

    新增车牌 = 新增车牌[~新增车牌.loc[:, '车牌号'].isin(本季度前3次以上车牌)]

    新增车牌 = 新增车牌.sort_values(by=['车牌号','处罚决定日期'],
                                    ascending=True).reset_index(drop=True)

    return 新增车牌



if __name__ == "__main__":
    data1 = pd.read_excel(r'C:\Users\stayhungary\Desktop\部级黑名单交警现场.xlsx')
    交警现场=格式处理(data1)
    data1 = pd.read_excel(r'C:\Users\stayhungary\Desktop\部级黑名单.xlsx')
    交通现场=格式处理(data1)
    部级黑名单车辆=pd.concat([交通现场,交警现场])
    全部3次以上人员= 部级黑名单车辆.groupby(['身份证号',"驾驶员姓名"])[
        '身份证号'].count().reset_index(name='全部3次以上人员')
    全部3次以上人员=全部3次以上人员[(全部3次以上人员['全部3次以上人员'] >=3)]
    全部3次以上人员=全部3次以上人员['身份证号']
    部级黑名单人员 = 部级黑名单车辆[部级黑名单车辆.loc[:, '身份证号'].isin(全部3次以上人员)]
    部级黑名单人员 = 部级黑名单人员.sort_values(by=['身份证号','处罚决定日期'],
                                    ascending=True).reset_index(drop=True)

    部级黑名单车辆['处罚决定日期'] = 部级黑名单车辆['处罚决定日期'].map(lambda x: x.strftime('%Y-%m-%d'))
    部级黑名单人员['处罚决定日期'] = 部级黑名单人员['处罚决定日期'].map(lambda x: x.strftime('%Y-%m-%d'))
    # 统计每列的不同值出现的次数

    部级黑名单车辆 = pd.DataFrame(部级黑名单车辆, columns=['序号', '车牌号','道路运输证号','违章次数','驾驶员姓名','身份证号','违法时间','行政处罚决定书文号','处罚决定日期','罚款金额','执法机构名称'])
    部级黑名单人员 = pd.DataFrame(部级黑名单人员, columns=['序号', '驾驶员姓名','身份证号','违章次数','车牌号','道路运输证号','违法时间','行政处罚决定书文号','处罚决定日期','罚款金额','执法机构名称'])
    with pd.ExcelWriter(r'C:\Users\stayhungary\Desktop\部级黑名单全部v4.0.xlsx') as writer1:
        部级黑名单车辆.to_excel(writer1, sheet_name='车辆', index=True)
        部级黑名单人员.to_excel(writer1, sheet_name='人员', index=True)

